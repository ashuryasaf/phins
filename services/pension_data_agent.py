"""
Pension Data Agent Service - Enhanced Mislaka Support
=====================================================
Processes Israeli pension and insurance XML data files according to
the Mislaka (מסלקה) interface standards based on official XSD schemas.

Based on ChatGPT analysis of:
- mivneachid_holdings_kupotgemel_xsd_schema_009.xsd
- mivneachid_holdings_karnotpensiavatikot_xsd_schema_009.xsd
- mivneachid_holdings_karnotpensiahadashot_xsd_schema_009.xsd
- mivneachid_holdings_hevrotbituah_xsd_schema_009.xsd
- mivneachid_mimshak_pitzuim_xsd_schema_9300-9306_005.XSD
- Holdings Interface V9.7.7
- Severance Interface V5.9.38
- Event Interface V7.6.30
- Transference Interface V3.7.2

Supported Interfaces:
- Holdings Interface (v9.7.7):
  * kupotgemel - קופות גמל (Provident Funds)
  * karnotpensiavatikot - קרנות פנסיה ותיקות (Old Pension Funds)  
  * karnotpensiahadashot - קרנות פנסיה חדשות (New Pension Funds)
  * hevrotbituah - חברות ביטוח (Insurance Companies)

- Severance Interface (v5.9.38):
  * Interface codes 9300, 9301, 9302, 9303, 9305, 9306

- Event Interface (v7.6.30)
- Transference Interface (v3.7.2)

Features:
- Full XSD schema field mappings from official Mislaka specs
- Automatic interface type and product type detection via SUG-MIMSHAK
- Comprehensive Hebrew field extraction with proper encoding
- Data enrichment with financial metrics
- Professional report generation matching Mislaka standards (like PDF format)
- AI-powered recommendations
- ClientProfile aggregation for multiple XML files in ZIP
- Section 14 (סעיף 14) analysis and rights explanation

Data Source: swiftness.co.il / Mislaka clearinghouse
Author: PHINS Platform
"""

import os
import json
import io
import re
import zipfile
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Any, Optional, Tuple, Set
from dataclasses import dataclass, asdict, field

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import lxml for better XML handling
try:
    from lxml import etree
    LXML_AVAILABLE = True
except ImportError:
    import xml.etree.ElementTree as etree
    LXML_AVAILABLE = False


# ============================================================================
# MISLAKA SCHEMA MAPPINGS - Based on official XSD schemas
# ============================================================================

class MislakaSchemaMapping:
    """
    Field mappings from Mislaka XSD schemas based on ChatGPT analysis.
    Based on:
    - mivneachid_holdings_*.xsd (v9.7.7)
    - mivneachid_mimshak_pitzuim_*.XSD (v5.9.38)
    - Holdings Interface V9.7.7 Excel specs
    - Severance Interface V5.9.38 Excel specs
    """
    
    # Interface type codes (SUG-MIMSHAK)
    INTERFACE_CODES = {
        # Holdings interfaces
        1: {'name': 'Holdings', 'he': 'אחזקות', 'schema': 'holdings_v9'},
        2: {'name': 'PreAdvice', 'he': 'הודעה מקדימה', 'schema': 'holdings_v9'},
        3: {'name': 'HoldingsPreAdvice', 'he': 'אחזקות + הודעה מקדימה', 'schema': 'holdings_v9'},
        
        # Severance interfaces (pitzuim)
        17: {'name': 'Severance', 'he': 'פיצויים', 'schema': 'pitzuim_v5'},
        9300: {'name': 'SeveranceRequest', 'he': 'בקשה לנתוני פיצויים', 'schema': 'pitzuim_9300'},
        9301: {'name': 'SeveranceResponse', 'he': 'תשובה לבקשת פיצויים', 'schema': 'pitzuim_9301'},
        9302: {'name': 'SeveranceQuery', 'he': 'שאילתת פיצויים', 'schema': 'pitzuim_9302'},
        9303: {'name': 'SeveranceData', 'he': 'נתוני פיצויים', 'schema': 'pitzuim_9303'},
        9305: {'name': 'SeveranceUpdate', 'he': 'עדכון פיצויים', 'schema': 'pitzuim_9305'},
        9306: {'name': 'SeveranceConfirm', 'he': 'אישור פיצויים', 'schema': 'pitzuim_9306'},
        
        # Event interface
        6: {'name': 'Events', 'he': 'אירועים', 'schema': 'events_v7'},
        21: {'name': 'Events', 'he': 'אירועים', 'schema': 'events_v7'},
        
        # Transference interface
        22: {'name': 'Transference', 'he': 'העברה', 'schema': 'transference_v3'},
        33: {'name': 'Transference', 'he': 'העברה', 'schema': 'transference_v3'},
    }
    
    # Product type codes (SUG-MUTZAR / KOD-SUG-MUTZAR)
    PRODUCT_TYPE_CODES = {
        # Pension funds - קרנות פנסיה
        '1': {'name': 'pension_fund_new', 'he': 'קרן פנסיה חדשה', 'en': 'New Pension Fund'},
        '2': {'name': 'pension_fund_old', 'he': 'קרן פנסיה ותיקה', 'en': 'Old Pension Fund'},
        '3': {'name': 'pension_fund_comprehensive', 'he': 'קרן פנסיה מקיפה', 'en': 'Comprehensive Pension'},
        
        # Provident funds (Gemel) - קופות גמל
        '4': {'name': 'provident_fund', 'he': 'קופת גמל', 'en': 'Provident Fund'},
        '5': {'name': 'central_severance_fund', 'he': 'קופה מרכזית לפיצויים', 'en': 'Central Severance Fund'},
        '6': {'name': 'education_fund', 'he': 'קרן השתלמות', 'en': 'Education Fund'},
        
        # Insurance - ביטוח
        '7': {'name': 'managers_insurance', 'he': 'ביטוח מנהלים', 'en': 'Managers Insurance'},
        '8': {'name': 'life_insurance', 'he': 'ביטוח חיים', 'en': 'Life Insurance'},
        '9': {'name': 'pension_insurance', 'he': 'ביטוח פנסיוני', 'en': 'Pension Insurance'},
        
        # Others
        '10': {'name': 'savings_policy', 'he': 'פוליסת חיסכון', 'en': 'Savings Policy'},
        '11': {'name': 'risk_insurance', 'he': 'ביטוח ריסק', 'en': 'Risk Insurance'},
        '12': {'name': 'disability_insurance', 'he': 'ביטוח אובדן כושר עבודה', 'en': 'Disability Insurance'},
    }
    
    # Status codes (STATUS-POLISA-O-CHESHBON)
    STATUS_CODES = {
        '1': {'name': 'active', 'he': 'פעיל', 'en': 'Active'},
        '2': {'name': 'frozen', 'he': 'מוקפא', 'en': 'Frozen'},
        '3': {'name': 'closed', 'he': 'סגור', 'en': 'Closed'},
        '4': {'name': 'paid_up', 'he': 'משולם', 'en': 'Paid Up'},
        '5': {'name': 'transferred', 'he': 'הועבר', 'en': 'Transferred'},
        '6': {'name': 'pending', 'he': 'בהמתנה', 'en': 'Pending'},
    }
    
    # Environment codes (KOD-SVIVAT-AVODA)
    ENVIRONMENT_CODES = {
        '1': 'Test',
        '2': 'Production',
    }
    
    # Sender/Recipient type codes (KOD-SHOLECH / KOD-NIMAAN)
    ENTITY_TYPE_CODES = {
        '1': {'he': 'יצרן', 'en': 'Provider/Institution'},
        '2': {'he': 'מסלקה', 'en': 'Clearinghouse'},
        '3': {'he': 'מפיץ/סוכן', 'en': 'Distributor/Agent'},
        '4': {'he': 'עמית/חוסך', 'en': 'Saver/Client'},
        '5': {'he': 'מעסיק', 'en': 'Employer'},
        '6': {'he': 'לשכת שירות', 'en': 'Service Bureau'},
    }
    
    # ID type codes (SUG-MEZAHE-SHOLECH / SUG-ZIHUI-LAKOACH)
    ID_TYPE_CODES = {
        '1': {'he': 'ח.פ (חברה)', 'en': 'Company ID'},
        '2': {'he': 'ח.צ (שותפות)', 'en': 'Partnership ID'},
        '3': {'he': 'תעודת זהות', 'en': 'ID Card'},
        '4': {'he': 'דרכון', 'en': 'Passport'},
        '5': {'he': 'רישיון עסק', 'en': 'Business License'},
        '7': {'he': 'עמותה', 'en': 'Non-profit'},
        '8': {'he': 'אגודה שיתופית', 'en': 'Cooperative'},
        '9': {'he': 'חברה ממשלתית', 'en': 'Government Company'},
    }
    
    # Header field mappings (KoteretKovetz)
    HEADER_FIELDS = {
        'SUG-MIMSHAK': 'interface_code',
        'SugMimshak': 'interface_code',
        'MISPAR-GIRSAT-XML': 'schema_version',
        'MisparGirsatXml': 'schema_version',
        'TAARICH-BITZUA': 'created_at',
        'TaarichBitzua': 'created_at',
        'TAARICH-HAFAKAT-HADOCH': 'report_date',
        'TaarichHafakatHadoch': 'report_date',
        'MISPAR-HAKOVETZ': 'file_id',
        'MisparHakovetz': 'file_id',
        'KOD-SVIVAT-AVODA': 'environment',
        'KodSvivatAvoda': 'environment',
        'KOD-SHOLEACH': 'sender_type',
        'KodSholeach': 'sender_type',
        'SUG-MEZAHE-SHOLECH': 'sender_id_type',
        'SugMezaheSholech': 'sender_id_type',
        'MISPAR-ZIHUI-SHOLECH': 'sender_id',
        'MisparZihuiSholech': 'sender_id',
        'SHEM-SHOLEACH': 'sender_name',
        'ShemSholeach': 'sender_name',
        'KOD-NIMAAN': 'recipient_type',
        'KodNimaan': 'recipient_type',
        'SHEM-MEKABEL': 'receiver_name',
        'ShemMekabel': 'receiver_name',
    }
    
    # Client field mappings (YeshutLakoach)
    CLIENT_FIELDS = {
        'MISPAR-ZIHUI-LAKOACH': 'id_number',
        'MisparZihuiLakoach': 'id_number',
        'MISPARZEHUT': 'id_number',
        'MisparZehut': 'id_number',
        'MISPAR-ZIHUY-MITPATZEACH': 'id_number',
        'MisparZihuiMitpatcheach': 'id_number',
        'SUG-ZIHUI-LAKOACH': 'id_type',
        'SugZihuiLakoach': 'id_type',
        'SugZihuiMitpatcheach': 'id_type',
        'SHEM-PRATI': 'first_name',
        'ShemPrati': 'first_name',
        'SHEM-MISHPACHA': 'last_name',
        'ShemMishpacha': 'last_name',
        'SHEM-LAKOACH': 'full_name',
        'ShemLakoach': 'full_name',
        'TAARICH-LEYDA': 'birth_date',
        'TaarichLeyda': 'birth_date',
        'MIN': 'gender',
        'KTOVET': 'address',
        'Ktovet': 'address',
        'YISHUV': 'city',
        'Yishuv': 'city',
        'TELEFON': 'phone',
        'Telefon': 'phone',
        'EMAIL': 'email',
        'Email': 'email',
    }
    
    # Provider field mappings (YeshutYatzran)
    PROVIDER_FIELDS = {
        'KOD-YATZRAN': 'code',
        'KodYatzran': 'code',
        'KOD-MEZAHE-YATZRAN': 'code',
        'KodMezaheYatzran': 'code',
        'SHEM-YATZRAN': 'name',
        'ShemYatzran': 'name',
        'SUG-YATZRAN': 'provider_type',
        'SugYatzran': 'provider_type',
    }
    
    # Product field mappings (Mutzar)
    PRODUCT_FIELDS = {
        'KOD-MUTZAR': 'code',
        'KodMutzar': 'code',
        'SHEM-MUTZAR': 'name',
        'ShemMutzar': 'name',
        'SUG-MUTZAR': 'product_type',
        'SugMutzar': 'product_type',
        'KOD-SUG-MUTZAR': 'product_type_code',
        'KodSugMutzar': 'product_type_code',
        'SUG-KUPA': 'sub_type',
        'SugKupa': 'sub_type',
        'STATUS-MUTZAR': 'status',
        'StatusMutzar': 'status',
        'TAARICH-TCHILAT-MUTZAR': 'start_date',
        'TaarichTchilatMutzar': 'start_date',
    }
    
    # Account field mappings (HeshbonOPolisa / PirteiHeshbon)
    ACCOUNT_FIELDS = {
        # Policy number
        'MISPAR-POLISA-O-HESHBON': 'policy_number',
        'MisparPolisaOHeshbon': 'policy_number',
        'MISPAR-POLISA': 'policy_number',
        'MisparPolisa': 'policy_number',
        'MISPAR-HESHBON': 'policy_number',
        'MisparHeshbon': 'policy_number',
        'MISPAR-CHESHBON': 'policy_number',
        'MisparCheshbon': 'policy_number',
        
        # Status
        'STATUS-POLISA-O-CHESHBON': 'status_code',
        'StatusPolisaOCheshbon': 'status_code',
        'STATUS-HESHBON': 'status_code',
        'StatusHeshbon': 'status_code',
        
        # Start date
        'TAARICH-TCHILAT-HESHBON': 'start_date',
        'TaarichTchilatHeshbon': 'start_date',
        'TAARICH-TCHILAT-BITUACH': 'start_date',
        'TaarichTchilatBituach': 'start_date',
        
        # Balances
        'SALDO': 'total_balance',
        'Saldo': 'total_balance',
        'YITRA-KOLELET': 'total_balance',
        'YitraKolelet': 'total_balance',
        'YITROT': 'total_balance',
        'Yitrot': 'total_balance',
        'SCHUM-HATZBARA': 'total_balance',
        'SchumHatzbara': 'total_balance',
        'YITRA-CHISACHON': 'savings_balance',
        'YitraChisachon': 'savings_balance',
        'YITRA-PITZUIM': 'severance_balance',
        'YitraPitzuim': 'severance_balance',
        'KFIFA-PITZUIM': 'severance_balance',
        'KfifaPitzuim': 'severance_balance',
        'PITZUEY-MAASIK': 'employer_severance',
        'PitzueyMaasik': 'employer_severance',
        'YITRA-TAGMULIM': 'compensation_balance',
        'YitraTagmulim': 'compensation_balance',
        
        # Investment track
        'MASLUL-HASHKAA': 'investment_track',
        'MaslulHashkaa': 'investment_track',
        'KOD-MASLUL': 'investment_track_code',
        'KodMaslul': 'investment_track_code',
        
        # Management fees
        'DMEY-NIHUL-CHISACHON': 'management_fee_savings',
        'DmeyNihulChisachon': 'management_fee_savings',
        'DMEY-NIHUL-HAFKADOT': 'management_fee_deposits',
        'DmeyNihulHafkadot': 'management_fee_deposits',
        
        # Employer
        'KOD-MAASIK': 'employer_id',
        'KodMaasik': 'employer_id',
        'SHEM-MAASIK': 'employer_name',
        'ShemMaasik': 'employer_name',
        'MPR-MAASIK-BE-YATZRAN': 'employer_internal_id',
        'MprMaasikBeYatzran': 'employer_internal_id',
        
        # Coverage (for insurance)
        'SACH-KISUY': 'coverage_amount',
        'SachKisuy': 'coverage_amount',
        'KITZBA-CHODSHIT': 'monthly_pension',
        'KitzbaChodshit': 'monthly_pension',
        'KISUY-MAVET': 'death_coverage',
        'KisuyMavet': 'death_coverage',
        'KISUY-NECHUT': 'disability_coverage',
        'KisuyNechut': 'disability_coverage',
        
        # Section 14
        'SEIF-14': 'section14',
        'Seif14': 'section14',
        'ARTICLE14': 'section14',
        'article14': 'section14',
        'SI14': 'section14',
        'SACHIF-14': 'section14',
        'TAARICH-SEIF-14': 'section14_date',
        'TaarichSeif14': 'section14_date',
    }
    
    # Contribution field mappings (NetuneiHafrasha / PirteiHafrasha)
    CONTRIBUTION_FIELDS = {
        'CHODESH-DIO': 'period',
        'ChodeshDio': 'period',
        'CHODESH': 'period',
        'Chodesh': 'period',
        'TKUFA': 'period',
        'Tkufa': 'period',
        
        'MISPAR-ZIHUI-LAKOACH': 'employee_id',
        'KOD-MAASIK': 'employer_id',
        'SHEM-MAASIK': 'employer_name',
        
        'HAFRASHA-OVED': 'employee_amount',
        'HafrashaOved': 'employee_amount',
        'HAFRASHA-MAASIK': 'employer_amount',
        'HafrashaMaasik': 'employer_amount',
        'HAFRASHA-PITZUIM': 'severance_amount',
        'HafrashaPitzuim': 'severance_amount',
        'SACH-HAFRASHA': 'total_amount',
        'SachHafrasha': 'total_amount',
        
        'SACHAR-KOVEA': 'salary_base',
        'SacharKovea': 'salary_base',
        
        'STATUS-HAFRASHA': 'status',
        'StatusHafrasha': 'status',
        'TAARICH-KLITA': 'received_date',
        'TaarichKlita': 'received_date',
    }
    
    # Severance field mappings (NetuneiPitzuim)
    SEVERANCE_FIELDS = {
        'MISPAR-ZIHUI-LAKOACH': 'employee_id',
        'MISPAR-POLISA': 'policy_number',
        'KOD-MAASIK': 'employer_id',
        'SHEM-MAASIK': 'employer_name',
        
        'KSF-PITZUIM-TZVUR': 'total_severance',
        'KsfPitzuimTzvur': 'total_severance',
        'SACH-PITZUIM': 'total_severance',
        'SachPitzuim': 'total_severance',
        'ERECH-PIDYON-PITZUIM-MAASEK-NOCHECHI': 'total_severance',
        'PITZUIM-LMSHICHA': 'available_severance',
        'PitzuimLmshicha': 'available_severance',
        'PITZUIM-SEIF14': 'section14_amount',
        'PitzuimSeif14': 'section14_amount',
        
        'SEIF-14': 'section14',
        'article14': 'section14',
        'ARTICLE14': 'section14',
        'ACHUZ-SEIF-14': 'section14_percentage',
        'AchuzSeif14': 'section14_percentage',
        
        'TAARICH-TCHILAT-AVODA': 'employment_start',
        'TaarichTchilatAvoda': 'employment_start',
        'TAARICH-SIUM-AVODA': 'employment_end',
        'TaarichSiumAvoda': 'employment_end',
    }
    
    # Known insurance company codes
    INSURANCE_COMPANIES = {
        '1': 'מגדל',
        '2': 'הראל',
        '3': 'כלל',
        '4': 'פניקס',
        '5': 'הפניקס',
        '6': 'מנורה מבטחים',
        '7': 'איילון',
        '8': 'ביטוח ישיר',
        '9': 'שירביט',
        '10': 'הכשרה',
        '11': 'ליברה',
        '12': 'אקסא',
    }
    
    # Known pension fund codes
    PENSION_FUNDS = {
        '512': 'מיטב דש',
        '513': 'אלטשולר שחם',
        '514': 'מור',
        '515': 'הלמן אלדובי',
        '516': 'אנליסט',
        '517': 'פסגות',
        '518': 'מנורה מבטחים פנסיה',
        '519': 'הראל פנסיה',
        '520': 'מגדל מקפת',
        '521': 'כלל פנסיה',
    }


# ============================================================================
# CLIENT PROFILE - Aggregated data from multiple XML files
# ============================================================================

@dataclass
class ClientProfile:
    """
    Aggregated client profile containing all data from multiple Mislaka sources.
    Used for ZIP files containing multiple XML files.
    """
    # Basic client info
    client_name: str = ""
    client_id: str = ""
    id_type: str = ""
    birth_date: str = ""
    gender: str = ""
    address: str = ""
    city: str = ""
    phone: str = ""
    email: str = ""
    
    # Accounts and holdings
    accounts: List[Dict] = field(default_factory=list)
    
    # Severance details
    severance_balance: float = 0.0
    section14: bool = False
    section14_date: str = ""
    
    # Contributions/events
    contributions: List[Dict] = field(default_factory=list)
    events: List[Dict] = field(default_factory=list)
    
    # Employers
    employers: Set[str] = field(default_factory=set)
    
    # Providers
    providers: Set[str] = field(default_factory=set)
    
    # Derived metrics
    total_balance: float = 0.0
    total_savings: float = 0.0
    total_severance: float = 0.0
    total_coverage: float = 0.0
    
    # Metadata
    last_update: str = ""
    file_count: int = 0
    interface_types: Set[str] = field(default_factory=set)
    
    # Anomalies/alerts
    anomalies: List[str] = field(default_factory=list)
    
    def merge_data(self, data: Dict[str, Any]):
        """Merge data from a single parsed file into this profile."""
        # Merge client info
        client = data.get('client', {})
        if client:
            if isinstance(client, list) and client:
                client = client[0]
            
            if client.get('id_number'):
                if self.client_id and self.client_id != client['id_number']:
                    self.anomalies.append(f"ID mismatch: {self.client_id} vs {client['id_number']}")
                else:
                    self.client_id = client['id_number']
            
            if client.get('full_name'):
                self.client_name = client['full_name']
            elif client.get('first_name') or client.get('last_name'):
                self.client_name = f"{client.get('first_name', '')} {client.get('last_name', '')}".strip()
            
            if client.get('id_type'):
                self.id_type = client['id_type']
            if client.get('birth_date'):
                self.birth_date = client['birth_date']
            if client.get('phone'):
                self.phone = client['phone']
            if client.get('email'):
                self.email = client['email']
        
        # Merge accounts
        for acct in data.get('accounts', []):
            self.accounts.append(acct)
            if acct.get('provider'):
                self.providers.add(acct['provider'])
            if acct.get('employer_name'):
                self.employers.add(acct['employer_name'])
        
        # Merge contributions
        for contrib in data.get('contributions', []):
            self.contributions.append(contrib)
            if contrib.get('employer_name'):
                self.employers.add(contrib['employer_name'])
        
        # Merge severance
        for sev in data.get('severance', []):
            if sev.get('total_severance'):
                self.severance_balance += float(sev['total_severance'] or 0)
            if sev.get('section14'):
                self.section14 = True
            if sev.get('employer_name'):
                self.employers.add(sev['employer_name'])
        
        # Merge providers from providers list
        for prov in data.get('providers', []):
            if prov.get('name'):
                self.providers.add(prov['name'])
        
        # Update metadata
        header = data.get('header', {})
        if header.get('report_date') or header.get('created_at'):
            file_date = header.get('report_date') or header.get('created_at')
            if not self.last_update or file_date > self.last_update:
                self.last_update = file_date
        
        if data.get('interface_type'):
            self.interface_types.add(data['interface_type'])
        
        self.file_count += 1
    
    def finalize(self):
        """Compute derived totals after all merges."""
        self.total_balance = 0.0
        self.total_savings = 0.0
        self.total_severance = 0.0
        self.total_coverage = 0.0
        
        for acct in self.accounts:
            self.total_balance += float(acct.get('total_balance', 0) or 0)
            self.total_savings += float(acct.get('savings_balance', 0) or 0)
            self.total_severance += float(acct.get('severance_balance', 0) or 0)
            self.total_coverage += float(acct.get('coverage_amount', 0) or 0)
            
            # Check for Section 14
            if acct.get('section14'):
                self.section14 = True
        
        # Add external severance balance
        self.total_severance += self.severance_balance
        
        # Check for anomalies
        if self.total_balance < 0:
            self.anomalies.append("Total balance is negative")
        if not self.accounts:
            self.anomalies.append("No accounts found")
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            'client': {
                'full_name': self.client_name,
                'id_number': self.client_id,
                'id_type': self.id_type,
                'birth_date': self.birth_date,
                'phone': self.phone,
                'email': self.email,
            },
            'accounts': self.accounts,
            'contributions': self.contributions,
            'severance': [{
                'total_severance': self.total_severance,
                'section14': self.section14,
            }],
            'providers': list(self.providers),
            'employers': list(self.employers),
            'totals': {
                'total_balance': self.total_balance,
                'total_balance_formatted': f"₪{self.total_balance:,.2f}",
                'total_savings': self.total_savings,
                'total_savings_formatted': f"₪{self.total_savings:,.2f}",
                'total_severance': self.total_severance,
                'total_severance_formatted': f"₪{self.total_severance:,.2f}",
                'total_coverage': self.total_coverage,
                'account_count': len(self.accounts),
                'provider_count': len(self.providers),
                'providers': list(self.providers),
                'section14_coverage': self.section14,
            },
            'header': {
                'report_date': self.last_update,
                'file_count': self.file_count,
                'interface_types': list(self.interface_types),
            },
            'anomalies': self.anomalies,
        }


# ============================================================================
# PENSION DATA AGENT - Main processing class
# ============================================================================

class PensionDataAgent:
    """
    Enhanced Pension Data Agent for processing Mislaka (מסלקה) XML/ZIP data.
    Based on ChatGPT analysis of official XSD schemas and interface specs.
    """
    
    def __init__(self, schema_dir: str = None):
        """Initialize the agent with optional schema directory."""
        self.schema_dir = schema_dir or os.path.join(os.path.dirname(__file__), 'schemas')
        self.schema_mapping = MislakaSchemaMapping()
        self.schema_cache = {}
    
    def process_xml_content(self, xml_content: bytes) -> Dict[str, Any]:
        """
        Process a single XML content and generate report.
        
        Args:
            xml_content: Raw XML bytes from Mislaka file
            
        Returns:
            Dictionary with parsed data and generated report
        """
        # Parse XML
        data = self._parse_mislaka_xml(xml_content)
        
        # Enrich with derived metrics
        data = self._enrich_data(data)
        
        # Generate professional report
        report = self._generate_professional_report(data)
        
        return {
            'data': data,
            'report': report,
            'language': 'hebrew',
            'interface_type': data.get('interface_type', 'Unknown'),
            'schema_version': data.get('header', {}).get('schema_version', 'Unknown'),
        }
    
    def process_zip_content(self, zip_content: bytes) -> Dict[str, Any]:
        """
        Process a ZIP file containing Mislaka XML and Excel files.
        
        Args:
            zip_content: Raw ZIP file bytes
            
        Returns:
            Dictionary with aggregated data and generated report
        """
        profile = ClientProfile()
        
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zf:
                for filename in zf.namelist():
                    # Skip hidden files
                    if filename.startswith('__') or filename.startswith('.'):
                        continue
                    
                    name_lower = filename.lower()
                    ext = name_lower.split('.')[-1] if '.' in name_lower else ''
                    
                    logger.info(f"Processing file: {filename}")
                    
                    try:
                        file_bytes = zf.read(filename)
                        
                        if ext == 'xml':
                            # Process XML file
                            file_data = self._parse_mislaka_xml(file_bytes)
                            profile.merge_data(file_data)
                        elif ext in ['xls', 'xlsx']:
                            # Process Excel file
                            file_data = self._parse_mislaka_excel(file_bytes, filename, ext)
                            if file_data:
                                profile.merge_data(file_data)
                        elif ext == 'csv':
                            # Process CSV file
                            file_data = self._parse_mislaka_csv(file_bytes, filename)
                            if file_data:
                                profile.merge_data(file_data)
                    except Exception as e:
                        logger.error(f"Error processing {filename}: {e}")
                        profile.anomalies.append(f"Failed to parse {filename}: {str(e)}")
        except zipfile.BadZipFile:
            raise ValueError("Invalid ZIP file format")
        
        # Finalize profile calculations
        profile.finalize()
        
        # Convert to standard data format
        data = profile.to_dict()
        
        # Enrich with health score
        data = self._enrich_data(data)
        
        # Generate report
        report = self._generate_professional_report(data)
        
        return {
            'data': data,
            'report': report,
            'language': 'hebrew',
            'interface_type': ', '.join(profile.interface_types) if profile.interface_types else 'Unknown',
            'file_count': profile.file_count,
        }
    
    def _parse_mislaka_xml(self, xml_content: bytes) -> Dict[str, Any]:
        """Parse Mislaka XML into structured data with proper encoding handling."""
        # Try to parse with automatic encoding detection
        try:
            if LXML_AVAILABLE:
                parser = etree.XMLParser(recover=True, encoding='utf-8')
                root = etree.fromstring(xml_content, parser)
            else:
                # Try UTF-8 first
                try:
                    root = etree.fromstring(xml_content.decode('utf-8', errors='replace'))
                except:
                    # Fall back to Windows-1255 (Hebrew)
                    root = etree.fromstring(xml_content.decode('windows-1255', errors='replace'))
        except Exception as e:
            # Last resort: try Windows-1255
            try:
                text = xml_content.decode('windows-1255', errors='replace')
                root = etree.fromstring(text)
            except Exception as e2:
                raise ValueError(f"Failed to parse XML: {e2}")
        
        data = {
            'header': {},
            'client': {},
            'providers': [],
            'accounts': [],
            'contributions': [],
            'severance': [],
            'employers': [],
            'raw_elements': {}
        }
        
        # Parse header
        data['header'] = self._parse_header(root)
        
        # Determine interface type
        interface_code = data['header'].get('interface_code', 1)
        data['interface_code'] = interface_code
        interface_info = self.schema_mapping.INTERFACE_CODES.get(
            interface_code, 
            {'name': f'Type{interface_code}', 'he': f'סוג {interface_code}'}
        )
        data['interface_type'] = interface_info['name']
        data['interface_type_he'] = interface_info['he']
        
        # Parse client
        data['client'] = self._parse_client(root)
        
        # Parse providers and accounts
        providers, accounts = self._parse_providers_and_accounts(root)
        data['providers'] = providers
        data['accounts'] = accounts
        
        # Parse contributions
        data['contributions'] = self._parse_contributions(root)
        
        # Parse severance data
        data['severance'] = self._parse_severance(root, interface_code)
        
        # Parse employers
        data['employers'] = self._parse_employers(root, accounts)
        
        # Extract all raw text elements for additional analysis
        data['raw_elements'] = self._extract_all_elements(root)
        
        return data
    
    def _parse_mislaka_excel(self, content: bytes, filename: str, ext: str) -> Optional[Dict[str, Any]]:
        """
        Parse Mislaka Excel files (.xls and .xlsx) into structured data.
        
        Args:
            content: Raw Excel file bytes
            filename: Original filename
            ext: File extension ('xls' or 'xlsx')
            
        Returns:
            Dictionary with parsed data or None if parsing fails
        """
        try:
            rows = []
            columns = []
            
            if ext == 'xlsx':
                # Use openpyxl for .xlsx files
                try:
                    import openpyxl
                    from openpyxl import load_workbook
                    
                    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
                    
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        sheet_rows = list(sheet.iter_rows(values_only=True))
                        
                        if not sheet_rows:
                            continue
                        
                        # First row as headers
                        header_row = sheet_rows[0] if sheet_rows else []
                        sheet_columns = [str(h).strip() if h else f'עמודה_{i}' for i, h in enumerate(header_row)]
                        
                        # Add columns that don't exist yet
                        for col in sheet_columns:
                            if col and col not in columns:
                                columns.append(col)
                        
                        # Process data rows
                        for row in sheet_rows[1:]:
                            if row and any(cell is not None for cell in row):
                                row_dict = {}
                                for i, cell in enumerate(row):
                                    if i < len(sheet_columns):
                                        col_name = sheet_columns[i]
                                        row_dict[col_name] = cell if cell is not None else ''
                                rows.append(row_dict)
                    
                    wb.close()
                    
                except ImportError:
                    logger.warning("openpyxl not available for xlsx parsing")
                    return None
                    
            elif ext == 'xls':
                # Use xlrd for older .xls files
                try:
                    import xlrd
                    
                    wb = xlrd.open_workbook(file_contents=content)
                    
                    for sheet_idx in range(wb.nsheets):
                        sheet = wb.sheet_by_index(sheet_idx)
                        
                        if sheet.nrows == 0:
                            continue
                        
                        # First row as headers
                        header_row = sheet.row_values(0) if sheet.nrows > 0 else []
                        sheet_columns = [str(h).strip() if h else f'עמודה_{i}' for i, h in enumerate(header_row)]
                        
                        # Add columns that don't exist yet
                        for col in sheet_columns:
                            if col and col not in columns:
                                columns.append(col)
                        
                        # Process data rows
                        for row_idx in range(1, sheet.nrows):
                            row = sheet.row_values(row_idx)
                            if row and any(cell for cell in row):
                                row_dict = {}
                                for i, cell in enumerate(row):
                                    if i < len(sheet_columns):
                                        col_name = sheet_columns[i]
                                        row_dict[col_name] = cell if cell else ''
                                rows.append(row_dict)
                    
                except ImportError:
                    logger.warning("xlrd not available for xls parsing")
                    return None
            
            if not rows and not columns:
                return None
            
            # Map Excel columns to pension data structure
            return self._map_excel_to_pension_data(columns, rows, filename)
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {filename}: {e}")
            return None
    
    def _parse_mislaka_csv(self, content: bytes, filename: str) -> Optional[Dict[str, Any]]:
        """
        Parse Mislaka CSV files into structured data.
        
        Args:
            content: Raw CSV file bytes
            filename: Original filename
            
        Returns:
            Dictionary with parsed data or None if parsing fails
        """
        try:
            # Try to decode with various encodings
            text_content = None
            for encoding in ['utf-8', 'windows-1255', 'iso-8859-8', 'cp1255']:
                try:
                    text_content = content.decode(encoding)
                    break
                except:
                    continue
            
            if not text_content:
                text_content = content.decode('utf-8', errors='replace')
            
            rows = []
            columns = []
            
            lines = text_content.strip().split('\n')
            if not lines:
                return None
            
            # Detect delimiter
            first_line = lines[0]
            delimiter = ','
            for delim in [',', '\t', ';', '|']:
                if delim in first_line:
                    delimiter = delim
                    break
            
            # Parse header
            columns = [col.strip().strip('"') for col in first_line.split(delimiter)]
            
            # Parse data rows
            for line in lines[1:]:
                if not line.strip():
                    continue
                values = [v.strip().strip('"') for v in line.split(delimiter)]
                if values and any(v for v in values):
                    row_dict = {}
                    for i, val in enumerate(values):
                        if i < len(columns):
                            row_dict[columns[i]] = val
                    rows.append(row_dict)
            
            if not rows:
                return None
            
            # Map CSV columns to pension data structure
            return self._map_excel_to_pension_data(columns, rows, filename)
            
        except Exception as e:
            logger.error(f"Error parsing CSV file {filename}: {e}")
            return None
    
    def _map_excel_to_pension_data(self, columns: List[str], rows: List[Dict], filename: str) -> Dict[str, Any]:
        """
        Map Excel/CSV column data to Mislaka pension data structure.
        
        Args:
            columns: List of column names
            rows: List of row dictionaries
            filename: Source filename
            
        Returns:
            Dictionary in standard Mislaka pension data format
        """
        # Hebrew column name mappings
        column_mappings = {
            # Client fields
            'שם': 'full_name',
            'שם מלא': 'full_name', 
            'שם פרטי': 'first_name',
            'שם משפחה': 'last_name',
            'תעודת זהות': 'id_number',
            'ת.ז': 'id_number',
            'ת"ז': 'id_number',
            'מספר זהות': 'id_number',
            'מספר ת.ז': 'id_number',
            'תאריך לידה': 'birth_date',
            'טלפון': 'phone',
            'נייד': 'mobile',
            'דוא"ל': 'email',
            'אימייל': 'email',
            'כתובת': 'address',
            
            # Provider/Product fields
            'יצרן': 'provider',
            'שם יצרן': 'provider',
            'חברה': 'provider',
            'שם חברה': 'provider',
            'גוף מוסדי': 'provider',
            'מוצר': 'product_name',
            'שם מוצר': 'product_name',
            'סוג מוצר': 'product_type',
            'סוג קופה': 'product_type',
            'סוג תוכנית': 'product_type',
            
            # Policy fields
            'מספר פוליסה': 'policy_number',
            'מס פוליסה': 'policy_number',
            "מס' פוליסה": 'policy_number',
            'מספר חשבון': 'policy_number',
            'מס חשבון': 'policy_number',
            
            # Balance fields
            'יתרה': 'total_balance',
            'יתרה כוללת': 'total_balance',
            'סך צבירה': 'total_balance',
            'צבירה': 'total_balance',
            'סה"כ צבירה': 'total_balance',
            'סכום צבירה': 'total_balance',
            'יתרת תגמולים': 'savings_balance',
            'תגמולים': 'savings_balance',
            'חיסכון': 'savings_balance',
            'יתרת פיצויים': 'severance_balance',
            'פיצויים': 'severance_balance',
            'סכום פיצויים': 'severance_balance',
            
            # Fee fields
            'דמי ניהול': 'management_fee',
            'דמי ניהול מצבירה': 'management_fee_savings',
            'ד"נ מצבירה': 'management_fee_savings',
            'דמי ניהול מהפקדות': 'management_fee_deposits',
            'ד"נ מהפקדות': 'management_fee_deposits',
            'עמלה': 'management_fee',
            
            # Status fields
            'סטטוס': 'status',
            'מצב': 'status',
            'סטטוס פוליסה': 'status',
            'מצב חשבון': 'status',
            
            # Section 14
            'סעיף 14': 'section14',
            'סעיף14': 'section14',
            
            # Employer
            'מעסיק': 'employer_name',
            'שם מעסיק': 'employer_name',
            
            # Insurance coverage
            'ביטוח חיים': 'death_coverage',
            'כיסוי מוות': 'death_coverage',
            'אובדן כושר': 'disability_coverage',
            'כיסוי נכות': 'disability_coverage',
            
            # Dates
            'תאריך תחילה': 'start_date',
            'תחילת ביטוח': 'start_date',
            'תאריך הצטרפות': 'start_date',
        }
        
        # Check if this looks like pension data
        pension_indicators = ['יצרן', 'פוליסה', 'צבירה', 'יתרה', 'תגמולים', 'פיצויים', 
                              'קופה', 'פנסיה', 'ביטוח', 'גמל', 'חיסכון', 'קרן']
        columns_lower = [str(c).lower() for c in columns]
        
        is_pension_data = any(
            any(indicator in col for indicator in pension_indicators) 
            for col in columns_lower
        )
        
        if not is_pension_data:
            logger.info(f"File {filename} does not appear to contain pension data")
            return None
        
        # Map columns to standardized names
        # Prioritize exact matches over substring matches to avoid confusion
        # (e.g., "צבירה" should not match "דמי ניהול מצבירה")
        mapped_columns = {}
        for col in columns:
            col_str = str(col).strip()
            
            # First try exact match
            if col_str in column_mappings:
                mapped_columns[col] = column_mappings[col_str]
                continue
            
            # Then try substring match, but prefer longer matches first
            # Sort potential matches by length (longest first)
            matches = []
            for hebrew_name, english_name in column_mappings.items():
                if hebrew_name in col_str:
                    matches.append((len(hebrew_name), hebrew_name, english_name))
            
            if matches:
                # Sort by length descending, take the longest match
                matches.sort(key=lambda x: x[0], reverse=True)
                mapped_columns[col] = matches[0][2]
        
        # Extract client and account data
        client_info = {}
        accounts = []
        employers = []
        
        for row in rows:
            account = {
                'source': 'excel',
                'source_file': filename
            }
            
            for original_col, value in row.items():
                if value is None or str(value).strip() == '':
                    continue
                    
                if original_col in mapped_columns:
                    mapped_name = mapped_columns[original_col]
                    value_str = str(value).strip()
                    
                    # Convert value based on field type
                    if mapped_name in ['total_balance', 'savings_balance', 'severance_balance', 
                                      'management_fee', 'management_fee_savings', 'management_fee_deposits',
                                      'death_coverage', 'disability_coverage']:
                        try:
                            # Clean numeric value
                            clean_val = value_str.replace(',', '').replace('₪', '').replace('ש"ח', '').strip()
                            account[mapped_name] = float(clean_val)
                        except:
                            account[mapped_name] = 0
                    elif mapped_name == 'section14':
                        account[mapped_name] = value_str.lower() in ['כן', 'yes', '1', 'true', 'v', '✓', 'y']
                    elif mapped_name in ['full_name', 'first_name', 'last_name', 'id_number', 
                                        'birth_date', 'phone', 'mobile', 'email', 'address']:
                        # Client info
                        if not client_info.get(mapped_name):
                            client_info[mapped_name] = value_str
                    elif mapped_name == 'employer_name':
                        if value_str and value_str not in employers:
                            employers.append(value_str)
                        account[mapped_name] = value_str
                    else:
                        account[mapped_name] = value_str
            
            # Only add if we have some account data
            if account.get('provider') or account.get('policy_number') or account.get('total_balance'):
                accounts.append(account)
        
        # Build full name if we have parts
        if client_info.get('first_name') or client_info.get('last_name'):
            parts = [client_info.get('first_name', ''), client_info.get('last_name', '')]
            client_info['full_name'] = ' '.join(p for p in parts if p)
        
        # Translate product types
        for account in accounts:
            if account.get('product_type'):
                pt = account['product_type']
                for code, info in self.schema_mapping.PRODUCT_TYPE_CODES.items():
                    if pt in [info['he'], info['en'], code]:
                        account['product_type_name'] = info['he']
                        account['product_type_en'] = info['en']
                        break
            
            # Set default status if not present
            if not account.get('status'):
                account['status'] = 'פעיל'
                account['status_en'] = 'Active'
        
        return {
            'header': {
                'source': 'Excel/CSV',
                'filename': filename,
                'interface_type': 'Excel Import',
                'interface_type_he': 'יבוא מאקסל',
            },
            'client': client_info,
            'providers': [],  # Will be derived from accounts
            'accounts': accounts,
            'contributions': [],
            'severance': [],
            'employers': [{'name': e} for e in employers],
            'interface_type': 'Excel',
            'interface_type_he': 'יבוא מאקסל',
        }
    
    def _parse_header(self, root) -> Dict[str, Any]:
        """Parse header (KoteretKovetz) from XML."""
        header = {}
        
        # Find header element
        header_elem = root.find('.//KoteretKovetz')
        if header_elem is None:
            header_elem = root.find('.//Header')
        if header_elem is None:
            header_elem = root
        
        # Extract fields using mapping
        for xml_tag, field_name in self.schema_mapping.HEADER_FIELDS.items():
            value = self._find_text(header_elem, xml_tag)
            if value:
                if field_name == 'interface_code':
                    try:
                        header[field_name] = int(value)
                    except:
                        header[field_name] = 1
                else:
                    header[field_name] = value
        
        # Add interface type name
        interface_code = header.get('interface_code', 1)
        interface_info = self.schema_mapping.INTERFACE_CODES.get(interface_code, {})
        header['interface_type'] = interface_info.get('name', f'Type{interface_code}')
        header['interface_type_he'] = interface_info.get('he', f'סוג {interface_code}')
        
        # Format dates
        if header.get('created_at') and len(header['created_at']) == 14:
            try:
                dt = datetime.strptime(header['created_at'], '%Y%m%d%H%M%S')
                header['created_at_formatted'] = dt.strftime('%d/%m/%Y %H:%M')
            except:
                pass
        
        return header
    
    def _parse_client(self, root) -> Dict[str, Any]:
        """Parse client (YeshutLakoach) from XML."""
        client = {}
        
        # Try multiple element names
        client_tags = ['YeshutLakoach', 'YeshutLakohach', 'Lakoach', 'Mevutach', 'Client', 'ClientDetails']
        client_elem = None
        
        for tag in client_tags:
            client_elem = root.find(f'.//{tag}')
            if client_elem is not None:
                break
        
        if client_elem is None:
            return client
        
        # Extract fields
        for xml_tag, field_name in self.schema_mapping.CLIENT_FIELDS.items():
            value = self._find_text(client_elem, xml_tag)
            if value:
                client[field_name] = value
        
        # Build full name if not present
        if not client.get('full_name') and (client.get('first_name') or client.get('last_name')):
            parts = [client.get('first_name', ''), client.get('last_name', '')]
            client['full_name'] = ' '.join(p for p in parts if p)
        
        # Translate ID type
        if client.get('id_type'):
            id_type_info = self.schema_mapping.ID_TYPE_CODES.get(client['id_type'], {})
            client['id_type_name'] = id_type_info.get('he', client['id_type'])
        
        return client
    
    def _parse_providers_and_accounts(self, root) -> Tuple[List[Dict], List[Dict]]:
        """Parse providers (YeshutYatzran) and accounts (HeshbonOPolisa)."""
        providers = []
        accounts = []
        
        # Find all providers
        for provider_elem in root.findall('.//YeshutYatzran'):
            provider = {}
            
            for xml_tag, field_name in self.schema_mapping.PROVIDER_FIELDS.items():
                value = self._find_text(provider_elem, xml_tag)
                if value:
                    provider[field_name] = value
            
            if provider:
                providers.append(provider)
            
            # Find products under this provider
            for product_elem in provider_elem.findall('.//Mutzar'):
                product_info = {}
                
                for xml_tag, field_name in self.schema_mapping.PRODUCT_FIELDS.items():
                    value = self._find_text(product_elem, xml_tag)
                    if value:
                        product_info[field_name] = value
                
                # Find accounts under this product
                account_tags = ['HeshbonOPolisa', 'PirteiHeshbon', 'Account', 'Policy', 'Plan', 'ReshimatKupa']
                for tag in account_tags:
                    for account_elem in product_elem.findall(f'.//{tag}'):
                        account = self._parse_account(account_elem, provider, product_info)
                        if account:
                            accounts.append(account)
        
        # Also find standalone accounts
        standalone_tags = ['HeshbonOPolisa', 'Account', 'Policy']
        for tag in standalone_tags:
            for account_elem in root.findall(f'.//{tag}'):
                policy_num = self._find_text(account_elem, 'MISPAR-POLISA-O-HESHBON')
                if policy_num and not any(a.get('policy_number') == policy_num for a in accounts):
                    account = self._parse_account(account_elem, {}, {})
                    if account:
                        accounts.append(account)
        
        return providers, accounts
    
    def _parse_account(self, elem, provider: Dict, product_info: Dict) -> Dict[str, Any]:
        """Parse a single account element."""
        account = {
            'provider': provider.get('name', ''),
            'provider_code': provider.get('code', ''),
            'product_type': product_info.get('product_type', ''),
            'product_type_code': product_info.get('product_type_code', ''),
            'product_name': product_info.get('name', ''),
        }
        
        # Extract account fields
        for xml_tag, field_name in self.schema_mapping.ACCOUNT_FIELDS.items():
            value = self._find_text(elem, xml_tag)
            if value:
                # Convert numeric fields
                if field_name in ['total_balance', 'savings_balance', 'severance_balance',
                                  'employer_severance', 'compensation_balance', 'coverage_amount',
                                  'monthly_pension', 'management_fee_savings', 'management_fee_deposits',
                                  'death_coverage', 'disability_coverage']:
                    account[field_name] = self._parse_number(value)
                elif field_name == 'section14':
                    account[field_name] = value in ['1', 'כן', 'true', 'True', 'Y', 'yes']
                else:
                    account[field_name] = value
        
        # Translate product type
        product_type_code = account.get('product_type_code', '') or account.get('product_type', '')
        if product_type_code in self.schema_mapping.PRODUCT_TYPE_CODES:
            type_info = self.schema_mapping.PRODUCT_TYPE_CODES[product_type_code]
            account['product_type_name'] = type_info['he']
            account['product_type_en'] = type_info['en']
        
        # Translate status
        status_code = account.get('status_code', '')
        if status_code in self.schema_mapping.STATUS_CODES:
            status_info = self.schema_mapping.STATUS_CODES[status_code]
            account['status'] = status_info['he']
            account['status_en'] = status_info['en']
        elif not account.get('status'):
            account['status'] = 'פעיל'
            account['status_en'] = 'Active'
        
        return account
    
    def _parse_contributions(self, root) -> List[Dict[str, Any]]:
        """Parse contributions (NetuneiHafrasha / PirteiHafrasha)."""
        contributions = []
        
        contrib_tags = ['NetuneiHafrasha', 'PirteiHafrasha', 'Hafrasha', 'ReshimatHafrashot', 'Peula', 'Event', 'Transaction']
        
        for tag in contrib_tags:
            for elem in root.findall(f'.//{tag}'):
                contrib = {}
                
                for xml_tag, field_name in self.schema_mapping.CONTRIBUTION_FIELDS.items():
                    value = self._find_text(elem, xml_tag)
                    if value:
                        if field_name in ['employee_amount', 'employer_amount', 
                                         'severance_amount', 'total_amount', 'salary_base']:
                            contrib[field_name] = self._parse_number(value)
                        else:
                            contrib[field_name] = value
                
                if contrib:
                    contributions.append(contrib)
        
        return contributions
    
    def _parse_severance(self, root, interface_code: int) -> List[Dict[str, Any]]:
        """Parse severance data (NetuneiPitzuim)."""
        severance_list = []
        
        sev_tags = ['NetuneiPitzuim', 'PirteiPitzuim', 'Pitzuim', 'SeveranceDetails']
        
        for tag in sev_tags:
            for elem in root.findall(f'.//{tag}'):
                sev = {}
                
                for xml_tag, field_name in self.schema_mapping.SEVERANCE_FIELDS.items():
                    value = self._find_text(elem, xml_tag)
                    if value:
                        if field_name in ['total_severance', 'available_severance', 
                                         'section14_amount', 'section14_percentage']:
                            sev[field_name] = self._parse_number(value)
                        elif field_name == 'section14':
                            # Section 14 codes: 1 = Yes, 2 = No (from schema)
                            sev[field_name] = value in ['1', 'כן', 'true', 'True', 'Y', 'yes']
                        else:
                            sev[field_name] = value
                
                if sev:
                    severance_list.append(sev)
        
        return severance_list
    
    def _parse_employers(self, root, accounts: List[Dict]) -> List[Dict[str, Any]]:
        """Parse employer information."""
        employers = []
        seen = set()
        
        # Collect from accounts
        for acct in accounts:
            emp_id = acct.get('employer_id', '')
            emp_name = acct.get('employer_name', '')
            if emp_name and emp_name not in seen:
                seen.add(emp_name)
                employers.append({'id': emp_id, 'name': emp_name})
        
        # Find standalone employer elements
        for elem in root.findall('.//YeshutMaasik'):
            emp_name = self._find_text(elem, 'SHEM-MAASIK') or self._find_text(elem, 'ShemMaasik')
            emp_id = self._find_text(elem, 'KOD-MAASIK') or self._find_text(elem, 'KodMaasik')
            if emp_name and emp_name not in seen:
                seen.add(emp_name)
                employers.append({'id': emp_id, 'name': emp_name})
        
        return employers
    
    def _extract_all_elements(self, root) -> Dict[str, List[str]]:
        """Extract all text elements for additional analysis."""
        elements = {}
        
        def extract(elem, path=''):
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            
            if elem.text and elem.text.strip():
                if tag not in elements:
                    elements[tag] = []
                elements[tag].append(elem.text.strip())
            
            for child in elem:
                extract(child)
        
        extract(root)
        return elements
    
    def _find_text(self, elem, tag: str) -> Optional[str]:
        """Find text content of a tag, trying multiple naming conventions."""
        if elem is None:
            return None
        
        # Try exact match
        found = elem.find(f'.//{tag}')
        if found is not None and found.text:
            return found.text.strip()
        
        # Try without hyphens
        tag_no_hyphen = tag.replace('-', '')
        found = elem.find(f'.//{tag_no_hyphen}')
        if found is not None and found.text:
            return found.text.strip()
        
        # Try CamelCase
        tag_camel = ''.join(word.capitalize() for word in tag.split('-'))
        found = elem.find(f'.//{tag_camel}')
        if found is not None and found.text:
            return found.text.strip()
        
        return None
    
    def _parse_number(self, value: str) -> float:
        """Parse numeric string to float."""
        if not value:
            return 0.0
        try:
            # Remove formatting
            cleaned = str(value).replace(',', '').replace(' ', '')
            cleaned = cleaned.replace('₪', '').replace('$', '').replace('€', '')
            cleaned = cleaned.replace("'", '')  # Hebrew thousands separator
            return float(cleaned)
        except:
            return 0.0
    
    def _enrich_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Enrich parsed data with calculated metrics."""
        accounts = data.get('accounts', [])
        contributions = data.get('contributions', [])
        severance = data.get('severance', [])
        
        # Initialize totals if not present
        if 'totals' not in data:
            data['totals'] = {}
        
        totals = data['totals']
        
        # Calculate totals
        totals['total_balance'] = sum(float(a.get('total_balance', 0) or 0) for a in accounts)
        totals['total_savings'] = sum(float(a.get('savings_balance', 0) or 0) for a in accounts)
        totals['total_severance'] = sum(float(a.get('severance_balance', 0) or 0) for a in accounts)
        totals['total_severance'] += sum(float(s.get('total_severance', 0) or 0) for s in severance)
        totals['total_coverage'] = sum(float(a.get('coverage_amount', 0) or 0) for a in accounts)
        totals['account_count'] = len(accounts)
        totals['provider_count'] = len(set(a.get('provider', '') for a in accounts if a.get('provider')))
        totals['providers'] = list(set(a.get('provider', '') for a in accounts if a.get('provider')))
        
        # Format totals
        totals['total_balance_formatted'] = f"₪{totals['total_balance']:,.2f}"
        totals['total_savings_formatted'] = f"₪{totals['total_savings']:,.2f}"
        totals['total_severance_formatted'] = f"₪{totals['total_severance']:,.2f}"
        
        # Contribution analysis
        if contributions:
            total_employee = sum(float(c.get('employee_amount', 0) or 0) for c in contributions)
            total_employer = sum(float(c.get('employer_amount', 0) or 0) for c in contributions)
            total_sev_contrib = sum(float(c.get('severance_amount', 0) or 0) for c in contributions)
            
            totals['contributions'] = {
                'employee_total': total_employee,
                'employer_total': total_employer,
                'severance_total': total_sev_contrib,
                'grand_total': total_employee + total_employer + total_sev_contrib,
                'periods_count': len(contributions),
            }
            
            # Contribution trend
            try:
                sorted_contribs = sorted(contributions, key=lambda x: x.get('period', ''))
                if len(sorted_contribs) >= 2:
                    first = sorted_contribs[0]
                    last = sorted_contribs[-1]
                    first_total = float(first.get('employee_amount', 0) or 0) + float(first.get('employer_amount', 0) or 0)
                    last_total = float(last.get('employee_amount', 0) or 0) + float(last.get('employer_amount', 0) or 0)
                    
                    if last_total > first_total * 1.1:
                        totals['contribution_trend'] = 'increasing'
                        totals['contribution_trend_he'] = 'עולה'
                    elif last_total < first_total * 0.9:
                        totals['contribution_trend'] = 'decreasing'
                        totals['contribution_trend_he'] = 'יורד'
                    else:
                        totals['contribution_trend'] = 'stable'
                        totals['contribution_trend_he'] = 'יציב'
            except:
                pass
        
        # Section 14 status
        section14_accounts = [a for a in accounts if a.get('section14')]
        section14_severance = [s for s in severance if s.get('section14')]
        totals['section14_coverage'] = len(section14_accounts) > 0 or len(section14_severance) > 0
        totals['section14_accounts'] = len(section14_accounts)
        
        # Health score
        totals['health_score'] = self._calculate_health_score(totals)
        
        return data
    
    def _calculate_health_score(self, totals: Dict) -> Dict[str, Any]:
        """Calculate financial health score."""
        score = {
            'overall': 0,
            'savings': 0,
            'diversification': 0,
            'section14': 0,
            'rating': 'unknown',
            'rating_he': 'לא ידוע',
        }
        
        total_balance = totals.get('total_balance', 0)
        provider_count = totals.get('provider_count', 0)
        
        # Savings score
        if total_balance >= 1000000:
            score['savings'] = 100
        elif total_balance >= 500000:
            score['savings'] = 80
        elif total_balance >= 200000:
            score['savings'] = 60
        elif total_balance >= 50000:
            score['savings'] = 40
        else:
            score['savings'] = 20
        
        # Diversification score
        if provider_count >= 3:
            score['diversification'] = 70
        elif provider_count == 2:
            score['diversification'] = 90
        elif provider_count == 1:
            score['diversification'] = 70
        else:
            score['diversification'] = 50
        
        # Section 14 score
        if totals.get('section14_coverage'):
            score['section14'] = 100
        elif totals.get('total_severance', 0) > 0:
            score['section14'] = 70
        else:
            score['section14'] = 50
        
        # Overall
        score['overall'] = int(
            score['savings'] * 0.5 +
            score['diversification'] * 0.2 +
            score['section14'] * 0.3
        )
        
        if score['overall'] >= 80:
            score['rating'] = 'excellent'
            score['rating_he'] = 'מצוין'
        elif score['overall'] >= 60:
            score['rating'] = 'good'
            score['rating_he'] = 'טוב'
        elif score['overall'] >= 40:
            score['rating'] = 'fair'
            score['rating_he'] = 'סביר'
        else:
            score['rating'] = 'needs_attention'
            score['rating_he'] = 'דורש תשומת לב'
        
        return score
    
    def _generate_professional_report(self, data: Dict[str, Any]) -> str:
        """
        Generate professional data-focused Hebrew pension statement.
        Shows actual Mislaka data clearly: name, ID, policies, savings, fees, status.
        Format similar to professional pension portals (sms2010.co.il style).
        """
        lines = []
        header = data.get('header', {})
        client = data.get('client', {})
        accounts = data.get('accounts', [])
        totals = data.get('totals', {})
        contributions = data.get('contributions', [])
        severance = data.get('severance', [])
        
        # Format report date
        report_date = header.get('report_date') or header.get('created_at') or datetime.now().strftime('%Y%m%d')
        if len(str(report_date)) >= 8:
            try:
                formatted_date = f"{str(report_date)[6:8]}/{str(report_date)[4:6]}/{str(report_date)[:4]}"
            except:
                formatted_date = datetime.now().strftime('%d/%m/%Y')
        else:
            formatted_date = datetime.now().strftime('%d/%m/%Y')
        
        # Get client info
        client_name = client.get('full_name') or f"{client.get('first_name', '')} {client.get('last_name', '')}".strip() or 'לא זמין'
        id_number = client.get('id_number', '') or 'לא זמין'
        birth_date = client.get('birth_date', '')
        
        # Calculate age if birth date available
        age_str = ''
        if birth_date and len(str(birth_date)) >= 8:
            try:
                birth_year = int(str(birth_date)[:4])
                current_year = datetime.now().year
                age = current_year - birth_year
                age_str = f", גיל {age}"
            except:
                pass
        
        # ===== REPORT HEADER =====
        lines.extend([
            "══════════════════════════════════════════════════════════════════════",
            "                    דו״ח נתוני פנסיה וביטוח                           ",
            "                    מסלקת הביטוח והפנסיה                              ",
            "══════════════════════════════════════════════════════════════════════",
            "",
            f"📅 נכון לתאריך: {formatted_date}",
            "",
        ])
        
        # ===== CLIENT SUMMARY LINE (like sms2010 style) =====
        total_balance = totals.get('total_balance', 0)
        account_count = totals.get('account_count', 0)
        
        lines.extend([
            "══════════════════════════════════════════════════════════════════════",
            "                         פרטי המבוטח                                  ",
            "══════════════════════════════════════════════════════════════════════",
            "",
            f"  👤 שם:              {client_name}",
            f"  🪪 ת.ז:             {id_number}{age_str}",
        ])
        
        if birth_date:
            if len(str(birth_date)) >= 8:
                try:
                    bd_formatted = f"{str(birth_date)[6:8]}/{str(birth_date)[4:6]}/{str(birth_date)[:4]}"
                    lines.append(f"  🎂 תאריך לידה:      {bd_formatted}")
                except:
                    lines.append(f"  🎂 תאריך לידה:      {birth_date}")
        
        if client.get('phone'):
            lines.append(f"  📱 טלפון:           {client.get('phone')}")
        if client.get('email'):
            lines.append(f"  📧 דוא״ל:           {client.get('email')}")
        
        lines.append("")
        
        # ===== TOTAL SUMMARY =====
        lines.extend([
            "══════════════════════════════════════════════════════════════════════",
            "                       סיכום כולל                                     ",
            "══════════════════════════════════════════════════════════════════════",
            "",
            f"  💰 סה״כ צבירה:                    ₪{total_balance:,.0f}",
            f"  📊 מספר מוצרים:                   {account_count}",
        ])
        
        if totals.get('total_severance', 0) > 0:
            lines.append(f"  💼 סה״כ פיצויים:                  ₪{totals.get('total_severance', 0):,.0f}")
        
        if totals.get('total_coverage', 0) > 0:
            lines.append(f"  🛡️ סה״כ כיסויים ביטוחיים:        ₪{totals.get('total_coverage', 0):,.0f}")
        
        lines.append("")
        
        # ===== DETAILED POLICY LIST =====
        if accounts:
            lines.extend([
                "══════════════════════════════════════════════════════════════════════",
                "                      פירוט מוצרים                                    ",
                "══════════════════════════════════════════════════════════════════════",
                "",
            ])
            
            for i, acct in enumerate(accounts, 1):
                balance = float(acct.get('total_balance', 0) or 0)
                savings = float(acct.get('savings_balance', 0) or 0)
                severance_bal = float(acct.get('severance_balance', 0) or 0)
                mgmt_fee = float(acct.get('management_fee_savings', 0) or 0)
                mgmt_fee_deposits = float(acct.get('management_fee_deposits', 0) or 0)
                
                provider = acct.get('provider', 'לא זמין')
                policy_num = acct.get('policy_number', 'לא זמין')
                product_type = acct.get('product_type_name', acct.get('product_type', ''))
                status = acct.get('status', 'פעיל')
                status_en = acct.get('status_en', 'active')
                
                # Determine liquidity status
                liquidity = 'נזיל' if status_en in ['active', 'Active'] else 'מוקפא'
                if status == 'מוקפא' or status_en == 'frozen':
                    liquidity = 'מוקפא'
                
                # Section 14 status
                section14_text = '✅ כן' if acct.get('section14') else '❌ לא'
                
                lines.append(f"  ┌─────────────────────────────────────────────────────────────────")
                lines.append(f"  │  📋 מוצר {i}: {product_type}")
                lines.append(f"  ├─────────────────────────────────────────────────────────────────")
                lines.append(f"  │  🏢 יצרן:           {provider}")
                lines.append(f"  │  🔢 מספר פוליסה:   {policy_num}")
                lines.append(f"  │  📌 סטטוס:         {status} ({liquidity})")
                lines.append(f"  │")
                lines.append(f"  │  💰 צבירה כוללת:   ₪{balance:,.0f}")
                
                if savings > 0 and savings != balance:
                    lines.append(f"  │     • תגמולים:     ₪{savings:,.0f}")
                if severance_bal > 0:
                    lines.append(f"  │     • פיצויים:     ₪{severance_bal:,.0f}")
                
                if mgmt_fee > 0 or mgmt_fee_deposits > 0:
                    lines.append(f"  │")
                    lines.append(f"  │  💳 דמי ניהול:")
                    if mgmt_fee > 0:
                        lines.append(f"  │     • מצבירה:      {mgmt_fee:.2f}%")
                    if mgmt_fee_deposits > 0:
                        lines.append(f"  │     • מהפקדות:     {mgmt_fee_deposits:.2f}%")
                
                lines.append(f"  │")
                lines.append(f"  │  📌 סעיף 14:       {section14_text}")
                
                # Employer info
                if acct.get('employer_name'):
                    lines.append(f"  │  🏭 מעסיק:         {acct.get('employer_name')}")
                
                # Insurance coverage
                death_coverage = float(acct.get('death_coverage', 0) or 0)
                disability_coverage = float(acct.get('disability_coverage', 0) or 0)
                
                if death_coverage > 0 or disability_coverage > 0:
                    lines.append(f"  │")
                    lines.append(f"  │  🛡️ כיסויים ביטוחיים:")
                    if death_coverage > 0:
                        lines.append(f"  │     • ביטוח חיים:  ₪{death_coverage:,.0f}")
                    if disability_coverage > 0:
                        lines.append(f"  │     • אובדן כושר:  ₪{disability_coverage:,.0f}/חודש")
                
                # Start date if available
                if acct.get('start_date'):
                    start_date = acct.get('start_date')
                    if len(str(start_date)) >= 8:
                        try:
                            sd_formatted = f"{str(start_date)[6:8]}/{str(start_date)[4:6]}/{str(start_date)[:4]}"
                            lines.append(f"  │  📅 תחילת ביטוח:   {sd_formatted}")
                        except:
                            pass
                
                lines.append(f"  └─────────────────────────────────────────────────────────────────")
                lines.append("")
        
        # ===== CONTRIBUTIONS SUMMARY =====
        contrib_totals = totals.get('contributions', {})
        if contrib_totals and contrib_totals.get('grand_total', 0) > 0:
            lines.extend([
                "══════════════════════════════════════════════════════════════════════",
                "                      הפקדות אחרונות                                  ",
                "══════════════════════════════════════════════════════════════════════",
                "",
                f"  📥 הפקדות עובד:                   ₪{contrib_totals.get('employee_total', 0):,.0f}",
                f"  📤 הפקדות מעסיק:                  ₪{contrib_totals.get('employer_total', 0):,.0f}",
            ])
            if contrib_totals.get('severance_total', 0) > 0:
                lines.append(f"  💼 הפקדות לפיצויים:               ₪{contrib_totals.get('severance_total', 0):,.0f}")
            lines.append(f"  ─────────────────────────────────────────────────────────────")
            lines.append(f"  📊 סה״כ הפקדות:                   ₪{contrib_totals.get('grand_total', 0):,.0f}")
            lines.append("")
        
        # ===== SEVERANCE & TAX INFO =====
        if totals.get('total_severance', 0) > 0 or totals.get('section14_coverage'):
            lines.extend([
                "══════════════════════════════════════════════════════════════════════",
                "                    פיצויים ומיסוי                                    ",
                "══════════════════════════════════════════════════════════════════════",
                "",
            ])
            
            if totals.get('total_severance', 0) > 0:
                lines.append(f"  💼 סה״כ פיצויים צבורים:           ₪{totals.get('total_severance', 0):,.0f}")
            
            if totals.get('section14_coverage'):
                lines.extend([
                    f"  📌 סעיף 14:                       ✅ פעיל",
                    f"",
                    f"  📋 משמעות סעיף 14:",
                    f"     • פיצויי הפיטורין שייכים לעובד במלואם",
                    f"     • לא נדרש אישור מעסיק למשיכה",
                    f"     • הכספים מוגנים גם בהתפטרות",
                ])
            else:
                lines.extend([
                    f"  📌 סעיף 14:                       ❌ לא פעיל",
                    f"",
                    f"  ⚠️ שים לב:",
                    f"     • פיצויים עשויים להיות תלויים באישור מעסיק",
                ])
            
            # Tax info
            lines.extend([
                f"",
                f"  💰 מידע מיסויי:",
                f"     • משיכת כספי פיצויים עד תקרה פטורה ממס",
                f"     • תקרת פטור: כ-12,640 ש״ח לכל שנת עבודה",
                f"     • מעל התקרה - מס שולי לפי מדרגות",
            ])
            lines.append("")
        
        # ===== SUMMARY BOX =====
        lines.extend([
            "══════════════════════════════════════════════════════════════════════",
            "                         סיכום                                        ",
            "══════════════════════════════════════════════════════════════════════",
            "",
            f"  {client_name}, ת.ז {id_number}",
            "",
        ])
        
        # Generate summary sentences for each account
        for acct in accounts[:5]:
            provider = acct.get('provider', '')
            product = acct.get('product_type_name', acct.get('product_type', ''))
            balance = float(acct.get('total_balance', 0) or 0)
            mgmt_fee = float(acct.get('management_fee_savings', 0) or 0)
            status = acct.get('status', 'פעיל')
            status_en = acct.get('status_en', 'active')
            liquidity = 'נזיל' if status_en in ['active', 'Active'] else 'מוקפא'
            
            summary_line = f"  • {product} ב{provider}"
            if balance > 0:
                summary_line += f" עם ₪{balance:,.0f} צבירה"
            if mgmt_fee > 0:
                summary_line += f", {mgmt_fee:.2f}% דמי ניהול"
            summary_line += f", סטטוס {liquidity}"
            
            lines.append(summary_line)
        
        if len(accounts) > 5:
            lines.append(f"  • ועוד {len(accounts) - 5} מוצרים נוספים...")
        
        lines.append("")
        lines.append(f"  💰 סה״כ צבירה בכל המוצרים: ₪{total_balance:,.0f}")
        lines.append("")
        
        # ===== FOOTER =====
        lines.extend([
            "══════════════════════════════════════════════════════════════════════",
            f"  📅 דו״ח הופק: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            f"  📊 מקור נתונים: מסלקת הביטוח והפנסיה",
            f"  🔗 PHINS - פלטפורמת ניהול פנסיה וביטוח",
            "══════════════════════════════════════════════════════════════════════",
        ])
        
        return '\n'.join(lines)
    
    def _generate_recommendations(self, data: Dict, totals: Dict, health: Dict) -> List[Dict]:
        """Generate AI recommendations based on comprehensive data analysis."""
        recommendations = []
        
        total_balance = totals.get('total_balance', 0)
        provider_count = totals.get('provider_count', 0)
        section14 = totals.get('section14_coverage', False)
        accounts = data.get('accounts', [])
        
        # Low savings warning
        if total_balance < 100000:
            recommendations.append({
                'priority': 'high',
                'title': 'הגדלת החיסכון הפנסיוני',
                'description': 'החיסכון הנוכחי נמוך מהמומלץ. שקול להגדיל את אחוזי ההפקדה או להפקיד סכומים נוספים באופן עצמאי.'
            })
        
        # Too many providers - consolidation
        if provider_count > 3:
            recommendations.append({
                'priority': 'medium',
                'title': 'איחוד חשבונות פנסיה',
                'description': f'יש לך חשבונות ב-{provider_count} יצרנים שונים. איחוד יכול להפחית דמי ניהול ולפשט את הניהול והמעקב.'
            })
        
        # No Section 14 coverage
        if not section14 and totals.get('total_severance', 0) > 0:
            recommendations.append({
                'priority': 'high',
                'title': 'בדיקת סעיף 14',
                'description': 'מומלץ לבדוק אפשרות להסדר סעיף 14 עם המעסיק להבטחת כספי הפיצויים והגנה במקרה של עזיבה.'
            })
        
        # Check for inactive accounts
        inactive_accounts = [a for a in accounts if a.get('status_en') in ['Frozen', 'Closed', 'Transferred']]
        if inactive_accounts:
            recommendations.append({
                'priority': 'medium',
                'title': 'בדיקת חשבונות לא פעילים',
                'description': f'נמצאו {len(inactive_accounts)} חשבונות לא פעילים. מומלץ לבדוק את מצבם ולשקול העברה או איחוד.'
            })
        
        # High management fees check
        high_fee_accounts = [a for a in accounts if float(a.get('management_fee_savings', 0) or 0) > 1.0]
        if high_fee_accounts:
            recommendations.append({
                'priority': 'medium',
                'title': 'בדיקת דמי ניהול גבוהים',
                'description': f'{len(high_fee_accounts)} חשבונות עם דמי ניהול מעל 1%. מומלץ לבדוק ולהשוות מול הצעות מתחרות.'
            })
        
        # Good standing message
        if health.get('overall', 0) >= 70:
            recommendations.append({
                'priority': 'low',
                'title': 'המשך מעקב שוטף',
                'description': 'המצב הפיננסי טוב. המשך לעקוב אחר ההפקדות ובצע בדיקה שנתית של תנאי הקופות.'
            })
        
        # Always recommend annual review
        recommendations.append({
            'priority': 'low',
            'title': 'בדיקה שנתית מקיפה',
            'description': 'מומלץ לבצע בדיקה שנתית של כל הקופות, להשוות דמי ניהול ותשואות, ולעדכן מוטבים.'
        })
        
        return recommendations[:6]
    
    def _mask_id(self, id_number: str) -> str:
        """Mask ID number for privacy."""
        if not id_number or len(id_number) < 5:
            return id_number or '***'
        return id_number[:2] + '*' * (len(id_number) - 4) + id_number[-2:]
    
    # ===== PUBLIC API =====
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """Process XML file from disk."""
        with open(file_path, 'rb') as f:
            content = f.read()
        
        # Check if ZIP
        if file_path.lower().endswith('.zip') or content[:4] == b'PK\x03\x04':
            return self.process_zip_content(content)
        else:
            return self.process_xml_content(content)
    
    def to_csv_format(self, data: Dict[str, Any]) -> Tuple[List[str], List[Dict]]:
        """Convert to CSV format for AI analysis integration."""
        columns = [
            'מספר פוליסה', 'יצרן', 'סוג מוצר', 'שם מוצר',
            'סטטוס', 'יתרה כוללת', 'חיסכון', 'פיצויים', 'מעסיק', 'סעיף 14'
        ]
        
        rows = []
        for acct in data.get('accounts', []):
            rows.append({
                'מספר פוליסה': acct.get('policy_number', ''),
                'יצרן': acct.get('provider', ''),
                'סוג מוצר': acct.get('product_type_name', acct.get('product_type', '')),
                'שם מוצר': acct.get('product_name', ''),
                'סטטוס': acct.get('status', 'פעיל'),
                'יתרה כוללת': acct.get('total_balance', 0),
                'חיסכון': acct.get('savings_balance', 0),
                'פיצויים': acct.get('severance_balance', 0),
                'מעסיק': acct.get('employer_name', ''),
                'סעיף 14': 'כן' if acct.get('section14') else 'לא'
            })
        
        # Summary row
        totals = data.get('totals', {})
        rows.append({
            'מספר פוליסה': 'סה״כ',
            'יצרן': '',
            'סוג מוצר': '',
            'שם מוצר': '',
            'סטטוס': '',
            'יתרה כוללת': totals.get('total_balance', 0),
            'חיסכון': totals.get('total_savings', 0),
            'פיצויים': totals.get('total_severance', 0),
            'מעסיק': '',
            'סעיף 14': ''
        })
        
        return columns, rows
    
    def generate_report_text(self, data: Dict[str, Any], language: str = 'hebrew') -> str:
        """Generate report text (alias for compatibility)."""
        return self._generate_professional_report(data)


# ============================================================================
# SINGLETON AND HELPERS
# ============================================================================

_pension_agent = None


def get_pension_agent() -> PensionDataAgent:
    """Get or create PensionDataAgent singleton."""
    global _pension_agent
    if _pension_agent is None:
        _pension_agent = PensionDataAgent()
    return _pension_agent


def is_pension_xml(content: bytes) -> bool:
    """Check if content appears to be Mislaka pension XML."""
    try:
        if not content.strip().startswith(b'<?xml') and not content.strip().startswith(b'<'):
            return False
        
        # Try multiple encodings
        try:
            content_str = content.decode('utf-8', errors='replace')[:5000]
        except:
            content_str = content.decode('windows-1255', errors='replace')[:5000]
        
        markers = [
            'SUG-MIMSHAK', 'SugMimshak',
            'KoteretKovetz', 'YeshutYatzran',
            'HeshbonOPolisa', 'PirteiHeshbon',
            'MISPAR-POLISA', 'MisparPolisa',
            'SHEM-YATZRAN', 'ShemYatzran',
            'HAFRASHA-OVED', 'HafrashaOved',
            'NetuneiPitzuim', 'PITZUIM',
            'YeshutLakoach', 'Mutzar',
            'MISPAR-ZIHUI-LAKOACH', 'MisparZihuiLakoach',
        ]
        
        return any(marker in content_str for marker in markers)
    except:
        return False
