"""
Document Processing & Metadata Mining Service
===============================================
Universal document handling pipeline that works across the entire PHINS platform.

Capabilities:
- Persistent file storage (survives restarts via disk + database)
- SHA-256 / MD5 integrity verification on every upload and retrieval
- Automatic MIME type detection and category classification
- Metadata extraction from images, PDFs, Office docs, CSV/XLS tables, audio, video
- Identity document analysis (ID cards, passports, driver licences)
- Medical document parsing
- Legal document parsing
- Table / spreadsheet extraction
- AI-powered tagging and summarisation
- Processing job queue with status tracking
- Duplicate detection via content-addressed checksums
- Version management for updated documents

Designed as the single entry point for *all* file operations so that every
upload path (policy creation, claim filing, underwriting, reports, media)
gets the same integrity guarantees, persistence, and processing pipeline.
"""

import base64
import hashlib
import json
import logging
import mimetypes
import os
import re
import shutil
import tempfile
import time
import uuid
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ── Configuration ─────────────────────────────────────────────────────────────

def _resolve_document_storage_root() -> str:
    """Pick a Railway-volume-aware storage path for uploaded documents.

    Priority:
      1. ``PHINS_DOCUMENT_STORAGE`` (explicit override)
      2. ``RAILWAY_VOLUME_MOUNT_PATH/documents`` (Railway volume)
      3. ``/data/documents`` (Docker volume mount, gated by Railway-or-
         opt-in detection so dev machines with a writable ``/data`` are
         not hijacked - see ``_data_volume_eligible``)
      4. ``<repo>/data/documents`` (developer fallback - ephemeral)
    """
    explicit = os.environ.get('PHINS_DOCUMENT_STORAGE')
    if explicit:
        return explicit

    railway_mount = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', '').strip()
    if railway_mount and os.path.isdir(railway_mount):
        return os.path.join(railway_mount, 'documents')

    # Reuse the eligibility gate from assessment_center_service so a single
    # signal (Railway env vars or PHINS_USE_DATA_VOLUME=1) governs both
    # persistence paths consistently.
    try:
        from services.assessment_center_service import _data_volume_eligible
        if _data_volume_eligible():
            return '/data/documents'
    except Exception:
        pass

    fallback = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'data', 'documents',
    )
    if not os.environ.get('PHINS_TEST_MODE'):
        print(
            f"⚠️  [doc-service] Using ephemeral document storage {fallback} - "
            "set PHINS_DOCUMENT_STORAGE or mount a Railway volume at /data "
            "for durable persistence.",
            flush=True,
        )
    return fallback


DOCUMENT_STORAGE_ROOT = _resolve_document_storage_root()

MAX_DOCUMENT_SIZE_BYTES = int(os.environ.get('PHINS_MAX_DOCUMENT_SIZE', 50 * 1024 * 1024))
MAX_BATCH_SIZE = int(os.environ.get('PHINS_MAX_BATCH_UPLOAD', 25))

ALLOWED_EXTENSIONS = {
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.csv', '.txt', '.rtf',
    '.json', '.xml', '.html', '.htm',
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg',
    '.mp4', '.avi', '.mov', '.wmv', '.mkv', '.webm', '.flv',
    '.mp3', '.wav', '.ogg', '.flac', '.aac', '.m4a', '.wma',
    '.zip', '.gz', '.tar', '.7z', '.rar',
    '.dicom', '.dcm',
    '.xsd',
}

CATEGORY_BY_MIME_PREFIX = {
    'image/': 'identity',
    'video/': 'media',
    'audio/': 'media',
    'text/csv': 'table',
    'application/vnd.ms-excel': 'table',
    'application/vnd.openxmlformats-officedocument.spreadsheetml': 'table',
    'application/pdf': 'general',
}

CATEGORY_BY_EXTENSION = {
    '.csv': 'table', '.xls': 'table', '.xlsx': 'table',
    '.pdf': 'general', '.doc': 'legal', '.docx': 'legal',
    '.dicom': 'medical', '.dcm': 'medical',
    '.mp4': 'media', '.avi': 'media', '.mov': 'media',
    '.mp3': 'media', '.wav': 'media', '.ogg': 'media',
    '.jpg': 'identity', '.jpeg': 'identity', '.png': 'identity',
    '.xsd': 'policy',
}


ALLOWED_CATEGORIES = {
    'identity', 'legal', 'medical', 'financial', 'policy', 'claim',
    'underwriting', 'report', 'media', 'table', 'general',
}


class ProcessingJobType(str, Enum):
    METADATA_EXTRACTION = 'metadata_extraction'
    TEXT_EXTRACTION = 'text_extraction'
    TABLE_EXTRACTION = 'table_extraction'
    IMAGE_ANALYSIS = 'image_analysis'
    IDENTITY_VERIFICATION = 'identity_verification'
    MEDICAL_ANALYSIS = 'medical_analysis'
    LEGAL_ANALYSIS = 'legal_analysis'
    AUDIO_TRANSCRIPTION = 'audio_transcription'
    VIDEO_ANALYSIS = 'video_analysis'
    INTEGRITY_CHECK = 'integrity_check'
    AI_SUMMARISATION = 'ai_summarisation'
    AI_TAGGING = 'ai_tagging'
    # Full upload enrichment (metadata + text/OCR + summary/tags/confidence)
    # run as a single queued unit by the async document worker.
    DOCUMENT_ENRICHMENT = 'document_enrichment'


def async_processing_enabled() -> bool:
    """True when uploads should enqueue enrichment instead of running inline.

    Controlled by ``PHINS_DOC_ASYNC``. Defaults to off so the synchronous
    behaviour (and the pytest embedded-server semantics) is unchanged until
    an operator opts in.
    """
    return str(os.environ.get('PHINS_DOC_ASYNC', '')).lower() in ('1', 'true', 'yes', 'y')


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class UploadResult:
    document_id: str
    file_name: str
    file_size: int
    mime_type: str
    sha256: str
    category: str
    status: str
    storage_path: str
    duplicate_of: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class ProcessingResult:
    job_id: str
    document_id: str
    job_type: str
    status: str
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    processing_time_ms: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


# ── Core Service ──────────────────────────────────────────────────────────────

class DocumentProcessingService:
    """Unified document storage, integrity, and processing pipeline."""

    def __init__(self, storage_root: Optional[str] = None, db_manager=None):
        self.storage_root = storage_root or DOCUMENT_STORAGE_ROOT
        self.db_manager = db_manager
        self._inmemory_store: Dict[str, Dict[str, Any]] = {}
        os.makedirs(self.storage_root, exist_ok=True)

    # ── Upload ────────────────────────────────────────────────────────────────

    def upload_document(
        self,
        *,
        file_name: str,
        file_data_b64: str,
        mime_type: Optional[str] = None,
        category: Optional[str] = None,
        document_type: Optional[str] = None,
        description: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[str] = None,
        customer_id: Optional[str] = None,
        uploaded_by: Optional[str] = None,
        uploaded_by_role: Optional[str] = None,
        parent_document_id: Optional[str] = None,
        skip_processing: bool = False,
    ) -> UploadResult:
        """Store a document persistently and queue processing jobs."""

        if not file_data_b64:
            raise ValueError('Missing file content')
        if not file_name:
            raise ValueError('Missing file name')

        raw_bytes = base64.b64decode(file_data_b64, validate=True)
        if not raw_bytes:
            raise ValueError('Document content cannot be empty')
        if len(raw_bytes) > MAX_DOCUMENT_SIZE_BYTES:
            raise ValueError(f'File exceeds maximum allowed size ({MAX_DOCUMENT_SIZE_BYTES // (1024*1024)}MB)')

        ext = self._get_extension(file_name)
        if ext and ext not in ALLOWED_EXTENSIONS:
            raise ValueError(f'File type {ext} not allowed')

        resolved_mime = mime_type or mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
        resolved_category = self._sanitise_category(
            category or self._classify_category(file_name, resolved_mime)
        )

        sha256 = hashlib.sha256(raw_bytes).hexdigest()
        md5 = hashlib.md5(raw_bytes).hexdigest()

        duplicate_id = self._check_duplicate(sha256, entity_type, entity_id)

        doc_id = f"DOC-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8].upper()}"
        safe_name = self._safe_filename(doc_id, file_name)
        storage_path = self._write_to_disk(safe_name, raw_bytes, resolved_category)

        if self._verify_on_disk(storage_path, sha256):
            status = 'uploaded'
        else:
            status = 'integrity_error'
            logger.error(f"Integrity check failed for {doc_id} at {storage_path}")

        doc_record = {
            'id': doc_id,
            'file_name': safe_name,
            'original_file_name': file_name,
            'mime_type': resolved_mime,
            'file_size': len(raw_bytes),
            'file_extension': ext,
            'storage_path': storage_path,
            'sha256_checksum': sha256,
            'md5_checksum': md5,
            'category': resolved_category,
            'document_type': document_type or 'general',
            'description': description or '',
            'entity_type': entity_type or '',
            'entity_id': entity_id or '',
            'customer_id': customer_id or '',
            'uploaded_by': uploaded_by or 'system',
            'uploaded_by_role': uploaded_by_role or '',
            'status': status,
            'processing_status': 'pending' if not skip_processing else None,
            'is_archived': False,
            'is_deleted': False,
            'version': 1,
            'parent_document_id': parent_document_id,
        }

        self._persist_record(doc_record)

        extracted = {}
        queued = False
        if not skip_processing and status == 'uploaded':
            if async_processing_enabled():
                queued = self._enqueue_enrichment(doc_id, sha256)
            if not queued:
                extracted = self._apply_enrichment(doc_id, raw_bytes, resolved_mime, ext)

        return UploadResult(
            document_id=doc_id,
            file_name=safe_name,
            file_size=len(raw_bytes),
            mime_type=resolved_mime,
            sha256=sha256,
            category=resolved_category,
            status='processed' if extracted else status,
            storage_path=storage_path,
            duplicate_of=duplicate_id,
            metadata={'queued': True} if queued else extracted,
        )

    def _enqueue_enrichment(self, doc_id: str, sha256: str) -> bool:
        """Queue the enrichment job for the async worker. Returns False on any
        failure so the caller falls back to synchronous processing (a document
        must never end up unprocessed just because the queue was unavailable)."""
        try:
            from services.document_job_worker import get_document_job_worker
            worker = get_document_job_worker(doc_service=self, db_manager=self.db_manager)
            worker.enqueue(
                document_id=doc_id,
                job_type=ProcessingJobType.DOCUMENT_ENRICHMENT.value,
                idempotency_key=f"{sha256}:{ProcessingJobType.DOCUMENT_ENRICHMENT.value}",
            )
            self._update_record(doc_id, {'processing_status': 'queued'})
            return True
        except Exception as exc:
            logger.error(f"Enqueue failed for {doc_id}, falling back to sync: {exc}")
            return False

    def _apply_enrichment(self, doc_id: str, raw_bytes: bytes, mime: str, ext: str) -> Dict[str, Any]:
        """Run full enrichment and persist the results onto the document row."""
        start = time.time()
        extracted = self._run_immediate_processing(doc_id, raw_bytes, mime, ext)
        self._update_record(doc_id, {
            'status': 'processed',
            'processing_status': 'completed',
            'extracted_metadata': json.dumps(extracted.get('metadata', {})),
            'extracted_text': extracted.get('text', ''),
            'ai_summary': extracted.get('summary', ''),
            'ai_tags': json.dumps(extracted.get('tags', [])),
            'confidence_score': extracted.get('confidence', None),
            'processed_date': datetime.now(),
        })
        self._meter_parse_usage(doc_id, extracted, int((time.time() - start) * 1000))
        return extracted

    def _meter_parse_usage(self, doc_id: str, extracted: Dict[str, Any],
                           duration_ms: int) -> None:
        """Record parse usage for cost accounting (self-hosted OCR meters $0
        unless a managed-parser page price is configured). Never fatal."""
        try:
            from services.ai_usage_service import get_ai_usage_service
            record = self._load_record(doc_id)
            customer_id = None
            if record is not None:
                customer_id = (record.get('customer_id') if isinstance(record, dict)
                               else record.customer_id) or None
            pages = (extracted.get('metadata') or {}).get('pages')
            get_ai_usage_service().record_usage(
                provider='self_hosted',
                operation='document_parse',
                customer_id=customer_id,
                document_id=doc_id,
                pages=len(pages) if isinstance(pages, list) else None,
                duration_ms=duration_ms,
            )
        except Exception as exc:
            logger.debug('Parse usage metering skipped: %s', exc)

    def run_enrichment(self, doc_id: str) -> Dict[str, Any]:
        """Load a stored document and run the full enrichment pipeline.

        Used by the async worker; idempotent — re-running simply recomputes
        and overwrites the derived fields (extracted text/summary/tags),
        never the source bytes or checksums.
        """
        record = self._load_record(doc_id)
        if not record:
            raise ValueError(f'Document {doc_id} not found')
        path = record.get('storage_path') if isinstance(record, dict) else record.storage_path
        mime = (record.get('mime_type') if isinstance(record, dict) else record.mime_type) or ''
        ext = (record.get('file_extension') if isinstance(record, dict) else record.file_extension) or ''
        expected_sha = record.get('sha256_checksum') if isinstance(record, dict) else record.sha256_checksum
        raw = self._read_from_disk(path)
        if raw is None:
            raise ValueError(f'File not found on disk for {doc_id}')
        if expected_sha and hashlib.sha256(raw).hexdigest() != expected_sha:
            raise ValueError(f'Integrity check failed for {doc_id} before enrichment')
        return self._apply_enrichment(doc_id, raw, mime, ext)

    def execute_job(self, document_id: str, job_type: str) -> Dict[str, Any]:
        """Execute one processing job payload for the async worker.

        Unlike :meth:`process_document` this does NOT create a new job row —
        the worker owns the job row lifecycle. Returns the raw result dict;
        raises on failure so the worker can apply retry/dead-letter handling.
        """
        if job_type == ProcessingJobType.DOCUMENT_ENRICHMENT.value:
            extracted = self.run_enrichment(document_id)
            return {
                'text_chars': len(extracted.get('text', '') or ''),
                'tags': extracted.get('tags', []),
                'confidence': extracted.get('confidence'),
            }
        record = self._load_record(document_id)
        if not record:
            raise ValueError(f'Document {document_id} not found')
        path = record.get('storage_path') if isinstance(record, dict) else record.storage_path
        mime = (record.get('mime_type') if isinstance(record, dict) else record.mime_type) or ''
        ext = (record.get('file_extension') if isinstance(record, dict) else record.file_extension) or ''
        raw = self._read_from_disk(path)
        if raw is None:
            raise ValueError(f'File not found on disk for {document_id}')
        handler = self._job_handlers.get(job_type)
        if not handler:
            raise ValueError(f'Unknown job type {job_type}')
        return handler(self, raw, mime, ext)

    def upload_batch(
        self,
        files: List[Dict[str, Any]],
        **common_kwargs,
    ) -> List[UploadResult]:
        """Upload multiple files in one call."""
        if len(files) > MAX_BATCH_SIZE:
            raise ValueError(f'Batch exceeds maximum of {MAX_BATCH_SIZE} files')
        results = []
        for f in files:
            try:
                kwargs = {**common_kwargs}
                kwargs['file_name'] = f.get('name') or f.get('file_name', 'unnamed')
                kwargs['file_data_b64'] = f.get('data') or f.get('file_data', '')
                kwargs['mime_type'] = f.get('type') or f.get('mime_type')
                results.append(self.upload_document(**kwargs))
            except Exception as exc:
                results.append(UploadResult(
                    document_id='',
                    file_name=f.get('name', 'unknown'),
                    file_size=0,
                    mime_type='',
                    sha256='',
                    category='',
                    status='error',
                    storage_path='',
                    metadata={'error': str(exc)},
                ))
        return results

    # ── Retrieval ─────────────────────────────────────────────────────────────

    def get_document(self, doc_id: str, include_data: bool = False) -> Optional[Dict[str, Any]]:
        record = self._load_record(doc_id)
        if not record:
            return None
        if isinstance(record, dict):
            result = dict(record)
            storage_path = record.get('storage_path', '')
            expected_sha = record.get('sha256_checksum')
        else:
            result = record.to_dict() if hasattr(record, 'to_dict') else {'id': record.id}
            storage_path = record.storage_path
            expected_sha = record.sha256_checksum
        if include_data:
            raw = self._read_from_disk(storage_path)
            if raw is not None:
                on_disk_sha = hashlib.sha256(raw).hexdigest()
                if on_disk_sha != expected_sha:
                    logger.error(f"Integrity mismatch for {doc_id}: expected "
                                 f"{expected_sha}, got {on_disk_sha}")
                    result['integrity_warning'] = True
                result['data'] = base64.b64encode(raw).decode('ascii')
                result['data_encoding'] = 'base64'
        return result

    def list_documents(
        self,
        entity_type: Optional[str] = None,
        entity_id: Optional[str] = None,
        customer_id: Optional[str] = None,
        category: Optional[str] = None,
        status: Optional[str] = None,
        search_query: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> Dict[str, Any]:
        offset = (page - 1) * page_size
        docs = self._search_records(
            entity_type=entity_type, entity_id=entity_id,
            customer_id=customer_id, category=category, status=status,
            query=search_query, limit=page_size, offset=offset,
        )
        total = self._count_records(
            entity_type=entity_type, entity_id=entity_id,
            customer_id=customer_id, category=category, status=status,
            query=search_query,
        )
        items = []
        for d in docs:
            if isinstance(d, dict):
                safe = {k: v for k, v in d.items() if k != 'data'}
                items.append(safe)
            else:
                items.append(d.to_dict())
        return {
            'items': items,
            'page': page,
            'page_size': page_size,
            'total': total,
        }

    def delete_document(self, doc_id: str, hard: bool = False) -> bool:
        record = self._load_record(doc_id)
        if not record:
            return False
        if hard:
            path = record.get('storage_path') if isinstance(record, dict) else record.storage_path
            if path and os.path.exists(path):
                os.remove(path)
            self._delete_record(doc_id)
        else:
            self._update_record(doc_id, {'is_deleted': True, 'status': 'archived'})
        return True

    def verify_integrity(self, doc_id: str) -> Dict[str, Any]:
        """Re-verify SHA-256 of stored file against the database record."""
        record = self._load_record(doc_id)
        if not record:
            return {'valid': False, 'error': 'Document not found'}
        path = record.get('storage_path') if isinstance(record, dict) else record.storage_path
        expected_sha = record.get('sha256_checksum') if isinstance(record, dict) else record.sha256_checksum
        raw = self._read_from_disk(path)
        if raw is None:
            return {'valid': False, 'error': 'File not found on disk'}
        actual = hashlib.sha256(raw).hexdigest()
        valid = actual == expected_sha
        return {
            'valid': valid,
            'expected_sha256': expected_sha,
            'actual_sha256': actual,
            'file_size': len(raw),
        }

    def get_statistics(self, customer_id: Optional[str] = None) -> Dict[str, Any]:
        if self.db_manager:
            try:
                return self.db_manager.documents.get_statistics(customer_id)
            except Exception as e:
                logger.error(f"DB stats error: {e}")
        return self._inmemory_stats(customer_id)

    # ── Processing pipeline ───────────────────────────────────────────────────

    def process_document(self, doc_id: str, job_types: Optional[List[str]] = None) -> List[ProcessingResult]:
        """Run one or more processing jobs on an already-uploaded document."""
        record = self._load_record(doc_id)
        if not record:
            raise ValueError(f'Document {doc_id} not found')

        path = record.get('storage_path') if isinstance(record, dict) else record.storage_path
        mime = record.get('mime_type') if isinstance(record, dict) else record.mime_type
        ext = record.get('file_extension') if isinstance(record, dict) else record.file_extension

        raw = self._read_from_disk(path)
        if raw is None:
            raise ValueError(f'File not found on disk for {doc_id}')

        if not job_types:
            job_types = self._default_jobs_for(mime, ext)

        results = []
        for jt in job_types:
            result = self._run_job(doc_id, jt, raw, mime, ext)
            results.append(result)

        return results

    def reprocess_document(self, doc_id: str) -> List[ProcessingResult]:
        """Re-run all default processing for a document."""
        return self.process_document(doc_id)

    # ── Internal: disk storage ────────────────────────────────────────────────

    @staticmethod
    def _sanitise_category(category: str) -> str:
        """Ensure category is a safe, whitelisted directory name (no path traversal)."""
        clean = re.sub(r'[^a-zA-Z0-9_-]', '', (category or 'general').strip())
        if clean not in ALLOWED_CATEGORIES:
            clean = 'general'
        return clean

    def _write_to_disk(self, safe_name: str, raw_bytes: bytes, category: str) -> str:
        safe_category = self._sanitise_category(category)
        cat_dir = os.path.join(self.storage_root, safe_category)
        os.makedirs(cat_dir, exist_ok=True)
        path = os.path.join(cat_dir, safe_name)
        resolved = os.path.realpath(path)
        if not resolved.startswith(os.path.realpath(self.storage_root)):
            raise ValueError('Invalid storage path')
        tmp_path = path + '.tmp'
        with open(tmp_path, 'wb') as f:
            f.write(raw_bytes)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, path)
        return path

    def _read_from_disk(self, path: str) -> Optional[bytes]:
        if not path or not os.path.exists(path):
            return None
        with open(path, 'rb') as f:
            return f.read()

    def _verify_on_disk(self, path: str, expected_sha256: str) -> bool:
        raw = self._read_from_disk(path)
        if raw is None:
            return False
        return hashlib.sha256(raw).hexdigest() == expected_sha256

    # ── Internal: record persistence (DB-first, in-memory fallback) ──────────

    def _persist_record(self, doc: Dict[str, Any]) -> None:
        if self.db_manager:
            try:
                self.db_manager.documents.create(**doc)
                return
            except Exception as e:
                logger.error(f"DB persist failed for {doc.get('id')}: {e}")
        self._inmemory_store[doc['id']] = doc

    def _load_record(self, doc_id: str):
        if self.db_manager:
            try:
                rec = self.db_manager.documents.get_by_id(doc_id)
                if rec:
                    return rec
            except Exception as e:
                logger.error(f"DB load failed for {doc_id}: {e}")
        return self._inmemory_store.get(doc_id)

    def _update_record(self, doc_id: str, updates: Dict[str, Any]) -> None:
        if self.db_manager:
            try:
                self.db_manager.documents.update(doc_id, **updates)
                return
            except Exception as e:
                logger.error(f"DB update failed for {doc_id}: {e}")
        rec = self._inmemory_store.get(doc_id)
        if rec:
            rec.update(updates)

    def _delete_record(self, doc_id: str) -> None:
        if self.db_manager:
            try:
                self.db_manager.documents.delete(doc_id)
                return
            except Exception as e:
                logger.error(f"DB delete failed for {doc_id}: {e}")
        self._inmemory_store.pop(doc_id, None)

    def _search_records(self, entity_type=None, entity_id=None, customer_id=None,
                        category=None, status=None, query=None,
                        limit=50, offset=0):
        if self.db_manager:
            try:
                return self.db_manager.documents.search_all(
                    query=query, entity_type=entity_type, entity_id=entity_id,
                    customer_id=customer_id, category=category, status=status,
                    limit=limit, offset=offset,
                )
            except Exception as e:
                logger.error(f"DB search error: {e}")

        docs = self._filter_inmemory(
            entity_type=entity_type, entity_id=entity_id,
            customer_id=customer_id, category=category, status=status,
        )
        if query:
            q_lower = query.lower()
            docs = [d for d in docs if q_lower in (d.get('file_name', '') + d.get('description', '')).lower()]
        return docs[offset:offset + limit]

    def _count_records(self, entity_type=None, entity_id=None, customer_id=None,
                       category=None, status=None, query=None, **_ignored) -> int:
        if self.db_manager:
            try:
                return self.db_manager.documents.count_filtered(
                    query=query, entity_type=entity_type, entity_id=entity_id,
                    customer_id=customer_id, category=category, status=status,
                )
            except Exception:
                pass
        docs = self._filter_inmemory(
            entity_type=entity_type, entity_id=entity_id,
            customer_id=customer_id, category=category, status=status,
        )
        if query:
            q_lower = query.lower()
            docs = [d for d in docs if q_lower in (d.get('file_name', '') + d.get('description', '')).lower()]
        return len(docs)

    def _filter_inmemory(self, entity_type=None, entity_id=None, customer_id=None,
                         category=None, status=None, **_ignored) -> List[Dict[str, Any]]:
        """Apply all filters to the in-memory store and return matching docs."""
        docs = [d for d in self._inmemory_store.values() if not d.get('is_deleted')]
        if entity_type:
            docs = [d for d in docs if d.get('entity_type') == entity_type]
        if entity_id:
            docs = [d for d in docs if d.get('entity_id') == entity_id]
        if customer_id:
            docs = [d for d in docs if d.get('customer_id') == customer_id]
        if category:
            docs = [d for d in docs if d.get('category') == category]
        if status:
            docs = [d for d in docs if d.get('status') == status]
        return docs

    def _check_duplicate(self, sha256: str, entity_type: Optional[str], entity_id: Optional[str]) -> Optional[str]:
        if self.db_manager:
            try:
                existing = self.db_manager.documents.get_by_checksum(sha256)
                if existing:
                    return existing.id
            except Exception:
                pass
        for doc_id, doc in self._inmemory_store.items():
            if doc.get('sha256_checksum') == sha256 and not doc.get('is_deleted'):
                return doc_id
        return None

    def _inmemory_stats(self, customer_id=None):
        docs = [d for d in self._inmemory_store.values() if not d.get('is_deleted')]
        if customer_id:
            docs = [d for d in docs if d.get('customer_id') == customer_id]
        by_cat = {}
        by_status = {}
        total_size = 0
        for d in docs:
            by_cat[d.get('category', 'general')] = by_cat.get(d.get('category', 'general'), 0) + 1
            by_status[d.get('status', 'uploaded')] = by_status.get(d.get('status', 'uploaded'), 0) + 1
            total_size += d.get('file_size', 0)
        return {
            'total_documents': len(docs),
            'total_size_bytes': total_size,
            'by_category': by_cat,
            'by_status': by_status,
        }

    # ── Internal: processing jobs ─────────────────────────────────────────────

    def _run_immediate_processing(self, doc_id: str, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        """Run lightweight processing synchronously on upload."""
        result: Dict[str, Any] = {}
        try:
            result['metadata'] = self._extract_metadata(raw, mime, ext)
        except Exception as e:
            logger.error(f"Metadata extraction failed for {doc_id}: {e}")
            result['metadata'] = {}

        try:
            # Prefer original filename for OCR language hint (e.g. *_he.pdf).
            name_hint = ''
            try:
                rec = self._inmemory_store.get(doc_id) or {}
                name_hint = (
                    rec.get('original_file_name')
                    or rec.get('file_name')
                    or ''
                )
            except Exception:
                name_hint = ''
            if mime.startswith('text/') or ext in ('.csv', '.txt', '.json', '.xml', '.html', '.htm'):
                result['text'] = self._extract_text_content(raw, mime, ext)
            elif mime == 'application/pdf' or ext == '.pdf':
                pdf_text, pdf_pages = self._extract_pdf_text_with_pages(
                    raw, lang_hint=name_hint or None)
                result['text'] = pdf_text
                if pdf_pages:
                    result.setdefault('metadata', {})['pages'] = pdf_pages
            elif ext == '.docx':
                result['text'] = self._extract_docx_text(raw)
            elif ext in ('.xls', '.xlsx'):
                result['text'] = self._extract_spreadsheet_summary(raw, ext)
            elif mime.startswith('audio/'):
                analysis = self._analyze_audio(raw, mime)
                self._merge_media_analysis(result, analysis)
            elif mime.startswith('video/'):
                analysis = self._analyze_video(raw, mime)
                self._merge_media_analysis(result, analysis)
        except Exception as e:
            logger.error(f"Text extraction failed for {doc_id}: {e}")

        try:
            result['summary'] = self._generate_summary(result.get('text', ''), result.get('metadata', {}), mime)
            result['tags'] = self._generate_tags(result.get('text', ''), result.get('metadata', {}), mime, ext)
            result['confidence'] = self._compute_confidence(result)
        except Exception as e:
            logger.error(f"AI enrichment failed for {doc_id}: {e}")

        return result

    @staticmethod
    def _merge_media_analysis(result: Dict[str, Any], analysis: Dict[str, Any]) -> None:
        """Fold audio/video analysis into the enrichment result, building a
        char-offset segment map so mined facts can cite timestamps."""
        text = analysis.get('text') or ''
        if text:
            result['text'] = text
        transcript = analysis.get('transcript')
        if not transcript:
            return
        meta = result.setdefault('metadata', {})
        segments_with_offsets = []
        cursor = 0
        transcript_text = transcript.get('text', '')
        for seg in transcript.get('segments') or []:
            seg_text = seg.get('text', '')
            if not seg_text:
                continue
            idx = transcript_text.find(seg_text, cursor)
            if idx < 0:
                idx = transcript_text.find(seg_text)
            entry = {
                'timestamp_start': seg.get('start'),
                'timestamp_end': seg.get('end'),
                'text': seg_text,
            }
            if idx >= 0:
                entry['char_start'] = idx
                entry['char_end'] = idx + len(seg_text)
                cursor = idx + len(seg_text)
            segments_with_offsets.append(entry)
        meta['transcript'] = {
            'language': transcript.get('language'),
            'provider': transcript.get('provider'),
            'model': transcript.get('model'),
            'duration_seconds': transcript.get('duration_seconds'),
            'segments': segments_with_offsets,
        }

    def _run_job(self, doc_id: str, job_type: str, raw: bytes, mime: str, ext: str) -> ProcessingResult:
        job_id = f"JOB-{uuid.uuid4().hex[:12].upper()}"
        start = time.time()
        result_data = None
        error = None
        status = 'completed'

        try:
            handler = self._job_handlers.get(job_type)
            if handler:
                result_data = handler(self, raw, mime, ext)
            else:
                result_data = {'note': f'No specialised handler for {job_type}'}
        except Exception as e:
            error = str(e)
            status = 'failed'

        elapsed_ms = int((time.time() - start) * 1000)

        if self.db_manager:
            try:
                self.db_manager.processing_jobs.create(
                    id=job_id,
                    document_id=doc_id,
                    job_type=job_type,
                    status=status,
                    result=json.dumps(result_data) if result_data else None,
                    error_message=error,
                    processing_time_ms=elapsed_ms,
                    completed_date=datetime.now(),
                )
            except Exception as e:
                logger.error(f"Failed to persist job {job_id}: {e}")

        return ProcessingResult(
            job_id=job_id,
            document_id=doc_id,
            job_type=job_type,
            status=status,
            result=result_data,
            error=error,
            processing_time_ms=elapsed_ms,
        )

    def _default_jobs_for(self, mime: str, ext: str) -> List[str]:
        jobs = [ProcessingJobType.METADATA_EXTRACTION.value]
        if mime.startswith('text/') or ext in ('.csv', '.txt', '.json', '.xml'):
            jobs.append(ProcessingJobType.TEXT_EXTRACTION.value)
        if ext in ('.csv', '.xls', '.xlsx'):
            jobs.append(ProcessingJobType.TABLE_EXTRACTION.value)
        if mime.startswith('image/') or ext in ('.jpg', '.jpeg', '.png', '.tiff', '.bmp'):
            jobs.append(ProcessingJobType.IMAGE_ANALYSIS.value)
        if mime == 'application/pdf' or ext == '.pdf':
            jobs.append(ProcessingJobType.TEXT_EXTRACTION.value)
        if mime.startswith('audio/') or ext in ('.mp3', '.wav', '.ogg', '.flac'):
            jobs.append(ProcessingJobType.AUDIO_TRANSCRIPTION.value)
        if mime.startswith('video/') or ext in ('.mp4', '.avi', '.mov', '.mkv', '.webm'):
            jobs.append(ProcessingJobType.VIDEO_ANALYSIS.value)
        if ext in ('.dicom', '.dcm'):
            jobs.append(ProcessingJobType.MEDICAL_ANALYSIS.value)
        jobs.append(ProcessingJobType.AI_TAGGING.value)
        return jobs

    # ── Processing handlers ───────────────────────────────────────────────────

    def _handle_metadata_extraction(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._extract_metadata(raw, mime, ext)

    def _handle_text_extraction(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        text = ''
        if mime == 'application/pdf' or ext == '.pdf':
            text = self._extract_pdf_text(raw)
        elif mime.startswith('text/') or ext in ('.csv', '.txt', '.json', '.xml', '.html', '.htm'):
            text = self._extract_text_content(raw, mime, ext)
        elif ext == '.docx':
            text = self._extract_docx_text(raw)
        elif ext in ('.xls', '.xlsx'):
            text = self._extract_spreadsheet_summary(raw, ext)
        return {'text': text, 'length': len(text)}

    def _handle_table_extraction(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._extract_table_data(raw, mime, ext)

    def _handle_image_analysis(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_image(raw, mime)

    def _handle_identity_verification(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_identity_document(raw, mime)

    def _handle_medical_analysis(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_medical_document(raw, mime, ext)

    def _handle_legal_analysis(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_legal_document(raw, mime, ext)

    def _handle_audio_transcription(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_audio(raw, mime)

    def _handle_video_analysis(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return self._analyze_video(raw, mime)

    def _handle_integrity_check(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        return {
            'sha256': hashlib.sha256(raw).hexdigest(),
            'md5': hashlib.md5(raw).hexdigest(),
            'size': len(raw),
            'valid': True,
        }

    def _handle_ai_summarisation(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        metadata = self._extract_metadata(raw, mime, ext)
        text = ''
        if mime.startswith('text/') or ext in ('.csv', '.txt', '.json', '.xml'):
            text = self._extract_text_content(raw, mime, ext)
        summary = self._generate_summary(text, metadata, mime)
        return {'summary': summary}

    def _handle_ai_tagging(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        metadata = self._extract_metadata(raw, mime, ext)
        text = ''
        if mime.startswith('text/') or ext in ('.csv', '.txt', '.json', '.xml'):
            text = self._extract_text_content(raw, mime, ext)
        tags = self._generate_tags(text, metadata, mime, ext)
        return {'tags': tags}

    _job_handlers = {
        ProcessingJobType.METADATA_EXTRACTION.value: _handle_metadata_extraction,
        ProcessingJobType.TEXT_EXTRACTION.value: _handle_text_extraction,
        ProcessingJobType.TABLE_EXTRACTION.value: _handle_table_extraction,
        ProcessingJobType.IMAGE_ANALYSIS.value: _handle_image_analysis,
        ProcessingJobType.IDENTITY_VERIFICATION.value: _handle_identity_verification,
        ProcessingJobType.MEDICAL_ANALYSIS.value: _handle_medical_analysis,
        ProcessingJobType.LEGAL_ANALYSIS.value: _handle_legal_analysis,
        ProcessingJobType.AUDIO_TRANSCRIPTION.value: _handle_audio_transcription,
        ProcessingJobType.VIDEO_ANALYSIS.value: _handle_video_analysis,
        ProcessingJobType.INTEGRITY_CHECK.value: _handle_integrity_check,
        ProcessingJobType.AI_SUMMARISATION.value: _handle_ai_summarisation,
        ProcessingJobType.AI_TAGGING.value: _handle_ai_tagging,
    }

    # ── Extraction helpers ────────────────────────────────────────────────────

    def _extract_metadata(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        meta: Dict[str, Any] = {
            'size_bytes': len(raw),
            'mime_type': mime,
            'extension': ext,
        }

        if mime.startswith('image/'):
            meta.update(self._image_metadata(raw))
        elif mime.startswith('video/'):
            meta['media_type'] = 'video'
            meta['estimated_duration_seconds'] = max(1, len(raw) // (500 * 1024))
        elif mime.startswith('audio/'):
            meta['media_type'] = 'audio'
            meta['estimated_duration_seconds'] = max(1, len(raw) // (16 * 1024))
        elif ext in ('.csv',):
            meta.update(self._csv_metadata(raw))
        elif ext in ('.json',):
            meta.update(self._json_metadata(raw))

        magic_bytes = raw[:16]
        if magic_bytes[:4] == b'%PDF':
            meta['format'] = 'PDF'
        elif magic_bytes[:2] in (b'PK',):
            meta['format'] = 'ZIP/Office'
        elif magic_bytes[:3] == b'ID3' or magic_bytes[:2] == b'\xff\xfb':
            meta['format'] = 'MP3'
        elif magic_bytes[:4] == b'RIFF':
            meta['format'] = 'WAV/AVI'
        elif magic_bytes[:4] == b'\x00\x00\x00\x1c' or magic_bytes[4:8] == b'ftyp':
            meta['format'] = 'MP4/MOV'

        return meta

    def _image_metadata(self, raw: bytes) -> Dict[str, Any]:
        info: Dict[str, Any] = {'media_type': 'image'}
        if raw[:2] == b'\xff\xd8':
            info['format'] = 'JPEG'
            info.update(self._parse_jpeg_dimensions(raw))
        elif raw[:8] == b'\x89PNG\r\n\x1a\n':
            info['format'] = 'PNG'
            if len(raw) >= 24:
                import struct
                w, h = struct.unpack('>II', raw[16:24])
                info['width'] = w
                info['height'] = h
        elif raw[:3] == b'GIF':
            info['format'] = 'GIF'
        elif raw[:4] in (b'II\x2a\x00', b'MM\x00\x2a'):
            info['format'] = 'TIFF'
        return info

    def _parse_jpeg_dimensions(self, raw: bytes) -> Dict[str, Any]:
        try:
            i = 2
            while i < len(raw) - 1:
                if raw[i] != 0xFF:
                    break
                marker = raw[i + 1]
                if marker == 0xD9:
                    break
                if marker in (0xC0, 0xC1, 0xC2):
                    if i + 9 < len(raw):
                        import struct
                        h, w = struct.unpack('>HH', raw[i + 5:i + 9])
                        return {'width': w, 'height': h}
                if i + 3 < len(raw):
                    import struct
                    length = struct.unpack('>H', raw[i + 2:i + 4])[0]
                    i += 2 + length
                else:
                    break
        except Exception:
            pass
        return {}

    def _csv_metadata(self, raw: bytes) -> Dict[str, Any]:
        try:
            text = raw.decode('utf-8', errors='replace')
            lines = text.strip().split('\n')
            headers = lines[0].split(',') if lines else []
            return {
                'format': 'CSV',
                'row_count': len(lines) - 1,
                'column_count': len(headers),
                'headers': [h.strip().strip('"') for h in headers[:50]],
            }
        except Exception:
            return {'format': 'CSV'}

    def _json_metadata(self, raw: bytes) -> Dict[str, Any]:
        try:
            data = json.loads(raw.decode('utf-8', errors='replace'))
            if isinstance(data, list):
                return {'format': 'JSON', 'type': 'array', 'item_count': len(data)}
            elif isinstance(data, dict):
                return {'format': 'JSON', 'type': 'object', 'top_keys': list(data.keys())[:20]}
            return {'format': 'JSON'}
        except Exception:
            return {'format': 'JSON'}

    def _extract_text_content(self, raw: bytes, mime: str, ext: str) -> str:
        try:
            text = raw.decode('utf-8', errors='replace')
            return text[:100_000]
        except Exception:
            return ''

    # Conservative caps so a single huge upload can never hang the worker.
    _OCR_LANGS = os.environ.get('PHINS_OCR_LANGS', 'heb+eng+ara')
    _OCR_DPI = int(os.environ.get('PHINS_OCR_DPI', '200'))
    _OCR_MAX_PDF_PAGES = int(os.environ.get('PHINS_OCR_MAX_PDF_PAGES', '15'))
    _OCR_MAX_IMAGE_BYTES = int(os.environ.get('PHINS_OCR_MAX_IMAGE_BYTES', 25 * 1024 * 1024))
    _OCR_MIN_TEXT_THRESHOLD = int(os.environ.get('PHINS_OCR_MIN_TEXT_THRESHOLD', '40'))

    @staticmethod
    def _has_meaningful_text(text: str) -> bool:
        """Heuristic: if pypdf returns mostly garbage from a scanned PDF
        we want to fall through to OCR. A real text layer normally has
        spaces and a high ratio of letters.

        Hebrew letters count as alphabetic (``str.isalpha``) so IL text-layer
        PDFs are not incorrectly escalated to OCR.
        """
        if not text:
            return False
        sample = text[:5000]
        letters = sum(1 for c in sample if c.isalpha() or ("\u0590" <= c <= "\u05FF"))
        return letters >= max(20, len(sample) // 4)

    @classmethod
    def _ocr_langs_for_hint(cls, hint: Optional[str] = None) -> str:
        """Return Tesseract ``-l`` language string, preferring Hebrew when hinted.

        Default remains ``heb+eng+ara`` (or ``PHINS_OCR_LANGS``). When the
        filename / prior text clearly looks Hebrew-dominant, put ``heb`` first
        so Tesseract's primary script model matches IL medical/insurance scans.
        """
        configured = cls._OCR_LANGS or "heb+eng+ara"
        parts = [p.strip() for p in configured.split("+") if p.strip()]
        if not parts:
            parts = ["heb", "eng", "ara"]
        try:
            from services.hebrew_assessment_lexicon import (
                contains_hebrew,
                hebrew_ratio,
            )
            hinted = bool(hint and (contains_hebrew(hint) or hebrew_ratio(hint) >= 0.2
                                    or any(tok in hint.lower() for tok in (
                                        "hebrew", "ivrit", "עברית", "_he.", "-he.",
                                    ))))
        except Exception:
            hinted = bool(hint and any("\u0590" <= ch <= "\u05FF" for ch in hint))
        if hinted and "heb" in parts:
            parts = ["heb"] + [p for p in parts if p != "heb"]
        return "+".join(parts)

    def _ocr_image_bytes(self, raw: bytes, *, lang_hint: Optional[str] = None) -> str:
        """Run Tesseract OCR on raw image bytes.

        Returns the extracted text, or '' when OCR is unavailable
        (system tesseract / Hebrew language pack missing) or the image
        is too large / corrupt. Languages are configurable via
        PHINS_OCR_LANGS (default ``heb+eng+ara``); Hebrew-hinted inputs
        put ``heb`` first for better IL medical/insurance scans.
        """
        if not raw or len(raw) > self._OCR_MAX_IMAGE_BYTES:
            return ''
        try:
            import pytesseract  # type: ignore
            from PIL import Image  # type: ignore
        except ImportError:
            return ''
        ocr_langs = self._ocr_langs_for_hint(lang_hint)
        try:
            import io as _io
            with Image.open(_io.BytesIO(raw)) as img:
                if img.mode not in ('RGB', 'L'):
                    img = img.convert('RGB')
                text = pytesseract.image_to_string(img, lang=ocr_langs)
            return (text or '').strip()
        except pytesseract.TesseractNotFoundError:
            return ''
        except Exception as exc:
            logger.debug('OCR failed: %s', exc)
            return ''

    def _ocr_pdf_pages(self, raw: bytes, *, lang_hint: Optional[str] = None) -> str:
        """Rasterise each page of a scanned PDF and OCR it (joined text)."""
        return '\n\n'.join(self._ocr_pdf_page_chunks(raw, lang_hint=lang_hint)).strip()

    def _ocr_pdf_page_chunks(self, raw: bytes, *, lang_hint: Optional[str] = None) -> List[str]:
        """Rasterise each page of a scanned PDF and OCR it, one chunk per page.

        Capped at ``PHINS_OCR_MAX_PDF_PAGES`` pages so a 100-page
        scanned document can't blow the request budget. Returns []
        if pdf2image / poppler / tesseract aren't all available.
        Empty pages are kept as '' so chunk index == page number - 1.
        """
        try:
            from pdf2image import convert_from_bytes  # type: ignore
            import pytesseract  # type: ignore
        except ImportError:
            return []
        try:
            pages = convert_from_bytes(raw, dpi=self._OCR_DPI,
                                       first_page=1, last_page=self._OCR_MAX_PDF_PAGES)
        except Exception as exc:
            logger.debug('pdf2image rasterisation failed: %s', exc)
            return []
        ocr_langs = self._ocr_langs_for_hint(lang_hint)
        chunks = []
        for page_img in pages:
            try:
                if page_img.mode not in ('RGB', 'L'):
                    page_img = page_img.convert('RGB')
                page_text = pytesseract.image_to_string(page_img, lang=ocr_langs)
                chunks.append((page_text or '').strip())
            except Exception as exc:
                logger.debug('OCR page failed: %s', exc)
                chunks.append('')
        if not any(chunks):
            return []
        return chunks

    def _extract_pdf_text(self, raw: bytes, *, lang_hint: Optional[str] = None) -> str:
        """Extract text from a PDF (see :meth:`_extract_pdf_text_with_pages`)."""
        text, _pages = self._extract_pdf_text_with_pages(raw, lang_hint=lang_hint)
        return text

    def _extract_pdf_text_with_pages(
        self, raw: bytes, *, lang_hint: Optional[str] = None,
    ) -> Tuple[str, List[Dict[str, int]]]:
        """Extract text from a PDF, escalating from cheapest to most powerful.

        Order:
          1. **pypdf** for proper text-layer PDFs (cheap, fast).
          2. **Regex fallback** over the raw bytes (cheap, last resort
             for tiny PDFs without proper text objects).
          3. **OCR via pdf2image + tesseract** for scanned PDFs (slow
             but extracts text from images embedded inside the PDF).

        Returns ``(text, pages)`` where ``pages`` maps 1-based page numbers to
        character offsets in the returned text
        (``[{'page': 1, 'char_start': 0, 'char_end': 1234}, ...]``) so
        downstream evidence extraction can cite the exact source page.
        ``pages`` is empty when per-page boundaries are unknown (regex path).
        The text is an explanatory marker when nothing could be extracted;
        the assessment center detects that and surfaces an extraction_hint
        fact so the user knows OCR is needed.
        """
        text = ''
        page_offsets: List[Dict[str, int]] = []

        def _join_with_offsets(chunks: List[str], separator: str) -> Tuple[str, List[Dict[str, int]]]:
            joined_parts: List[str] = []
            offsets: List[Dict[str, int]] = []
            cursor = 0
            for index, chunk in enumerate(chunks):
                if index:
                    cursor += len(separator)
                offsets.append({
                    'page': index + 1,
                    'char_start': cursor,
                    'char_end': cursor + len(chunk),
                })
                joined_parts.append(chunk)
                cursor += len(chunk)
            return separator.join(joined_parts), offsets

        try:
            from pypdf import PdfReader  # type: ignore
            from pypdf.errors import PdfReadError  # type: ignore
            import io as _io
            try:
                reader = PdfReader(_io.BytesIO(raw))
                if getattr(reader, 'is_encrypted', False):
                    try:
                        reader.decrypt('')
                    except Exception:
                        pass
                pages_text = []
                for page in reader.pages[:50]:
                    try:
                        page_text = page.extract_text() or ''
                    except Exception:
                        page_text = ''
                    # Keep empty pages in the list so page numbers stay true.
                    pages_text.append(page_text)
                text, page_offsets = _join_with_offsets(pages_text, '\n')
                text = text.strip() and text or ''
                if not text:
                    page_offsets = []
            except PdfReadError:
                text = ''
            except Exception as exc:
                logger.debug('pypdf extraction failed, falling back to regex: %s', exc)
                text = ''
        except ImportError:
            text = ''

        if not self._has_meaningful_text(text):
            try:
                latin = raw.decode('latin-1', errors='replace')
                parts = []
                for match in re.finditer(r'\(([^)]{1,500})\)', latin):
                    chunk = match.group(1)
                    if any(c.isalpha() for c in chunk):
                        parts.append(chunk)
                regex_text = ' '.join(parts)
                if self._has_meaningful_text(regex_text):
                    text = regex_text
                    page_offsets = []
            except Exception:
                pass

        if not self._has_meaningful_text(text):
            # Prefer any partial Hebrew from the weak text layer / filename so
            # Tesseract loads heb as the primary script for IL scans.
            ocr_hint = lang_hint or text or None
            ocr_chunks = self._ocr_pdf_page_chunks(raw, lang_hint=ocr_hint)
            if ocr_chunks:
                text, page_offsets = _join_with_offsets(ocr_chunks, '\n\n')

        if not text:
            return '[PDF content - extraction yielded no text; image-only or encrypted]', []
        if len(text) > 200_000:
            text = text[:200_000]
            page_offsets = [
                {**p, 'char_end': min(p['char_end'], len(text))}
                for p in page_offsets if p['char_start'] < len(text)
            ]
        return text, page_offsets

    def _extract_docx_text(self, raw: bytes) -> str:
        """Extract paragraph text from a DOCX (OOXML) file.

        Stdlib-only: unzip ``word/document.xml`` and walk the WordprocessingML
        text nodes (``w:t``), inserting newlines at paragraph boundaries and
        tabs/breaks where declared. Table cell text is included in document
        order. Uses defusedxml when available to guard against XML bombs in
        untrusted uploads. Returns '' when the file is not a valid DOCX.
        """
        import io as _io
        import zipfile as _zip
        try:
            from defusedxml import ElementTree as _ET  # type: ignore
        except ImportError:
            import xml.etree.ElementTree as _ET  # type: ignore
        try:
            with _zip.ZipFile(_io.BytesIO(raw)) as zf:
                with zf.open('word/document.xml') as fh:
                    xml_data = fh.read()
        except Exception as exc:
            logger.debug('DOCX unzip failed: %s', exc)
            return ''
        ns = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
        try:
            root = _ET.fromstring(xml_data)
        except Exception as exc:
            logger.debug('DOCX XML parse failed: %s', exc)
            return ''
        paragraphs: List[str] = []
        for para in root.iter(f'{ns}p'):
            parts: List[str] = []
            for node in para.iter():
                if node.tag == f'{ns}t' and node.text:
                    parts.append(node.text)
                elif node.tag == f'{ns}tab':
                    parts.append('\t')
                elif node.tag == f'{ns}br':
                    parts.append('\n')
            line = ''.join(parts)
            if line.strip():
                paragraphs.append(line)
        return '\n'.join(paragraphs)[:200_000]

    def _extract_zip_contents(self, raw: bytes) -> str:
        """Concatenate text from every supported file inside a ZIP.

        Limits: at most ``PHINS_ZIP_MAX_FILES`` entries, no recursion
        into nested zips, individual entries capped at 25MB. Each
        entry is dispatched back through the standard text-extraction
        helpers so PDFs and Excel files inside a zip get the same
        treatment as if they were uploaded directly.
        """
        try:
            import io as _io
            import zipfile as _zip
        except ImportError:
            return ''
        max_files = int(os.environ.get('PHINS_ZIP_MAX_FILES', '25'))
        max_entry_bytes = int(os.environ.get('PHINS_ZIP_MAX_ENTRY_BYTES', 25 * 1024 * 1024))
        chunks = []
        try:
            with _zip.ZipFile(_io.BytesIO(raw)) as zf:
                for info in zf.infolist()[:max_files]:
                    if info.is_dir() or info.file_size > max_entry_bytes:
                        continue
                    name = info.filename or ''
                    ext = ''
                    if '.' in name:
                        ext = '.' + name.rsplit('.', 1)[-1].lower()
                    if ext not in ALLOWED_EXTENSIONS or ext == '.zip':
                        continue
                    try:
                        entry_raw = zf.read(info)
                    except Exception as exc:
                        logger.debug('ZIP entry read failed (%s): %s', name, exc)
                        continue
                    entry_text = self._extract_text_for_file(name, ext, entry_raw)
                    if entry_text:
                        chunks.append(f'--- {name} ---\n{entry_text}')
        except _zip.BadZipFile:
            return ''
        except Exception as exc:
            logger.debug('ZIP traversal failed: %s', exc)
            return ''
        return ('\n\n'.join(chunks))[:400_000]

    def _extract_text_for_file(self, name: str, ext: str, raw: bytes) -> str:
        """Best-effort text extraction dispatcher used by the ZIP walker."""
        mime = mimetypes.guess_type(name)[0] or ''
        if ext == '.pdf' or mime == 'application/pdf':
            return self._extract_pdf_text(raw, lang_hint=name)
        if ext == '.docx':
            return self._extract_docx_text(raw)
        if ext in ('.xls', '.xlsx'):
            return self._extract_spreadsheet_summary(raw, ext)
        if ext in ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif'):
            return self._ocr_image_bytes(raw, lang_hint=name)
        return self._extract_text_content(raw, mime, ext)

    def _extract_spreadsheet_summary(self, raw: bytes, ext: str) -> str:
        """Pull every cell value across every sheet so insurance /
        savings / medical text inside an Excel file is mineable.

        openpyxl handles .xlsx; xlrd is the fallback for legacy .xls.
        Both are bounded by row / sheet caps so a 1M-row sheet can't
        blow the budget.
        """
        max_rows = int(os.environ.get('PHINS_XLSX_MAX_ROWS_PER_SHEET', '5000'))
        max_sheets = int(os.environ.get('PHINS_XLSX_MAX_SHEETS', '10'))
        chunks: List[str] = []
        if ext == '.xlsx':
            try:
                from openpyxl import load_workbook  # type: ignore
                import io as _io
                wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
                for sheet_name in wb.sheetnames[:max_sheets]:
                    ws = wb[sheet_name]
                    chunks.append(f'### {sheet_name}')
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= max_rows:
                            break
                        cells = [str(c) for c in row if c is not None and str(c).strip()]
                        if cells:
                            chunks.append(' | '.join(cells))
            except Exception as exc:
                logger.debug('openpyxl extraction failed: %s', exc)
        elif ext == '.xls':
            try:
                import xlrd  # type: ignore
                book = xlrd.open_workbook(file_contents=raw)
                for sheet_index in range(min(book.nsheets, max_sheets)):
                    sheet = book.sheet_by_index(sheet_index)
                    chunks.append(f'### {sheet.name}')
                    for r in range(min(sheet.nrows, max_rows)):
                        cells = [str(sheet.cell_value(r, c))
                                 for c in range(sheet.ncols)
                                 if str(sheet.cell_value(r, c)).strip()]
                        if cells:
                            chunks.append(' | '.join(cells))
            except Exception as exc:
                logger.debug('xlrd extraction failed: %s', exc)

        if not chunks:
            return f'[Spreadsheet content ({ext}) - {len(raw)} bytes; could not parse]'
        text = '\n'.join(chunks)
        return text[:200_000]

    def _extract_table_data(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        if ext == '.csv' or mime == 'text/csv':
            return self._parse_csv_table(raw)
        return {'format': ext, 'size': len(raw), 'note': 'Binary table format - specialised parser needed'}

    def _parse_csv_table(self, raw: bytes) -> Dict[str, Any]:
        try:
            import csv as csv_mod
            import io
            text = raw.decode('utf-8', errors='replace')
            reader = csv_mod.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return {'headers': [], 'rows': [], 'row_count': 0}
            headers = rows[0]
            data_rows = rows[1:]
            preview = data_rows[:20]
            numeric_cols = []
            for col_idx, hdr in enumerate(headers):
                nums = []
                for row in data_rows:
                    if col_idx < len(row):
                        try:
                            nums.append(float(row[col_idx].replace(',', '')))
                        except (ValueError, TypeError):
                            pass
                if nums:
                    numeric_cols.append({
                        'column': hdr,
                        'min': min(nums),
                        'max': max(nums),
                        'mean': sum(nums) / len(nums),
                        'count': len(nums),
                    })
            return {
                'headers': headers,
                'row_count': len(data_rows),
                'column_count': len(headers),
                'preview_rows': preview,
                'numeric_analysis': numeric_cols,
            }
        except Exception as e:
            return {'error': str(e)}

    # ── Specialised analysis ──────────────────────────────────────────────────

    def _analyze_image(self, raw: bytes, mime: str) -> Dict[str, Any]:
        meta = self._image_metadata(raw)
        result = {
            'type': 'image_analysis',
            'format': meta.get('format', 'unknown'),
            'dimensions': {
                'width': meta.get('width'),
                'height': meta.get('height'),
            },
            'file_size': len(raw),
        }
        aspect = None
        if meta.get('width') and meta.get('height'):
            aspect = round(meta['width'] / meta['height'], 3)
            result['aspect_ratio'] = aspect
            is_document = 0.6 < aspect < 0.85 or 1.2 < aspect < 1.6
            result['likely_document'] = is_document
            is_portrait = 0.6 < aspect < 0.9
            result['likely_portrait'] = is_portrait
        return result

    def _analyze_identity_document(self, raw: bytes, mime: str) -> Dict[str, Any]:
        analysis = self._analyze_image(raw, mime)
        analysis['type'] = 'identity_verification'
        analysis['checks'] = {
            'file_integrity': True,
            'minimum_resolution': (analysis.get('dimensions', {}).get('width', 0) or 0) >= 300,
            'format_acceptable': analysis.get('format', '') in ('JPEG', 'PNG', 'TIFF'),
        }
        analysis['verification_status'] = 'pending_review'
        return analysis

    def _analyze_medical_document(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            'type': 'medical_analysis',
            'format': ext,
            'file_size': len(raw),
        }
        if ext in ('.dicom', '.dcm'):
            result['dicom_detected'] = True
            result['note'] = 'DICOM medical imaging format detected'
        elif ext == '.pdf':
            text = self._extract_pdf_text(raw)
            medical_terms = ['diagnosis', 'patient', 'treatment', 'prescription',
                             'medication', 'symptom', 'clinical', 'medical', 'hospital',
                             'doctor', 'physician', 'surgery']
            found_terms = [t for t in medical_terms if t.lower() in text.lower()]
            result['detected_medical_terms'] = found_terms
            result['medical_relevance_score'] = min(1.0, len(found_terms) / 5)
        return result

    def _analyze_legal_document(self, raw: bytes, mime: str, ext: str) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            'type': 'legal_analysis',
            'format': ext,
            'file_size': len(raw),
        }
        text = ''
        if ext == '.pdf':
            text = self._extract_pdf_text(raw)
        elif mime.startswith('text/'):
            text = self._extract_text_content(raw, mime, ext)
        if text:
            legal_terms = ['contract', 'agreement', 'clause', 'liability',
                           'indemnity', 'warranty', 'insurance', 'policy',
                           'coverage', 'premium', 'deductible', 'beneficiary',
                           'claimant', 'arbitration', 'jurisdiction', 'statute']
            found = [t for t in legal_terms if t.lower() in text.lower()]
            result['detected_legal_terms'] = found
            result['legal_relevance_score'] = min(1.0, len(found) / 5)
        return result

    def _analyze_audio(self, raw: bytes, mime: str) -> Dict[str, Any]:
        """Transcribe audio through the configured provider.

        Falls back to the informational stub when no provider is configured
        (``PHINS_TRANSCRIPTION_PROVIDER=disabled``, the default) so existing
        deployments keep their current behaviour until an operator opts in.
        """
        base = {
            'type': 'audio_analysis',
            'file_size': len(raw),
            'estimated_duration_seconds': max(1, len(raw) // (16 * 1024)),
        }
        transcript = self._transcribe_media(raw, mime, default_ext='.mp3')
        if transcript is None:
            base['note'] = 'Audio transcription requires external ASR service integration'
            return base
        base['transcript'] = transcript
        base['text'] = transcript.get('text', '')
        base['language'] = transcript.get('language')
        if transcript.get('duration_seconds'):
            base['estimated_duration_seconds'] = transcript['duration_seconds']
        return base

    def _analyze_video(self, raw: bytes, mime: str) -> Dict[str, Any]:
        """Extract intelligence from video without sending raw video to LLMs.

        Cost-controlled pipeline (spec §10): the audio track is extracted with
        ffmpeg and transcribed; sampled keyframes are OCR'd for visible text
        (documents shown on screen, screen recordings). Degrades gracefully to
        the informational stub when ffmpeg or the transcription provider is
        unavailable.
        """
        base = {
            'type': 'video_analysis',
            'file_size': len(raw),
            'estimated_duration_seconds': max(1, len(raw) // (500 * 1024)),
        }
        audio_bytes = self._ffmpeg_extract_audio(raw)
        transcript = None
        if audio_bytes:
            transcript = self._transcribe_media(audio_bytes, 'audio/mpeg',
                                                default_ext='.mp3')
        if transcript is not None:
            base['transcript'] = transcript
            base['language'] = transcript.get('language')
            if transcript.get('duration_seconds'):
                base['estimated_duration_seconds'] = transcript['duration_seconds']

        keyframe_text = self._ffmpeg_keyframe_ocr(raw)
        if keyframe_text:
            base['visible_text'] = keyframe_text

        combined = '\n'.join(filter(None, [
            (transcript or {}).get('text', ''),
            keyframe_text,
        ])).strip()
        if combined:
            base['text'] = combined
        else:
            base['note'] = ('Video analysis limited: install ffmpeg and '
                            'configure PHINS_TRANSCRIPTION_PROVIDER for '
                            'spoken/visible-text extraction')
        return base

    def _transcribe_media(self, raw: bytes, mime: str, *,
                          default_ext: str = '.mp3') -> Optional[Dict[str, Any]]:
        """Best-effort transcription; None when no provider is configured or
        the provider call fails (callers keep their stub behaviour)."""
        try:
            from services.transcription_providers import (
                TranscriptionUnavailableError,
                get_transcription_provider,
            )
        except ImportError:
            return None
        try:
            return get_transcription_provider().transcribe(
                raw, file_name=f'media{default_ext}', mime_type=mime)
        except TranscriptionUnavailableError:
            return None
        except Exception as exc:
            logger.warning('Transcription failed (non-fatal): %s', exc)
            return None

    # Video helpers: ffmpeg is optional; every path degrades to ''/None.

    _VIDEO_MAX_BYTES = int(os.environ.get('PHINS_VIDEO_MAX_ANALYSIS_BYTES',
                                          200 * 1024 * 1024))
    _VIDEO_KEYFRAME_COUNT = int(os.environ.get('PHINS_VIDEO_KEYFRAMES', '6'))

    @staticmethod
    def _ffmpeg_available() -> bool:
        return shutil.which('ffmpeg') is not None

    def _ffmpeg_extract_audio(self, raw: bytes) -> Optional[bytes]:
        """Extract the audio track as MP3 via ffmpeg; None when unavailable."""
        if not raw or len(raw) > self._VIDEO_MAX_BYTES or not self._ffmpeg_available():
            return None
        import subprocess
        src = dst = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.video', delete=False) as f:
                f.write(raw)
                src = f.name
            dst = src + '.mp3'
            result = subprocess.run(
                ['ffmpeg', '-y', '-i', src, '-vn', '-acodec', 'libmp3lame',
                 '-b:a', '64k', dst],
                capture_output=True, timeout=300,
            )
            if result.returncode != 0 or not os.path.exists(dst):
                return None
            with open(dst, 'rb') as f:
                return f.read()
        except Exception as exc:
            logger.debug('ffmpeg audio extraction failed: %s', exc)
            return None
        finally:
            for path in (src, dst):
                if path and os.path.exists(path):
                    try:
                        os.remove(path)
                    except OSError:
                        pass

    def _ffmpeg_keyframe_ocr(self, raw: bytes) -> str:
        """Sample keyframes and OCR them for visible text; '' when unavailable."""
        if not raw or len(raw) > self._VIDEO_MAX_BYTES or not self._ffmpeg_available():
            return ''
        import subprocess
        src = None
        frames_dir = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.video', delete=False) as f:
                f.write(raw)
                src = f.name
            frames_dir = tempfile.mkdtemp(prefix='phins_frames_')
            result = subprocess.run(
                ['ffmpeg', '-y', '-i', src, '-vf',
                 f"select='eq(pict_type\\,I)',scale=1280:-1",
                 '-vsync', 'vfr', '-frames:v', str(self._VIDEO_KEYFRAME_COUNT),
                 os.path.join(frames_dir, 'frame_%03d.png')],
                capture_output=True, timeout=300,
            )
            if result.returncode != 0:
                return ''
            chunks = []
            for name in sorted(os.listdir(frames_dir)):
                frame_path = os.path.join(frames_dir, name)
                try:
                    with open(frame_path, 'rb') as f:
                        frame_text = self._ocr_image_bytes(f.read())
                    if frame_text:
                        chunks.append(frame_text)
                except Exception:
                    continue
            return '\n'.join(dict.fromkeys(chunks))[:100_000]
        except Exception as exc:
            logger.debug('ffmpeg keyframe OCR failed: %s', exc)
            return ''
        finally:
            if src and os.path.exists(src):
                try:
                    os.remove(src)
                except OSError:
                    pass
            if frames_dir and os.path.isdir(frames_dir):
                shutil.rmtree(frames_dir, ignore_errors=True)

    # ── AI enrichment ─────────────────────────────────────────────────────────

    def _generate_summary(self, text: str, metadata: Dict[str, Any], mime: str) -> str:
        parts = []
        fmt = metadata.get('format', mime.split('/')[-1] if '/' in mime else 'file')
        parts.append(f"{fmt} document")
        size = metadata.get('size_bytes', 0)
        if size:
            if size > 1024 * 1024:
                parts.append(f"({size / (1024*1024):.1f} MB)")
            else:
                parts.append(f"({size / 1024:.1f} KB)")
        if metadata.get('row_count'):
            parts.append(f"with {metadata['row_count']} rows and {metadata.get('column_count', '?')} columns")
        if metadata.get('width') and metadata.get('height'):
            parts.append(f"dimensions {metadata['width']}x{metadata['height']}")
        if text and len(text) > 20:
            preview = text[:200].replace('\n', ' ').strip()
            parts.append(f"- content preview: {preview}")
        return ' '.join(parts)

    def _generate_tags(self, text: str, metadata: Dict[str, Any],
                       mime: str, ext: str) -> List[str]:
        tags = []
        if mime.startswith('image/'):
            tags.append('image')
        elif mime.startswith('video/'):
            tags.append('video')
        elif mime.startswith('audio/'):
            tags.append('audio')
        elif mime == 'application/pdf':
            tags.append('pdf')
        elif ext in ('.csv', '.xls', '.xlsx'):
            tags.append('spreadsheet')
            tags.append('data')
        elif ext in ('.doc', '.docx'):
            tags.append('document')
        if ext in ('.json', '.xml'):
            tags.append('structured-data')
        if metadata.get('format'):
            tags.append(metadata['format'].lower())
        if text:
            text_lower = text.lower()
            domain_keywords = {
                'insurance': [
                    'policy', 'premium', 'coverage', 'claim', 'insured', 'underwriting',
                    # Hebrew insurance terms (Assessment Center / IL market).
                    'פוליסה', 'פרמיה', 'ביטוח', 'כיסוי', 'תביעה', 'מבוטח', 'מוטב',
                ],
                'medical': [
                    'patient', 'diagnosis', 'treatment', 'medication', 'clinical',
                    'סוכרת', 'יתר לחץ דם', 'אבחנה', 'תרופה', 'מטופל', 'רפואי',
                ],
                'legal': ['contract', 'agreement', 'liability', 'clause', 'jurisdiction',
                          'חוזה', 'הסכם'],
                'financial': [
                    'balance', 'transaction', 'payment', 'invoice', 'revenue',
                    'יתרה', 'הפקדה', 'צבירה', 'פנסיה', 'גמל',
                ],
                'identity': [
                    'passport', 'license', 'id card', 'identity', 'verification',
                    'תעודת זהות', 'ת.ז', 'דרכון',
                ],
                'hebrew': [],  # filled below when Hebrew script is detected
            }
            for domain, keywords in domain_keywords.items():
                if domain == 'hebrew':
                    continue
                if any(kw in text_lower or kw in text for kw in keywords):
                    tags.append(domain)
            # Language tag so dashboards can filter Hebrew-sourced artefacts.
            try:
                from services.hebrew_assessment_lexicon import contains_hebrew
                if contains_hebrew(text):
                    tags.append('hebrew')
            except ImportError:
                if any('\u0590' <= ch <= '\u05FF' for ch in text):
                    tags.append('hebrew')
        return list(dict.fromkeys(tags))

    def _compute_confidence(self, result: Dict[str, Any]) -> float:
        score = 0.5
        if result.get('metadata'):
            score += 0.1
        if result.get('text') and len(result['text']) > 10:
            score += 0.15
        if result.get('summary'):
            score += 0.1
        if result.get('tags') and len(result['tags']) > 1:
            score += 0.1
        return min(1.0, round(score, 2))

    # ── Helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _get_extension(filename: str) -> str:
        if '.' in filename:
            return '.' + filename.rsplit('.', 1)[-1].lower()
        return ''

    @staticmethod
    def _safe_filename(doc_id: str, original: str) -> str:
        ext = ''
        if '.' in original:
            ext = '.' + original.rsplit('.', 1)[-1].lower()
        safe_base = re.sub(r'[^a-zA-Z0-9_-]', '_', original.rsplit('.', 1)[0])[:80]
        return f"{doc_id}_{safe_base}{ext}"

    @staticmethod
    def _classify_category(filename: str, mime: str) -> str:
        ext = ''
        if '.' in filename:
            ext = '.' + filename.rsplit('.', 1)[-1].lower()
        if ext in CATEGORY_BY_EXTENSION:
            return CATEGORY_BY_EXTENSION[ext]
        for prefix, cat in CATEGORY_BY_MIME_PREFIX.items():
            if mime.startswith(prefix) or mime == prefix:
                return cat
        return 'general'


# ── Module-level singleton ────────────────────────────────────────────────────

_default_service: Optional[DocumentProcessingService] = None


def get_document_service(db_manager=None) -> DocumentProcessingService:
    """Return or create the module-level service singleton."""
    global _default_service
    if _default_service is None:
        _default_service = DocumentProcessingService(db_manager=db_manager)
    elif db_manager and _default_service.db_manager is None:
        _default_service.db_manager = db_manager
    return _default_service


def reset_document_service() -> None:
    """Reset the singleton (mainly for tests)."""
    global _default_service
    _default_service = None
