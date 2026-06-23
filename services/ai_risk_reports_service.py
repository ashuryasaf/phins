"""
AI Risk & Reports Analysis Service
===================================
Provides AI-powered analysis of uploaded documents (CSV/XLS/ZIP) to generate
intelligent reports with recommendations for insurance, investment, and risk assessment.

Features:
- Multi-language support (including Hebrew, Arabic, etc.)
- Auto-detection of data types (insurance, investment, risk, savings)
- Pattern recognition and anomaly detection
- Risk scoring and factor extraction
- Automated report generation with charts
- Personalized recommendations

Author: PHINS Platform
"""

import csv
import copy
import io
import json
import logging
import os
import re
import hashlib
import zipfile
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
from enum import Enum
from dataclasses import dataclass, field, asdict
import random
import base64

logger = logging.getLogger('phins.ai_risk_reports')


def _risk_report_audit(action: str, entity_id: Optional[str], details: Dict[str, Any]) -> None:
    """Mirror a risk-report lifecycle event into the durable audit store.

    Best-effort and non-fatal. The report artifacts themselves persist in the
    service's JSON store; this adds a durable compliance trail of who/what was
    produced, independent of that store. No-op without a database.
    """
    try:
        from services.ai_audit_bridge import record_ai_audit
        record_ai_audit(
            action=action,
            entity_type='risk_report',
            entity_id=entity_id,
            details=details,
            username='ai_risk_reports',
        )
    except Exception as exc:
        logger.warning("risk report audit mirror failed (non-fatal): %s", exc)


class DataType(Enum):
    INSURANCE = "insurance"
    INVESTMENT = "investment"
    RISK = "risk"
    SAVINGS = "savings"
    MIXED = "mixed"
    UNKNOWN = "unknown"


class ChartType(Enum):
    PIE = "pie"
    BAR = "bar"
    LINE = "line"
    GAUGE = "gauge"
    SCATTER = "scatter"
    DOUGHNUT = "doughnut"


class Priority(Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    URGENT = "urgent"


class Severity(Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"


@dataclass
class ColumnInfo:
    name: str
    data_type: str
    sample_values: List[Any]
    semantic_type: str = ""
    is_key_field: bool = False


@dataclass
class Factor:
    name: str
    value: Any
    importance: float
    category: str


@dataclass
class Pattern:
    type: str
    description: str
    affected_rows: List[int]
    significance: float


@dataclass
class Anomaly:
    type: str
    severity: Severity
    description: str
    affected_data: Dict
    recommendation: str


@dataclass
class ChartConfig:
    type: ChartType
    title: str
    data: Dict
    options: Dict = field(default_factory=dict)


@dataclass
class ReportSection:
    title: str
    content: str
    data_table: Optional[Dict] = None
    order: int = 0


@dataclass
class Recommendation:
    id: str
    category: str
    priority: Priority
    title: str
    description: str
    action_items: List[str]
    expected_impact: str


@dataclass
class AnalysisResult:
    id: str
    document_id: str
    language: str
    language_name: str
    data_classification: DataType
    extracted_factors: List[Factor]
    patterns_found: List[Pattern]
    anomalies: List[Anomaly]
    risk_score: float
    confidence: float
    processing_time_ms: int
    summary: str
    key_metrics: Dict[str, Any]


@dataclass
class GeneratedReport:
    id: str
    analysis_id: str
    report_type: str
    language: str
    title: str
    sections: List[ReportSection]
    charts: List[ChartConfig]
    recommendations: List[Recommendation]
    generated_at: str
    metadata: Dict[str, Any]


class LanguageDetector:
    """Detects language from text content"""
    
    # Language patterns - character ranges and common words
    LANGUAGE_PATTERNS = {
        'hebrew': {
            'chars': r'[\u0590-\u05FF]',  # Hebrew Unicode range
            'words': ['של', 'את', 'על', 'עם', 'לא', 'זה', 'או', 'כי', 'אם', 'גם'],
            'name': 'עברית (Hebrew)'
        },
        'arabic': {
            'chars': r'[\u0600-\u06FF]',  # Arabic Unicode range
            'words': ['من', 'في', 'على', 'إلى', 'أن', 'هذا', 'التي', 'مع'],
            'name': 'العربية (Arabic)'
        },
        'english': {
            'chars': r'[a-zA-Z]',
            'words': ['the', 'is', 'and', 'of', 'to', 'in', 'for', 'with', 'that', 'this'],
            'name': 'English'
        },
        'spanish': {
            'chars': r'[a-zA-ZáéíóúñüÁÉÍÓÚÑÜ]',
            'words': ['el', 'la', 'de', 'que', 'en', 'los', 'del', 'las', 'por', 'con'],
            'name': 'Español (Spanish)'
        },
        'french': {
            'chars': r'[a-zA-ZàâäéèêëïîôùûüÿçœæÀÂÄÉÈÊËÏÎÔÙÛÜŸÇŒÆ]',
            'words': ['le', 'la', 'de', 'et', 'est', 'en', 'que', 'les', 'des', 'du'],
            'name': 'Français (French)'
        },
        'german': {
            'chars': r'[a-zA-ZäöüßÄÖÜ]',
            'words': ['der', 'die', 'und', 'ist', 'von', 'den', 'das', 'mit', 'für', 'auf'],
            'name': 'Deutsch (German)'
        },
        'russian': {
            'chars': r'[\u0400-\u04FF]',  # Cyrillic Unicode range
            'words': ['и', 'в', 'на', 'не', 'что', 'он', 'как', 'это', 'по', 'но'],
            'name': 'Русский (Russian)'
        },
        'chinese': {
            'chars': r'[\u4e00-\u9fff]',  # CJK Unified Ideographs
            'words': [],  # Chinese doesn't use word boundaries the same way
            'name': '中文 (Chinese)'
        },
        'japanese': {
            'chars': r'[\u3040-\u309F\u30A0-\u30FF]',  # Hiragana and Katakana
            'words': [],
            'name': '日本語 (Japanese)'
        }
    }
    
    @classmethod
    def detect(cls, text: str) -> Tuple[str, str, float]:
        """
        Detect the primary language of text.
        Returns: (language_code, language_name, confidence)
        """
        if not text or len(text.strip()) < 5:
            return 'english', 'English', 0.5
        
        scores = {}
        text_lower = text.lower()
        
        for lang, patterns in cls.LANGUAGE_PATTERNS.items():
            score = 0
            
            # Check character patterns
            char_matches = len(re.findall(patterns['chars'], text))
            char_ratio = char_matches / max(len(text), 1)
            score += char_ratio * 60
            
            # Check common words
            if patterns['words']:
                word_matches = sum(1 for word in patterns['words'] if word in text_lower)
                word_score = (word_matches / len(patterns['words'])) * 40
                score += word_score
            
            scores[lang] = score
        
        # Get the highest scoring language
        best_lang = max(scores, key=scores.get)
        confidence = min(scores[best_lang] / 100, 1.0)
        
        # Default to English if confidence is too low
        if confidence < 0.2:
            return 'english', 'English', 0.5
        
        return best_lang, cls.LANGUAGE_PATTERNS[best_lang]['name'], confidence


class HebrewDocumentExtractor:
    """
    Extracts structured data from Hebrew insurance/financial documents.
    Uses pattern recognition to identify key fields.
    """
    
    # Hebrew field patterns for insurance documents
    FIELD_PATTERNS = {
        'policy_number': [
            r'מספר פוליס[הא][\s:]*([0-9\-/]+)',
            r'פוליס[הא]\s*מס[פ\'][\s:]*([0-9\-/]+)',
            r'policy[\s#:]*([0-9\-/]+)',
        ],
        'id_number': [
            r'ת\.?ז\.?[\s:]*([0-9]{9})',
            r'תעודת זהות[\s:]*([0-9]{9})',
            r'מספר זהות[\s:]*([0-9]{9})',
            r'ת"ז[\s:]*([0-9]{9})',
        ],
        'start_date': [
            r'תאריך תחילה[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'תחילת ביטוח[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'מתאריך[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'start date[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
        ],
        'end_date': [
            r'תאריך סיום[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'תום תקופה[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'עד תאריך[\s:]*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
        ],
        'premium': [
            r'פרמי[הא][\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'תשלום חודשי[\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'premium[\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'סכום לתשלום[\s:]*[₪$]?[\s]*([0-9,\.]+)',
        ],
        'cover_amount': [
            r'סכום ביטוח[\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'סכום כיסוי[\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'cover[\s:]*[₪$]?[\s]*([0-9,\.]+)',
            r'סכום מבוטח[\s:]*[₪$]?[\s]*([0-9,\.]+)',
        ],
        'insured_name': [
            r'שם המבוטח[\s:]*([א-ת\s]+)',
            r'שם מלא[\s:]*([א-ת\s]+)',
            r'מבוטח[\s:]*([א-ת\s]+)',
        ],
        'pension_type': [
            r'סוג פנסיה[\s:]*([א-ת\s]+)',
            r'תוכנית פנסיה[\s:]*([א-ת\s]+)',
            r'קרן פנסיה[\s:]*([א-ת\s]+)',
        ],
        'insurance_type': [
            r'סוג ביטוח[\s:]*([א-ת\s]+)',
            r'סוג פוליסה[\s:]*([א-ת\s]+)',
            r'סוג הכיסוי[\s:]*([א-ת\s]+)',
        ],
        'beneficiary': [
            r'מוטב[\s:]*([א-ת\s]+)',
            r'מוטבים[\s:]*([א-ת\s]+)',
            r'שם מוטב[\s:]*([א-ת\s]+)',
        ],
    }
    
    # Insurance product types (Hebrew)
    PRODUCT_TYPES = {
        'life': ['ביטוח חיים', 'חיים', 'life'],
        'health': ['ביטוח בריאות', 'בריאות', 'health'],
        'pension': ['פנסיה', 'גמל', 'pension', 'קרן פנסיה'],
        'car': ['ביטוח רכב', 'רכב', 'חובה', 'מקיף'],
        'home': ['ביטוח דירה', 'דירה', 'מבנה', 'תכולה'],
        'travel': ['ביטוח נסיעות', 'נסיעות לחו"ל'],
        'business': ['ביטוח עסק', 'עסקי', 'אחריות מקצועית'],
    }
    
    @classmethod
    def extract_fields(cls, text: str) -> Dict[str, Any]:
        """Extract structured fields from Hebrew document text."""
        extracted = {}
        
        for field_name, patterns in cls.FIELD_PATTERNS.items():
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    value = match.group(1).strip()
                    # Clean up numeric values
                    if field_name in ['premium', 'cover_amount']:
                        value = value.replace(',', '')
                        try:
                            value = float(value)
                        except ValueError:
                            pass
                    extracted[field_name] = value
                    break
        
        # Detect product type
        text_lower = text.lower()
        for product_type, keywords in cls.PRODUCT_TYPES.items():
            if any(kw in text_lower for kw in keywords):
                extracted['product_type'] = product_type
                break
        
        return extracted
    
    @classmethod
    def analyze_policy_age(cls, start_date_str: str) -> Dict[str, Any]:
        """Analyze policy age and status."""
        try:
            # Try various date formats
            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d/%m/%y', '%d-%m-%y']:
                try:
                    start_date = datetime.strptime(start_date_str, fmt)
                    break
                except ValueError:
                    continue
            else:
                return {}
            
            today = datetime.now()
            age_days = (today - start_date).days
            age_years = age_days / 365.25
            
            return {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'policy_age_years': round(age_years, 1),
                'policy_age_days': age_days,
                'is_mature': age_years > 10,
                'status': 'veteran' if age_years > 20 else ('mature' if age_years > 10 else ('established' if age_years > 5 else 'new'))
            }
        except Exception:
            return {}
    
    @classmethod
    def calculate_coverage_ratio(cls, premium: float, cover_amount: float) -> Dict[str, Any]:
        """Calculate coverage efficiency metrics."""
        if premium <= 0 or cover_amount <= 0:
            return {}
        
        ratio = cover_amount / premium
        annual_cost_per_1000 = (premium * 12) / (cover_amount / 1000)
        
        return {
            'coverage_ratio': round(ratio, 2),
            'annual_cost_per_1000_cover': round(annual_cost_per_1000, 2),
            'efficiency_rating': 'excellent' if ratio > 5000 else ('good' if ratio > 2000 else ('fair' if ratio > 1000 else 'review_recommended'))
        }


class DataClassifier:
    """Classifies the type of data based on column names and content"""
    
    INSURANCE_KEYWORDS = [
        'policy', 'premium', 'coverage', 'claim', 'insured', 'beneficiary',
        'deductible', 'underwriting', 'risk', 'פוליסה', 'ביטוח', 'כיסוי',
        'תביעה', 'פרמיה', 'מבוטח', 'סכום', 'השתתפות עצמית',
        # Extended Hebrew insurance terms
        'תאריך תחילה', 'תאריך סיום', 'סכום ביטוח', 'מוטב', 'מוטבים',
        'קרן פנסיה', 'גמל', 'ביטוח חיים', 'ביטוח בריאות', 'ביטוח רכב',
        'חובה', 'מקיף', 'צד ג', 'ביטוח דירה', 'תכולה', 'מבנה',
        'אחריות מקצועית', 'ביטוח נסיעות', 'ביטוח משכנתא'
    ]
    
    INVESTMENT_KEYWORDS = [
        'portfolio', 'stock', 'bond', 'fund', 'yield', 'return', 'asset',
        'equity', 'dividend', 'market', 'תיק', 'השקעה', 'מניה', 'אגרת חוב',
        'קרן', 'תשואה', 'נכס', 'דיבידנד',
        # Extended Hebrew investment terms
        'קופת גמל', 'קרן השתלמות', 'פיקדון', 'תיק ניירות ערך',
        'מדד', 'שוק ההון', 'ניהול תיקים', 'חיסכון לכל ילד'
    ]
    
    RISK_KEYWORDS = [
        'risk', 'score', 'assessment', 'rating', 'exposure', 'probability',
        'impact', 'mitigation', 'סיכון', 'ציון', 'הערכה', 'דירוג', 'חשיפה',
        # Extended Hebrew risk terms
        'הערכת סיכונים', 'ניהול סיכונים', 'סיכון תפעולי', 'סיכון שוק'
    ]
    
    SAVINGS_KEYWORDS = [
        'savings', 'balance', 'deposit', 'withdrawal', 'interest', 'account',
        'חיסכון', 'יתרה', 'הפקדה', 'משיכה', 'ריבית', 'חשבון',
        # Extended Hebrew savings terms
        'תוכנית חיסכון', 'חיסכון פנסיוני', 'קופת חיסכון', 'חיסכון לטווח ארוך'
    ]
    
    @classmethod
    def classify(cls, columns: List[str], sample_data: List[Dict]) -> Tuple[DataType, float]:
        """
        Classify the data type based on columns and content.
        Returns: (data_type, confidence)
        """
        # Combine all text for analysis
        all_text = ' '.join(columns).lower()
        if sample_data:
            for row in sample_data[:10]:
                all_text += ' ' + ' '.join(str(v).lower() for v in row.values() if v)
        
        scores = {
            DataType.INSURANCE: sum(1 for kw in cls.INSURANCE_KEYWORDS if kw.lower() in all_text),
            DataType.INVESTMENT: sum(1 for kw in cls.INVESTMENT_KEYWORDS if kw.lower() in all_text),
            DataType.RISK: sum(1 for kw in cls.RISK_KEYWORDS if kw.lower() in all_text),
            DataType.SAVINGS: sum(1 for kw in cls.SAVINGS_KEYWORDS if kw.lower() in all_text),
        }
        
        total_score = sum(scores.values())
        if total_score == 0:
            return DataType.UNKNOWN, 0.3
        
        best_type = max(scores, key=scores.get)
        confidence = scores[best_type] / max(total_score, 1)
        
        # Check for mixed data
        high_scores = [t for t, s in scores.items() if s > 0 and s >= scores[best_type] * 0.5]
        if len(high_scores) > 1:
            return DataType.MIXED, confidence * 0.8
        
        return best_type, min(confidence + 0.3, 1.0)


class AIRiskReportsService:
    """Main service for AI-powered risk and reports analysis"""
    
    def __init__(self):
        self.documents: Dict[str, Dict] = {}
        self.analyses: Dict[str, AnalysisResult] = {}
        self.reports: Dict[str, GeneratedReport] = {}
    
    def parse_file(self, filename: str, file_content: bytes, file_type: str, 
                   owner_id: str = None, owner_role: str = None) -> Dict[str, Any]:
        """
        Parse uploaded file and extract structured data.
        Supports CSV, XLS (as CSV), and ZIP containing CSV files.
        
        Args:
            filename: Name of the uploaded file
            file_content: Raw bytes of the file
            file_type: Type of file (csv, xls, xlsx, zip)
            owner_id: ID of the user who uploaded the file (for data isolation)
            owner_role: Role of the user (admin, customer, etc.)
        """
        doc_id = f"DOC-{datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(1000, 9999)}"
        
        result = {
            'document_id': doc_id,
            'filename': filename,
            'file_type': file_type,
            'file_size': len(file_content),
            'status': 'processing',
            'parsed_data': None,
            'error': None,
            'owner_id': owner_id,
            'owner_role': owner_role,
            'created_at': datetime.now().isoformat()
        }
        
        try:
            file_type_lower = file_type.lower()
            
            if file_type_lower == 'zip':
                # Handle ZIP file
                parsed = self._parse_zip(file_content)
                encoding = 'utf-8'
            elif file_type_lower in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
                # Handle image files
                parsed = self._parse_image(file_content, filename, file_type_lower)
                encoding = 'binary'
            elif file_type_lower == 'pdf':
                # Handle PDF files
                parsed = self._parse_pdf(file_content, filename)
                encoding = 'binary'
            elif file_type_lower == 'xml':
                # Handle Mislaka / pension XML files
                parsed = self._parse_pension_xml(file_content, filename)
                encoding = 'utf-8'
            elif file_type_lower in ['xls', 'xlsx']:
                # Handle Excel files properly using openpyxl / xlrd
                parsed = self._parse_excel(file_content, filename, file_type_lower)
                encoding = 'binary'
                # If _parse_excel returned empty (libraries missing), detect and warn
                if not parsed.get('rows') and not parsed.get('columns'):
                    print(f"[AI_REPORTS] Excel parse returned no data for {filename}. "
                          "Ensure openpyxl (xlsx) or xlrd (xls) is installed.")
            elif file_type_lower == 'csv':
                # Parse CSV text files
                encoding = self._detect_encoding(file_content)
                text_content = file_content.decode(encoding, errors='replace')
                parsed = self._parse_csv(text_content)
            else:
                # Unknown type - detect if binary before trying CSV
                if self._is_binary_content(file_content):
                    # Binary file: try Excel first, then fall back to metadata-only
                    parsed = self._parse_excel(file_content, filename, 'xlsx')
                    encoding = 'binary'
                    if not parsed.get('rows'):
                        parsed = {
                            'columns': ['filename', 'size_bytes', 'type', 'note'],
                            'rows': [{
                                'filename': filename,
                                'size_bytes': len(file_content),
                                'type': file_type,
                                'note': 'Binary file - use specific format for full parsing'
                            }],
                            'file_type': 'binary'
                        }
                else:
                    # Try to parse as CSV text
                    encoding = self._detect_encoding(file_content)
                    text_content = file_content.decode(encoding, errors='replace')
                    parsed = self._parse_csv(text_content)
            
            result['encoding'] = encoding
            result['parsed_data'] = parsed
            result['status'] = 'completed'
            result['row_count'] = len(parsed.get('rows', []))
            result['column_count'] = len(parsed.get('columns', []))
            
            # Store document
            self.documents[doc_id] = result
            
            # Auto-save for persistence
            self.save_data()

            _risk_report_audit('risk_report_document_uploaded', doc_id, {
                'row_count': result.get('row_count'),
                'column_count': result.get('column_count'),
                'status': result.get('status'),
            })
            
        except Exception as e:
            result['status'] = 'failed'
            result['error'] = str(e)
        
        return result
    
    def _detect_encoding(self, content: bytes) -> str:
        """Detect file encoding"""
        # Check for BOM markers
        if content.startswith(b'\xef\xbb\xbf'):
            return 'utf-8-sig'
        if content.startswith(b'\xff\xfe'):
            return 'utf-16-le'
        if content.startswith(b'\xfe\xff'):
            return 'utf-16-be'
        
        # Try common encodings
        for encoding in ['utf-8', 'windows-1255', 'iso-8859-8', 'windows-1252', 'latin-1']:
            try:
                content.decode(encoding)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue
        
        return 'utf-8'
    
    @staticmethod
    def _is_binary_content(content: bytes) -> bool:
        """
        Detect whether file content is binary (not parseable as plain text).
        Checks for common binary file signatures and control character density.
        """
        if not content:
            return False
        # Check known binary file signatures
        binary_signatures = [
            b'PK',           # ZIP / OOXML (xlsx, docx, etc.)
            b'\xd0\xcf\x11', # OLE2 (xls, doc, etc.)
            b'\x89PNG',      # PNG
            b'\xff\xd8\xff', # JPEG
            b'GIF8',         # GIF
            b'%PDF',         # PDF
            b'RIFF',         # RIFF (webp, wav, etc.)
        ]
        header = content[:8]
        for sig in binary_signatures:
            if header.startswith(sig):
                return True
        # Check for high density of non-text bytes in first 512 bytes
        sample = content[:512]
        non_text = sum(1 for b in sample if b < 0x09 or (0x0E <= b < 0x20 and b != 0x1B))
        return (non_text / max(len(sample), 1)) > 0.10
    
    def _parse_csv(self, text_content: str) -> Dict[str, Any]:
        """Parse CSV content"""
        # Try different delimiters
        for delimiter in [',', ';', '\t', '|']:
            try:
                reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
                rows = list(reader)
                if rows and len(rows[0]) > 1:
                    columns = list(rows[0].keys()) if rows else []
                    return {
                        'columns': columns,
                        'rows': rows,
                        'delimiter': delimiter
                    }
            except Exception:
                continue
        
        # Fallback: simple line parsing
        lines = text_content.strip().split('\n')
        if lines:
            columns = lines[0].split(',')
            rows = []
            for line in lines[1:]:
                values = line.split(',')
                rows.append(dict(zip(columns, values)))
            return {'columns': columns, 'rows': rows, 'delimiter': ','}
        
        return {'columns': [], 'rows': [], 'delimiter': ','}
    
    def _parse_excel(self, content: bytes, filename: str, ext: str) -> Dict[str, Any]:
        """
        Parse Excel files (.xls and .xlsx) from Mislaka data.
        Extracts pension and insurance data from Excel format.
        
        IMPORTANT: XLSX files are ZIP-based OOXML archives. They MUST be parsed
        with openpyxl (xlsx) or xlrd (xls) - never as plain text/CSV, which
        would produce garbled PK!... binary output.
        """
        rows = []
        columns = []
        pension_data = None
        sheet_names = []
        parse_method = None
        
        try:
            if ext == 'xlsx':
                # Use openpyxl for .xlsx files
                try:
                    import openpyxl
                    from openpyxl import load_workbook
                    
                    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
                    sheet_names = wb.sheetnames
                    parse_method = 'openpyxl'
                    
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        sheet_rows = list(sheet.iter_rows(values_only=True))
                        
                        if not sheet_rows:
                            continue
                        
                        # First row as headers
                        header_row = sheet_rows[0] if sheet_rows else []
                        sheet_columns = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(header_row)]
                        
                        # Add columns that don't exist yet
                        for col in sheet_columns:
                            if col and col not in columns:
                                columns.append(col)
                        
                        # Process data rows
                        for row in sheet_rows[1:]:
                            if row and any(cell is not None for cell in row):
                                row_dict = {}
                                for i, cell in enumerate(row):
                                    if i < len(sheet_columns):
                                        col_name = sheet_columns[i]
                                        # Convert dates, numbers etc. to string representation
                                        if cell is None:
                                            row_dict[col_name] = ''
                                        elif hasattr(cell, 'isoformat'):
                                            row_dict[col_name] = cell.isoformat()
                                        else:
                                            row_dict[col_name] = cell
                                rows.append(row_dict)
                    
                    wb.close()
                    print(f"[AI_REPORTS] Successfully parsed XLSX '{filename}': "
                          f"{len(sheet_names)} sheets, {len(columns)} columns, {len(rows)} rows")
                    
                except ImportError:
                    print("[AI_REPORTS] WARNING: openpyxl not installed. "
                          "Cannot parse XLSX files. Install with: pip install openpyxl")
                except Exception as e:
                    print(f"[AI_REPORTS] openpyxl error parsing '{filename}': {e}")
                    
            elif ext == 'xls':
                # Use xlrd for older .xls files
                try:
                    import xlrd
                    
                    wb = xlrd.open_workbook(file_contents=content)
                    parse_method = 'xlrd'
                    
                    for sheet_idx in range(wb.nsheets):
                        sheet = wb.sheet_by_index(sheet_idx)
                        sheet_names.append(sheet.name)
                        
                        if sheet.nrows == 0:
                            continue
                        
                        # First row as headers
                        header_row = sheet.row_values(0) if sheet.nrows > 0 else []
                        sheet_columns = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(header_row)]
                        
                        # Add columns that don't exist yet
                        for col in sheet_columns:
                            if col and col not in columns:
                                columns.append(col)
                        
                        # Process data rows
                        for row_idx in range(1, sheet.nrows):
                            row = sheet.row_values(row_idx)
                            if row and any(cell for cell in row):
                                row_dict = {}
                                for i, cell in enumerate(row):
                                    if i < len(sheet_columns):
                                        col_name = sheet_columns[i]
                                        row_dict[col_name] = cell if cell else ''
                                rows.append(row_dict)
                    
                    print(f"[AI_REPORTS] Successfully parsed XLS '{filename}': "
                          f"{len(sheet_names)} sheets, {len(columns)} columns, {len(rows)} rows")
                    
                except ImportError:
                    print("[AI_REPORTS] WARNING: xlrd not installed. "
                          "Cannot parse XLS files. Install with: pip install xlrd")
                except Exception as e:
                    print(f"[AI_REPORTS] xlrd error parsing '{filename}': {e}")
            
            # Try to detect if this is Mislaka pension data
            if rows and columns:
                pension_data = self._detect_mislaka_excel_data(columns, rows, filename)
            
        except Exception as e:
            print(f"[AI_REPORTS] Error parsing Excel file {filename}: {e}")
        
        return {
            'columns': columns,
            'rows': rows,
            'pension_data': pension_data,
            'file_type': 'excel',
            'original_filename': filename,
            'sheet_names': sheet_names,
            'parse_method': parse_method or 'none'
        }
    
    def _detect_mislaka_excel_data(self, columns: List[str], rows: List[Dict], filename: str) -> Optional[Dict[str, Any]]:
        """
        Detect and extract Mislaka pension data from Excel columns/rows.
        Maps Hebrew column names to pension data structure.
        """
        # Hebrew column name mappings for Mislaka data
        column_mappings = {
            # Client fields
            'שם': 'client_name',
            'שם מלא': 'client_name', 
            'שם פרטי': 'first_name',
            'שם משפחה': 'last_name',
            'תעודת זהות': 'id_number',
            'ת.ז': 'id_number',
            'ת"ז': 'id_number',
            'מספר זהות': 'id_number',
            'תאריך לידה': 'birth_date',
            
            # Provider/Product fields
            'יצרן': 'provider',
            'שם יצרן': 'provider',
            'חברה': 'provider',
            'שם חברה': 'provider',
            'מוצר': 'product_name',
            'שם מוצר': 'product_name',
            'סוג מוצר': 'product_type',
            'סוג קופה': 'product_type',
            
            # Policy fields
            'מספר פוליסה': 'policy_number',
            'מס פוליסה': 'policy_number',
            'מספר חשבון': 'policy_number',
            'מס חשבון': 'policy_number',
            
            # Balance fields
            'יתרה': 'balance',
            'יתרה כוללת': 'total_balance',
            'סך צבירה': 'total_balance',
            'צבירה': 'total_balance',
            'סה"כ צבירה': 'total_balance',
            'יתרת תגמולים': 'savings_balance',
            'תגמולים': 'savings_balance',
            'יתרת פיצויים': 'severance_balance',
            'פיצויים': 'severance_balance',
            
            # Fee fields
            'דמי ניהול': 'management_fee',
            'דמי ניהול מצבירה': 'management_fee_savings',
            'דמי ניהול מהפקדות': 'management_fee_deposits',
            'עמלה': 'management_fee',
            
            # Status fields
            'סטטוס': 'status',
            'מצב': 'status',
            'סטטוס פוליסה': 'status',
            
            # Section 14
            'סעיף 14': 'section14',
            'סעיף14': 'section14',
            
            # Employer
            'מעסיק': 'employer_name',
            'שם מעסיק': 'employer_name',
        }
        
        # Check if this looks like pension data
        pension_indicators = ['יצרן', 'פוליסה', 'צבירה', 'יתרה', 'תגמולים', 'פיצויים', 'קופה', 'פנסיה', 'ביטוח', 'גמל']
        columns_lower = [str(c).lower() for c in columns]
        
        is_pension_data = any(
            any(indicator in col for indicator in pension_indicators) 
            for col in columns_lower
        )
        
        if not is_pension_data:
            return None
        
        # Map columns to standardized names
        mapped_columns = {}
        for col in columns:
            col_str = str(col).strip()
            for hebrew_name, english_name in column_mappings.items():
                if hebrew_name in col_str or col_str == hebrew_name:
                    mapped_columns[col] = english_name
                    break
        
        # Extract client and account data
        client_info = {}
        accounts = []
        
        for row in rows:
            account = {}
            
            for original_col, value in row.items():
                if original_col in mapped_columns:
                    mapped_name = mapped_columns[original_col]
                    
                    # Convert value
                    if value is not None and value != '':
                        if mapped_name in ['total_balance', 'savings_balance', 'severance_balance', 
                                          'management_fee', 'management_fee_savings', 'management_fee_deposits']:
                            try:
                                account[mapped_name] = float(str(value).replace(',', '').replace('₪', '').strip())
                            except:
                                account[mapped_name] = 0
                        elif mapped_name == 'section14':
                            account[mapped_name] = str(value).lower() in ['כן', 'yes', '1', 'true', 'v', '✓']
                        elif mapped_name in ['client_name', 'first_name', 'last_name', 'id_number', 'birth_date']:
                            client_info[mapped_name] = str(value).strip()
                        else:
                            account[mapped_name] = str(value).strip()
            
            if account.get('provider') or account.get('policy_number') or account.get('total_balance'):
                accounts.append(account)
        
        # Build full name if we have parts
        if client_info.get('first_name') or client_info.get('last_name'):
            parts = [client_info.get('first_name', ''), client_info.get('last_name', '')]
            client_info['full_name'] = ' '.join(p for p in parts if p)
        elif client_info.get('client_name'):
            client_info['full_name'] = client_info['client_name']

        # Normalize customer identity fields for report consistency.
        client_info = self._normalize_client_profile_fields(client_info)
        
        if not accounts and not client_info:
            return None
        
        # Calculate totals
        total_balance = sum(a.get('total_balance', 0) for a in accounts)
        total_severance = sum(a.get('severance_balance', 0) for a in accounts)
        
        return {
            'client': client_info,
            'accounts': accounts,
            'totals': {
                'total_balance': total_balance,
                'total_balance_formatted': f"₪{total_balance:,.0f}",
                'total_severance': total_severance,
                'total_severance_formatted': f"₪{total_severance:,.0f}",
                'account_count': len(accounts),
                'provider_count': len(set(a.get('provider', '') for a in accounts if a.get('provider'))),
                'providers': list(set(a.get('provider', '') for a in accounts if a.get('provider'))),
            },
            'header': {
                'source': 'Excel',
                'filename': filename,
            }
        }
    
    def _parse_zip(self, content: bytes) -> Dict[str, Any]:
        """Parse ZIP file containing CSV, XML (pension), image, and PDF files"""
        combined_data = {
            'columns': [],
            'rows': [],
            'files': [],
            'pension_data': None,
            'integrity': {
                'zip_file_count': 0,
                'affiliated_files_processed': 0,
                'pension_sources': [],
                'issues': [],
            }
        }
        
        with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
            for name in zf.namelist():
                # Skip directories and hidden files
                if name.endswith('/') or name.startswith('__') or name.startswith('.'):
                    continue
                combined_data['integrity']['zip_file_count'] += 1
                
                name_lower = name.lower()
                ext = name_lower.split('.')[-1] if '.' in name_lower else ''
                
                with zf.open(name) as f:
                    file_content = f.read()
                
                parsed = None
                file_type = ext
                
                if ext == 'csv':
                    encoding = self._detect_encoding(file_content)
                    text_content = file_content.decode(encoding, errors='replace')
                    parsed = self._parse_csv(text_content)
                elif ext == 'xml':
                    # Check if it's a pension/insurance XML file
                    parsed = self._parse_pension_xml(file_content, name)
                    if parsed:
                        file_type = 'pension_xml'
                        # Store pension data separately for enhanced analysis
                        if parsed.get('pension_data'):
                            combined_data['integrity']['affiliated_files_processed'] += 1
                            combined_data['integrity']['pension_sources'].append(name)
                            combined_data['pension_data'] = self._merge_pension_data_records(
                                combined_data.get('pension_data'),
                                parsed.get('pension_data')
                            )
                elif ext in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
                    parsed = self._parse_image(file_content, name, ext)
                elif ext == 'pdf':
                    parsed = self._parse_pdf(file_content, name)
                elif ext in ['xls', 'xlsx']:
                    # Parse Excel files properly
                    parsed = self._parse_excel(file_content, name, ext)
                    if parsed:
                        file_type = 'excel'
                        # Check if this looks like Mislaka data
                        if parsed.get('pension_data'):
                            combined_data['integrity']['affiliated_files_processed'] += 1
                            combined_data['integrity']['pension_sources'].append(name)
                            combined_data['pension_data'] = self._merge_pension_data_records(
                                combined_data.get('pension_data'),
                                parsed.get('pension_data')
                            )
                
                if parsed:
                    combined_data['files'].append({
                        'name': name,
                        'type': file_type,
                        'columns': parsed.get('columns', []),
                        'row_count': len(parsed.get('rows', []))
                    })
                    
                    # Merge columns and rows
                    for col in parsed.get('columns', []):
                        if col not in combined_data['columns']:
                            combined_data['columns'].append(col)
                    combined_data['rows'].extend(parsed.get('rows', []))

        if (
            combined_data['integrity']['zip_file_count'] > 0
            and combined_data['integrity']['affiliated_files_processed'] == 0
        ):
            combined_data['integrity']['issues'].append(
                'No Swiftness-affiliated XML/Excel records detected in ZIP'
            )
        
        return combined_data

    @staticmethod
    def _normalize_customer_identifier(identifier: Any) -> str:
        """Normalize customer identifiers while preserving non-digit fallback values."""
        text = str(identifier or '').strip()
        if not text:
            return ''

        digits = re.sub(r'\D', '', text)
        # Israeli IDs are 9 digits; some sources drop leading zeroes.
        if 7 <= len(digits) <= 9:
            return digits.zfill(9)
        return text

    @staticmethod
    def _is_valid_israeli_id(identifier: str) -> bool:
        """Validate Israeli ID checksum for 9-digit identifiers."""
        if not identifier or not identifier.isdigit() or len(identifier) != 9:
            return False

        total = 0
        for index, char in enumerate(identifier):
            digit = int(char)
            factor = 1 if index % 2 == 0 else 2
            product = digit * factor
            if product > 9:
                product -= 9
            total += product

        return total % 10 == 0

    @staticmethod
    def _normalize_birth_date(value: Any) -> Tuple[str, str]:
        """
        Normalize birth dates into:
          - raw canonical format: YYYYMMDD
          - display format: DD/MM/YYYY
        """
        raw_input = str(value or '').strip()
        if not raw_input:
            return '', ''

        def _to_pair(dt: datetime) -> Tuple[str, str]:
            return dt.strftime('%Y%m%d'), dt.strftime('%d/%m/%Y')

        digits = re.sub(r'\D', '', raw_input)
        if len(digits) == 8:
            # Prefer YYYYMMDD (e.g. 19781111), fallback to DDMMYYYY.
            for year, month, day in [
                (digits[0:4], digits[4:6], digits[6:8]),
                (digits[4:8], digits[2:4], digits[0:2]),
            ]:
                try:
                    parsed = datetime(int(year), int(month), int(day))
                    if 1900 <= parsed.year <= 2100:
                        return _to_pair(parsed)
                except Exception:
                    continue

        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y.%m.%d', '%d.%m.%Y'):
            try:
                parsed = datetime.strptime(raw_input, fmt)
                if 1900 <= parsed.year <= 2100:
                    return _to_pair(parsed)
            except Exception:
                continue

        try:
            parsed = datetime.fromisoformat(raw_input.replace('Z', '+00:00'))
            if 1900 <= parsed.year <= 2100:
                return _to_pair(parsed)
        except Exception:
            pass

        return digits if len(digits) == 8 else raw_input, ''

    def _normalize_client_profile_fields(self, client_info: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Normalize and enrich client identity fields used in reports."""
        normalized = dict(client_info or {})

        normalized_id = self._normalize_customer_identifier(normalized.get('id_number'))
        if normalized_id:
            normalized['id_number'] = normalized_id
            normalized['id_israeli_valid'] = self._is_valid_israeli_id(normalized_id)

        birth_raw, birth_display = self._normalize_birth_date(normalized.get('birth_date'))
        if birth_raw:
            normalized['birth_date_raw'] = birth_raw
        if birth_display:
            normalized['birth_date'] = birth_display

        return normalized

    def _merge_pension_data_records(
        self,
        current: Optional[Dict[str, Any]],
        incoming: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Merge multiple pension-affiliated payloads from ZIP members while preserving
        customer profile and totals integrity.
        """
        if not isinstance(current, dict) or not current:
            base = copy.deepcopy(incoming or {})
            if isinstance(base.get('client'), dict):
                base['client'] = self._normalize_client_profile_fields(base.get('client', {}))
            return base
        if not isinstance(incoming, dict) or not incoming:
            return copy.deepcopy(current)

        merged = copy.deepcopy(current)
        incoming_copy = copy.deepcopy(incoming)

        def _dedupe_by_key(rows: List[Dict[str, Any]], key_fields: List[str]) -> List[Dict[str, Any]]:
            results: List[Dict[str, Any]] = []
            seen: set = set()
            for row in rows:
                if not isinstance(row, dict):
                    continue
                key = tuple(str(row.get(field, '') or '').strip() for field in key_fields)
                if not any(key):
                    payload = json.dumps(row, sort_keys=True, ensure_ascii=False)
                    key = ('hash', hashlib.sha256(payload.encode('utf-8')).hexdigest())
                if key in seen:
                    continue
                seen.add(key)
                results.append(row)
            return results

        merged_client = merged.get('client', {})
        incoming_client = incoming_copy.get('client', {})
        if isinstance(merged_client, list):
            merged_client = merged_client[0] if merged_client else {}
        if isinstance(incoming_client, list):
            incoming_client = incoming_client[0] if incoming_client else {}

        merged_client = self._normalize_client_profile_fields(merged_client if isinstance(merged_client, dict) else {})
        incoming_client = self._normalize_client_profile_fields(incoming_client if isinstance(incoming_client, dict) else {})
        for key, value in incoming_client.items():
            if value and not merged_client.get(key):
                merged_client[key] = value
        merged_anomalies = list(merged.get('anomalies', []) or [])
        if (
            merged_client.get('id_number')
            and incoming_client.get('id_number')
            and str(merged_client.get('id_number')) != str(incoming_client.get('id_number'))
        ):
            merged_anomalies.append(
                f"Client ID mismatch across affiliated files: {merged_client.get('id_number')} vs {incoming_client.get('id_number')}"
            )
        merged['anomalies'] = merged_anomalies
        merged['client'] = merged_client

        merged_accounts = list(merged.get('accounts', []) or [])
        incoming_accounts = list(incoming_copy.get('accounts', []) or [])
        merged['accounts'] = _dedupe_by_key(
            merged_accounts + incoming_accounts,
            ['policy_number', 'provider', 'product_type', 'start_date']
        )

        merged_contributions = list(merged.get('contributions', []) or [])
        incoming_contributions = list(incoming_copy.get('contributions', []) or [])
        merged['contributions'] = _dedupe_by_key(
            merged_contributions + incoming_contributions,
            ['period', 'policy_number', 'employer_name', 'employee_amount', 'employer_amount', 'severance_amount', 'total_amount']
        )

        merged_severance = list(merged.get('severance', []) or [])
        incoming_severance = list(incoming_copy.get('severance', []) or [])
        merged['severance'] = _dedupe_by_key(
            merged_severance + incoming_severance,
            ['employer_name', 'section14_date', 'total_severance']
        )

        merged_employers = list(merged.get('employers', []) or [])
        incoming_employers = list(incoming_copy.get('employers', []) or [])
        employer_rows: List[Dict[str, Any]] = []
        for employer in (merged_employers + incoming_employers):
            if isinstance(employer, dict):
                employer_rows.append(employer)
            else:
                name = str(employer or '').strip()
                if name:
                    employer_rows.append({'id': '', 'name': name})
        merged['employers'] = _dedupe_by_key(employer_rows, ['id', 'name'])

        merged_providers = list(merged.get('providers', []) or [])
        incoming_providers = list(incoming_copy.get('providers', []) or [])
        if merged_providers and isinstance(merged_providers[0], dict):
            merged['providers'] = _dedupe_by_key(merged_providers + incoming_providers, ['code', 'name'])
        else:
            merged['providers'] = sorted({
                str(p).strip()
                for p in (merged_providers + incoming_providers)
                if str(p).strip()
            })

        header = dict(merged.get('header', {}) or {})
        incoming_header = dict(incoming_copy.get('header', {}) or {})
        for key, value in incoming_header.items():
            if value and not header.get(key):
                header[key] = value
        merged['header'] = header

        # Recompute key totals after merge.
        total_balance = sum(self._to_float_amount(a.get('total_balance')) for a in merged.get('accounts', []))
        total_savings = sum(self._to_float_amount(a.get('savings_balance')) for a in merged.get('accounts', []))
        total_severance = (
            sum(self._to_float_amount(a.get('severance_balance')) for a in merged.get('accounts', [])) +
            sum(self._to_float_amount(s.get('total_severance')) for s in merged.get('severance', []))
        )
        provider_names = sorted({
            str(a.get('provider', '')).strip()
            for a in merged.get('accounts', [])
            if str(a.get('provider', '')).strip()
        })
        totals = dict(merged.get('totals', {}) or {})
        totals.update({
            'total_balance': round(total_balance, 2),
            'total_balance_formatted': f"₪{total_balance:,.2f}",
            'total_savings': round(total_savings, 2),
            'total_savings_formatted': f"₪{total_savings:,.2f}",
            'total_severance': round(total_severance, 2),
            'total_severance_formatted': f"₪{total_severance:,.2f}",
            'total_coverage': round(
                sum(
                    self._to_float_amount(a.get('coverage_amount'))
                    or self._to_float_amount(a.get('death_coverage')) + self._to_float_amount(a.get('disability_coverage'))
                    for a in merged.get('accounts', [])
                ),
                2
            ),
            'account_count': len(merged.get('accounts', [])),
            'provider_count': len(provider_names),
            'providers': provider_names,
            'section14_coverage': any(bool(a.get('section14')) for a in merged.get('accounts', [])),
        })
        merged['totals'] = totals

        return merged
    
    def _parse_pension_xml(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse Israeli pension/insurance XML files using the PensionDataAgent.
        
        Supports Mislaka (מסלקה) interface standards:
        - Type 1: Holdings (אחזקות)
        - Type 2: Pre-Advice (הודעה מקדימה)
        - Type 3: Holdings + Pre-Advice Combined
        - Type 17: Severance (פיצויים)
        """
        try:
            from services.pension_data_agent import get_pension_agent, is_pension_xml
            
            # Check if this is a pension XML file
            if not is_pension_xml(content):
                # Not a pension XML, try to parse as generic XML
                return self._parse_generic_xml(content, filename)
            
            # Process with PensionDataAgent
            agent = get_pension_agent()
            result = agent.process_xml_content(content)
            
            pension_data = result.get('data', {})
            if isinstance(pension_data.get('client'), dict):
                pension_data['client'] = self._normalize_client_profile_fields(pension_data.get('client', {}))
            report_text = result.get('report', '')
            
            # Convert to CSV-like format for AI analysis
            columns, rows = agent.to_csv_format(pension_data)
            
            # Add summary data as additional rows
            summary = pension_data.get('summary', {})
            header = pension_data.get('header', {})
            clients = pension_data.get('client', [])
            
            # Add header info as rows
            meta_rows = [
                {'מספר פוליסה': 'סוג ממשק', 'יצרן': pension_data.get('interface_type', ''), 'סוג מוצר': '', 'שם מוצר': '', 'סטטוס': '', 'יתרה': '', 'פיצויים': '', 'מעסיק': ''},
                {'מספר פוליסה': 'גרסת סכמה', 'יצרן': header.get('schema_version', ''), 'סוג מוצר': '', 'שם מוצר': '', 'סטטוס': '', 'יתרה': '', 'פיצויים': '', 'מעסיק': ''},
            ]
            
            # Add client info
            if clients:
                client = clients[0] if isinstance(clients, list) else clients
                meta_rows.append({
                    'מספר פוליסה': 'לקוח',
                    'יצרן': client.get('name', ''),
                    'סוג מוצר': '',
                    'שם מוצר': '',
                    'סטטוס': '',
                    'יתרה': '',
                    'פיצויים': '',
                    'מעסיק': ''
                })
            
            # Prepend meta rows
            all_rows = meta_rows + rows
            
            return {
                'columns': columns,
                'rows': all_rows,
                'delimiter': None,
                'file_type': 'pension_xml',
                'original_filename': filename,
                'pension_data': pension_data,
                'pension_report': report_text
            }
            
        except ImportError:
            print("[AI_REPORTS] PensionDataAgent not available, falling back to generic XML parsing")
            return self._parse_generic_xml(content, filename)
        except Exception as e:
            print(f"[AI_REPORTS] Error parsing pension XML: {e}")
            return self._parse_generic_xml(content, filename)
    
    def _parse_generic_xml(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse generic XML file into tabular format.
        """
        try:
            from defusedxml import ElementTree as DefusedET
            
            # Decode content
            encoding = self._detect_encoding(content)
            xml_str = content.decode(encoding, errors='replace')
            
            # Parse XML
            root = DefusedET.fromstring(xml_str)
            
            # Extract all leaf elements as rows
            columns = ['element', 'value', 'path']
            rows = []
            
            def extract_elements(elem, path=""):
                current_path = f"{path}/{elem.tag}" if path else elem.tag
                
                # If element has text content
                if elem.text and elem.text.strip():
                    rows.append({
                        'element': elem.tag,
                        'value': elem.text.strip(),
                        'path': current_path
                    })
                
                # Process attributes
                for attr, val in elem.attrib.items():
                    rows.append({
                        'element': f"{elem.tag}@{attr}",
                        'value': val,
                        'path': current_path
                    })
                
                # Process children
                for child in elem:
                    extract_elements(child, current_path)
            
            extract_elements(root)
            
            return {
                'columns': columns,
                'rows': rows,
                'delimiter': None,
                'file_type': 'xml',
                'original_filename': filename
            }
            
        except Exception as e:
            print(f"[AI_REPORTS] Error parsing generic XML: {e}")
            return {'columns': [], 'rows': [], 'file_type': 'xml', 'error': str(e)}
    
    def _parse_image(self, content: bytes, filename: str, file_type: str) -> Dict[str, Any]:
        """
        Parse image file and extract metadata for analysis.
        Creates a structured data format from image properties.
        """
        # Basic image metadata extraction
        file_size = len(content)
        
        # Try to detect image dimensions from header
        width, height = 0, 0
        if file_type in ['png']:
            # PNG header: width at bytes 16-19, height at bytes 20-23
            if len(content) >= 24:
                width = int.from_bytes(content[16:20], 'big')
                height = int.from_bytes(content[20:24], 'big')
        elif file_type in ['jpg', 'jpeg']:
            # JPEG - simplified dimension detection
            # Look for SOF0 marker (0xFF 0xC0)
            try:
                i = 0
                while i < len(content) - 10:
                    if content[i] == 0xFF and content[i+1] in [0xC0, 0xC1, 0xC2]:
                        height = int.from_bytes(content[i+5:i+7], 'big')
                        width = int.from_bytes(content[i+7:i+9], 'big')
                        break
                    i += 1
            except:
                pass
        
        # Extract any text from filename for language detection
        filename_text = filename.replace('_', ' ').replace('-', ' ')
        
        # Create structured data for analysis
        columns = ['property', 'value', 'category']
        rows = [
            {'property': 'filename', 'value': filename, 'category': 'metadata'},
            {'property': 'file_type', 'value': file_type.upper(), 'category': 'metadata'},
            {'property': 'file_size_bytes', 'value': str(file_size), 'category': 'metadata'},
            {'property': 'file_size_kb', 'value': str(round(file_size / 1024, 2)), 'category': 'metadata'},
            {'property': 'width_px', 'value': str(width), 'category': 'dimensions'},
            {'property': 'height_px', 'value': str(height), 'category': 'dimensions'},
            {'property': 'resolution', 'value': f'{width}x{height}', 'category': 'dimensions'},
            {'property': 'document_type', 'value': 'image', 'category': 'classification'},
            {'property': 'source_name', 'value': filename_text, 'category': 'context'},
        ]
        
        return {
            'columns': columns,
            'rows': rows,
            'delimiter': None,
            'file_type': 'image',
            'original_filename': filename
        }
    
    def _parse_pdf(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse PDF file and extract metadata for analysis.
        Creates a structured data format from PDF properties.
        """
        file_size = len(content)
        
        # Basic PDF metadata extraction
        page_count = 0
        title = filename
        author = ''
        creation_date = ''
        
        # Simple PDF parsing for metadata
        try:
            content_str = content[:4096].decode('latin-1', errors='ignore')
            
            # Count pages (approximate)
            page_count = content.count(b'/Type /Page') or content.count(b'/Type/Page')
            
            # Extract title if present
            if '/Title' in content_str:
                start = content_str.find('/Title')
                if start != -1:
                    # Try to extract title value
                    paren_start = content_str.find('(', start)
                    paren_end = content_str.find(')', paren_start) if paren_start != -1 else -1
                    if paren_start != -1 and paren_end != -1:
                        title = content_str[paren_start+1:paren_end][:100]
            
            # Extract author if present
            if '/Author' in content_str:
                start = content_str.find('/Author')
                if start != -1:
                    paren_start = content_str.find('(', start)
                    paren_end = content_str.find(')', paren_start) if paren_start != -1 else -1
                    if paren_start != -1 and paren_end != -1:
                        author = content_str[paren_start+1:paren_end][:100]
        except:
            pass
        
        # Extract text from filename for context
        filename_text = filename.replace('_', ' ').replace('-', ' ').replace('.pdf', '')
        
        # Detect if filename contains Hebrew
        has_hebrew = bool(re.search(r'[\u0590-\u05FF]', filename_text))
        
        # Create structured data for analysis
        columns = ['property', 'value', 'category']
        rows = [
            {'property': 'filename', 'value': filename, 'category': 'metadata'},
            {'property': 'file_type', 'value': 'PDF', 'category': 'metadata'},
            {'property': 'file_size_bytes', 'value': str(file_size), 'category': 'metadata'},
            {'property': 'file_size_kb', 'value': str(round(file_size / 1024, 2)), 'category': 'metadata'},
            {'property': 'file_size_mb', 'value': str(round(file_size / (1024*1024), 2)), 'category': 'metadata'},
            {'property': 'page_count', 'value': str(page_count), 'category': 'content'},
            {'property': 'title', 'value': title, 'category': 'metadata'},
            {'property': 'author', 'value': author, 'category': 'metadata'},
            {'property': 'document_type', 'value': 'pdf', 'category': 'classification'},
            {'property': 'source_name', 'value': filename_text, 'category': 'context'},
            {'property': 'has_hebrew', 'value': str(has_hebrew), 'category': 'language'},
        ]
        
        return {
            'columns': columns,
            'rows': rows,
            'delimiter': None,
            'file_type': 'pdf',
            'original_filename': filename,
            'page_count': page_count
        }
    
    def analyze(self, document_id: str) -> AnalysisResult:
        """
        Perform advanced AI/BI analysis on parsed document using inductive reasoning.
        
        This method learns from the uploaded data through:
        1. Statistical profiling of all columns
        2. Inductive pattern recognition
        3. Correlation analysis between fields
        4. Domain-specific semantic analysis
        5. Language-aware interpretation
        
        Returns comprehensive analysis with factors, patterns, and risk assessment.
        """
        start_time = datetime.now()
        
        if document_id not in self.documents:
            raise ValueError(f"Document {document_id} not found")
        
        doc = self.documents[document_id]
        parsed = doc.get('parsed_data', {})
        columns = parsed.get('columns', [])
        rows = parsed.get('rows', [])
        
        # =====================================================================
        # PHASE 1: INDUCTIVE DATA PROFILING
        # Learn the structure and semantics of the uploaded data
        # =====================================================================
        
        # Combine all text for language detection
        all_text = ' '.join(columns)
        for row in rows[:50]:  # Sample more rows for better detection
            all_text += ' ' + ' '.join(str(v) for v in row.values() if v)
        
        # Detect language with confidence
        lang_code, lang_name, lang_confidence = LanguageDetector.detect(all_text)
        
        # Classify data type using semantic analysis
        data_type, type_confidence = DataClassifier.classify(columns, rows)
        
        # =====================================================================
        # PHASE 2: HEBREW DOCUMENT EXTRACTION (for Hebrew insurance documents)
        # Extract structured fields from Hebrew documents using pattern recognition
        # =====================================================================
        
        hebrew_extracted = {}
        if lang_code == 'hebrew' or any(re.search(r'[\u0590-\u05FF]', str(v)) for row in rows[:10] for v in row.values()):
            hebrew_extracted = self._extract_hebrew_document_data(all_text, rows)
        
        # =====================================================================
        # PHASE 3: ADVANCED STATISTICAL ANALYSIS (BI)
        # Compute comprehensive statistics for each column
        # =====================================================================
        
        column_profiles = self._profile_columns(columns, rows)
        
        # =====================================================================
        # PHASE 4: CORRELATION & RELATIONSHIP DISCOVERY
        # Find relationships between different data fields
        # =====================================================================
        
        correlations = self._find_correlations(columns, rows, column_profiles)
        
        # =====================================================================
        # PHASE 5: INDUCTIVE PATTERN LEARNING
        # Discover patterns and rules from the data
        # =====================================================================
        
        # Extract factors with enhanced analysis
        factors = self._extract_factors_advanced(columns, rows, data_type, column_profiles)
        
        # Add Hebrew document factors if available
        if hebrew_extracted:
            factors.extend(self._create_hebrew_document_factors(hebrew_extracted, lang_code))
        
        # Find patterns using inductive reasoning
        patterns = self._find_patterns_advanced(rows, data_type, column_profiles, correlations)
        
        # Add Hebrew-specific patterns
        if hebrew_extracted:
            patterns.extend(self._find_hebrew_patterns(hebrew_extracted))
        
        # Detect anomalies with statistical backing
        anomalies = self._detect_anomalies_advanced(rows, data_type, column_profiles)
        
        # =====================================================================
        # PHASE 6: DOMAIN-SPECIFIC INSIGHTS
        # Apply domain knowledge based on detected data type
        # =====================================================================
        
        domain_insights = self._generate_domain_insights(data_type, column_profiles, rows, lang_code, hebrew_extracted)
        
        # =====================================================================
        # PHASE 6: RISK ASSESSMENT
        # Calculate comprehensive risk score
        # =====================================================================
        
        risk_score = self._calculate_risk_score_advanced(
            factors, patterns, anomalies, correlations, domain_insights
        )
        
        # Generate language-aware summary with insights
        summary = self._generate_summary_advanced(
            lang_code, data_type, len(rows), factors, risk_score, 
            column_profiles, domain_insights
        )
        
        # Extract comprehensive key metrics
        key_metrics = self._extract_key_metrics_advanced(
            rows, columns, data_type, column_profiles, correlations, domain_insights
        )
        
        processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
        
        analysis_id = f"ANA-{datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(1000, 9999)}"
        
        result = AnalysisResult(
            id=analysis_id,
            document_id=document_id,
            language=lang_code,
            language_name=lang_name,
            data_classification=data_type,
            extracted_factors=factors,
            patterns_found=patterns,
            anomalies=anomalies,
            risk_score=risk_score,
            confidence=min((lang_confidence + type_confidence) / 2, 1.0),
            processing_time_ms=processing_time,
            summary=summary,
            key_metrics=key_metrics
        )
        
        self.analyses[analysis_id] = result
        
        # Auto-save for persistence
        self.save_data()
        
        return result
    
    # =========================================================================
    # ADVANCED BI/AI ANALYSIS METHODS
    # These methods provide deep inductive analysis of uploaded data
    # =========================================================================
    
    def _profile_columns(self, columns: List[str], rows: List[Dict]) -> Dict[str, Dict]:
        """
        Create comprehensive statistical profiles for each column.
        This is the foundation of inductive data analysis.
        """
        profiles = {}
        
        for col in columns:
            profile = {
                'name': col,
                'type': 'unknown',
                'count': 0,
                'null_count': 0,
                'unique_count': 0,
                'numeric': False,
                'values': [],
                'stats': {}
            }
            
            values = []
            numeric_values = []
            
            for row in rows:
                val = row.get(col)
                if val is None or str(val).strip() == '':
                    profile['null_count'] += 1
                else:
                    profile['count'] += 1
                    values.append(str(val))
                    
                    # Try to parse as numeric
                    try:
                        clean_val = str(val).replace(',', '').replace('₪', '').replace('$', '').replace('€', '').replace('%', '')
                        num_val = float(clean_val)
                        numeric_values.append(num_val)
                    except (ValueError, TypeError):
                        pass
            
            profile['unique_count'] = len(set(values))
            profile['values'] = values[:100]  # Store sample
            
            # Determine column type and compute statistics
            if len(numeric_values) > len(values) * 0.5:  # More than 50% numeric
                profile['numeric'] = True
                profile['type'] = 'numeric'
                
                if numeric_values:
                    sorted_vals = sorted(numeric_values)
                    n = len(numeric_values)
                    mean_val = sum(numeric_values) / n
                    
                    # Variance and std dev
                    variance = sum((x - mean_val) ** 2 for x in numeric_values) / n if n > 0 else 0
                    std_dev = variance ** 0.5
                    
                    # Quartiles
                    q1_idx = int(n * 0.25)
                    q2_idx = int(n * 0.5)
                    q3_idx = int(n * 0.75)
                    
                    profile['stats'] = {
                        'min': round(min(numeric_values), 2),
                        'max': round(max(numeric_values), 2),
                        'sum': round(sum(numeric_values), 2),
                        'mean': round(mean_val, 2),
                        'median': round(sorted_vals[q2_idx] if n > 0 else 0, 2),
                        'std_dev': round(std_dev, 2),
                        'variance': round(variance, 2),
                        'q1': round(sorted_vals[q1_idx] if n > 0 else 0, 2),
                        'q3': round(sorted_vals[q3_idx] if n > 0 else 0, 2),
                        'iqr': round((sorted_vals[q3_idx] - sorted_vals[q1_idx]) if n > 0 else 0, 2),
                        'count': n,
                        'range': round(max(numeric_values) - min(numeric_values), 2)
                    }
                    
                    # Detect distribution shape
                    if std_dev > 0:
                        skewness = sum((x - mean_val) ** 3 for x in numeric_values) / (n * std_dev ** 3)
                        profile['stats']['skewness'] = round(skewness, 3)
                        profile['stats']['distribution'] = 'normal' if abs(skewness) < 0.5 else ('right_skewed' if skewness > 0 else 'left_skewed')
            else:
                profile['type'] = 'categorical'
                # Frequency distribution for categorical
                freq = {}
                for v in values:
                    freq[v] = freq.get(v, 0) + 1
                
                sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)
                profile['stats'] = {
                    'top_values': sorted_freq[:10],
                    'unique_ratio': round(profile['unique_count'] / max(len(values), 1), 3),
                    'mode': sorted_freq[0][0] if sorted_freq else None,
                    'mode_count': sorted_freq[0][1] if sorted_freq else 0
                }
            
            # Semantic type detection
            col_lower = col.lower()
            if any(x in col_lower for x in ['date', 'time', 'תאריך']):
                profile['semantic_type'] = 'datetime'
            elif any(x in col_lower for x in ['email', 'מייל']):
                profile['semantic_type'] = 'email'
            elif any(x in col_lower for x in ['phone', 'טלפון', 'נייד']):
                profile['semantic_type'] = 'phone'
            elif any(x in col_lower for x in ['price', 'amount', 'premium', 'מחיר', 'סכום', 'פרמיה']):
                profile['semantic_type'] = 'currency'
            elif any(x in col_lower for x in ['percent', 'rate', 'אחוז', 'שיעור']):
                profile['semantic_type'] = 'percentage'
            elif any(x in col_lower for x in ['id', 'number', 'מספר', 'מזהה']):
                profile['semantic_type'] = 'identifier'
            elif any(x in col_lower for x in ['name', 'שם']):
                profile['semantic_type'] = 'name'
            elif any(x in col_lower for x in ['status', 'סטטוס', 'מצב']):
                profile['semantic_type'] = 'status'
            else:
                profile['semantic_type'] = 'general'
            
            profiles[col] = profile
        
        return profiles
    
    def _find_correlations(self, columns: List[str], rows: List[Dict], 
                          profiles: Dict[str, Dict]) -> List[Dict]:
        """
        Find correlations between numeric columns.
        Uses Pearson correlation coefficient.
        """
        correlations = []
        numeric_cols = [col for col, p in profiles.items() if p['numeric']]
        
        if len(numeric_cols) < 2 or len(rows) < 3:
            return correlations
        
        # Extract numeric values for each column
        col_values = {}
        for col in numeric_cols:
            values = []
            for row in rows:
                try:
                    val = float(str(row.get(col, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                    values.append(val)
                except:
                    values.append(0)
            col_values[col] = values
        
        # Calculate correlations between pairs
        for i, col1 in enumerate(numeric_cols):
            for col2 in numeric_cols[i+1:]:
                vals1 = col_values[col1]
                vals2 = col_values[col2]
                
                n = len(vals1)
                mean1 = sum(vals1) / n
                mean2 = sum(vals2) / n
                
                # Covariance
                cov = sum((vals1[j] - mean1) * (vals2[j] - mean2) for j in range(n)) / n
                
                # Standard deviations
                std1 = (sum((x - mean1) ** 2 for x in vals1) / n) ** 0.5
                std2 = (sum((x - mean2) ** 2 for x in vals2) / n) ** 0.5
                
                # Pearson correlation
                if std1 > 0 and std2 > 0:
                    corr = cov / (std1 * std2)
                    
                    if abs(corr) > 0.3:  # Only significant correlations
                        correlations.append({
                            'column1': col1,
                            'column2': col2,
                            'correlation': round(corr, 3),
                            'strength': 'strong' if abs(corr) > 0.7 else ('moderate' if abs(corr) > 0.5 else 'weak'),
                            'direction': 'positive' if corr > 0 else 'negative'
                        })
        
        # Sort by absolute correlation
        correlations.sort(key=lambda x: abs(x['correlation']), reverse=True)
        return correlations[:10]  # Top 10 correlations
    
    def _extract_factors_advanced(self, columns: List[str], rows: List[Dict], 
                                  data_type: DataType, profiles: Dict[str, Dict]) -> List[Factor]:
        """
        Extract key factors using advanced statistical analysis.
        IMPORTANT: Excludes ID/policy numbers from statistical analysis - 
        these are for display only, not for statistical calculations.
        """
        factors = []
        
        # Columns to EXCLUDE from statistical analysis (IDs, policy numbers, etc.)
        # These should only be used for display/identification, not statistics
        exclude_patterns = [
            'id', 'מספר', 'תז', 'ת.ז', 'ת"ז', 'זהות', 'פוליסה', 'חשבון', 
            'policy', 'account', 'number', 'num', 'code', 'קוד', 'מזהה',
            'phone', 'טלפון', 'נייד', 'zip', 'מיקוד', 'index', 'row'
        ]
        
        # Columns that ARE meaningful for statistical analysis (financial data)
        meaningful_patterns = [
            'balance', 'יתרה', 'צבירה', 'תגמולים', 'פיצויים', 'חיסכון',
            'premium', 'פרמיה', 'הפקדה', 'תשלום',
            'coverage', 'כיסוי', 'ביטוח', 'סכום',
            'fee', 'דמי', 'עמלה', 'ניהול',
            'return', 'תשואה', 'רווח', 'הפסד',
            'salary', 'שכר', 'משכורת',
            'amount', 'סכום', 'ערך', 'שווי'
        ]
        
        def should_exclude_column(col_name: str) -> bool:
            """Check if column should be excluded from statistical analysis."""
            col_lower = str(col_name).lower()
            
            # Check if it's an identifier/number column
            for pattern in exclude_patterns:
                if pattern in col_lower:
                    return True
            
            return False
        
        def is_meaningful_financial_column(col_name: str, semantic_type: str) -> bool:
            """Check if column is meaningful for financial analysis."""
            col_lower = str(col_name).lower()
            
            # Currency and percentage are always meaningful
            if semantic_type in ['currency', 'percentage']:
                return True
            
            # Check for financial keywords
            for pattern in meaningful_patterns:
                if pattern in col_lower:
                    return True
            
            return False
        
        # Add statistical factors ONLY for meaningful numeric columns
        for col, profile in profiles.items():
            if profile['numeric'] and profile['stats']:
                # Skip ID/policy number columns - CRITICAL
                if should_exclude_column(col):
                    continue
                
                # Only analyze columns that are meaningful for financial analysis
                semantic_type = profile.get('semantic_type', '')
                if not is_meaningful_financial_column(col, semantic_type):
                    continue
                
                stats = profile['stats']
                
                # Determine importance based on variance and semantic type
                importance = 0.5
                if semantic_type == 'currency':
                    importance = 0.9
                elif semantic_type == 'percentage':
                    importance = 0.8
                elif stats.get('std_dev', 0) > stats.get('mean', 1) * 0.5:
                    importance = 0.7  # High variability is important
                
                factors.append(Factor(
                    name=f"{col} Analysis",
                    value={
                        'mean': stats.get('mean'),
                        'median': stats.get('median'),
                        'range': f"{stats.get('min')} - {stats.get('max')}",
                        'std_dev': stats.get('std_dev'),
                        'distribution': stats.get('distribution', 'unknown')
                    },
                    importance=importance,
                    category='statistical'
                ))
        
        # Add categorical distribution factors
        for col, profile in profiles.items():
            if not profile['numeric'] and profile['stats'].get('top_values'):
                top_vals = profile['stats']['top_values'][:5]
                
                factors.append(Factor(
                    name=f"{col} Distribution",
                    value={
                        'unique_values': profile['unique_count'],
                        'top_categories': [{'value': v, 'count': c} for v, c in top_vals],
                        'concentration': profile['stats'].get('unique_ratio', 0)
                    },
                    importance=0.6 if profile.get('semantic_type') == 'status' else 0.4,
                    category='categorical'
                ))
        
        # Data type specific factors
        if data_type == DataType.INSURANCE:
            factors.append(Factor(
                name='Insurance Data Profile',
                value={
                    'record_count': len(rows),
                    'data_completeness': round(sum(1 for p in profiles.values() if p['null_count'] == 0) / max(len(profiles), 1) * 100, 1),
                    'domain': 'insurance'
                },
                importance=0.95,
                category='domain'
            ))
        elif data_type == DataType.INVESTMENT:
            factors.append(Factor(
                name='Investment Data Profile',
                value={
                    'record_count': len(rows),
                    'numeric_fields': sum(1 for p in profiles.values() if p['numeric']),
                    'domain': 'investment'
                },
                importance=0.95,
                category='domain'
            ))
        
        return factors[:15]  # Return top 15 factors
    
    def _find_patterns_advanced(self, rows: List[Dict], data_type: DataType,
                               profiles: Dict[str, Dict], correlations: List[Dict]) -> List[Pattern]:
        """
        Find patterns using inductive reasoning.
        """
        patterns = []
        
        if len(rows) < 2:
            return patterns
        
        # Pattern 1: Data completeness patterns
        incomplete_cols = [col for col, p in profiles.items() if p['null_count'] > len(rows) * 0.1]
        if incomplete_cols:
            patterns.append(Pattern(
                type='data_quality',
                description=f"Incomplete data in {len(incomplete_cols)} columns: {', '.join(incomplete_cols[:3])}",
                affected_rows=list(range(len(rows))),
                significance=0.8
            ))
        
        # Pattern 2: Value concentration (potential data issues)
        for col, profile in profiles.items():
            if not profile['numeric'] and profile['stats'].get('unique_ratio', 1) < 0.1:
                mode = profile['stats'].get('mode')
                mode_count = profile['stats'].get('mode_count', 0)
                if mode_count > len(rows) * 0.5:
                    patterns.append(Pattern(
                        type='value_concentration',
                        description=f"High concentration in '{col}': '{mode}' appears in {mode_count}/{len(rows)} records ({round(mode_count/len(rows)*100)}%)",
                        affected_rows=[],
                        significance=0.6
                    ))
        
        # Pattern 3: Correlation-based patterns
        for corr in correlations[:3]:
            direction = "increases" if corr['direction'] == 'positive' else "decreases"
            patterns.append(Pattern(
                type='correlation',
                description=f"{corr['strength'].capitalize()} {corr['direction']} correlation: When '{corr['column1']}' increases, '{corr['column2']}' {direction} (r={corr['correlation']})",
                affected_rows=[],
                significance=abs(corr['correlation'])
            ))
        
        # Pattern 4: Distribution patterns
        for col, profile in profiles.items():
            if profile['numeric'] and profile['stats'].get('distribution'):
                dist = profile['stats']['distribution']
                if dist != 'normal':
                    patterns.append(Pattern(
                        type='distribution',
                        description=f"'{col}' shows {dist.replace('_', ' ')} distribution (skewness: {profile['stats'].get('skewness', 0)})",
                        affected_rows=[],
                        significance=0.5
                    ))
        
        # Pattern 5: Outlier patterns
        for col, profile in profiles.items():
            if profile['numeric'] and profile['stats']:
                q1 = profile['stats'].get('q1', 0)
                q3 = profile['stats'].get('q3', 0)
                iqr = profile['stats'].get('iqr', 0)
                if iqr > 0:
                    lower_bound = q1 - 1.5 * iqr
                    upper_bound = q3 + 1.5 * iqr
                    
                    outlier_count = 0
                    for row in rows:
                        try:
                            val = float(str(row.get(col, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                            if val < lower_bound or val > upper_bound:
                                outlier_count += 1
                        except:
                            pass
                    
                    if outlier_count > 0:
                        patterns.append(Pattern(
                            type='outliers',
                            description=f"'{col}' has {outlier_count} outlier values outside normal range [{round(lower_bound,2)}, {round(upper_bound,2)}]",
                            affected_rows=[],
                            significance=min(outlier_count / len(rows) + 0.3, 1.0)
                        ))
        
        return patterns[:10]
    
    def _detect_anomalies_advanced(self, rows: List[Dict], data_type: DataType,
                                   profiles: Dict[str, Dict]) -> List[Anomaly]:
        """
        Detect anomalies using statistical methods.
        """
        anomalies = []
        
        if len(rows) < 3:
            return anomalies
        
        # Z-score based anomaly detection for numeric columns
        for col, profile in profiles.items():
            if profile['numeric'] and profile['stats']:
                mean = profile['stats'].get('mean', 0)
                std_dev = profile['stats'].get('std_dev', 0)
                
                if std_dev > 0:
                    extreme_values = []
                    for i, row in enumerate(rows):
                        try:
                            val = float(str(row.get(col, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                            z_score = abs(val - mean) / std_dev
                            if z_score > 3:  # More than 3 standard deviations
                                extreme_values.append({'row': i, 'value': val, 'z_score': round(z_score, 2)})
                        except:
                            pass
                    
                    if extreme_values:
                        severity = Severity.CRITICAL if len(extreme_values) > 5 else (
                            Severity.HIGH if len(extreme_values) > 2 else Severity.MEDIUM
                        )
                        anomalies.append(Anomaly(
                            type='statistical_outlier',
                            severity=severity,
                            description=f"Found {len(extreme_values)} extreme values in '{col}' (>3 standard deviations from mean)",
                            affected_data={'column': col, 'outliers': extreme_values[:5]},
                            recommendation=f"Review extreme values in '{col}' for data accuracy"
                        ))
        
        # Data quality anomalies
        high_null_cols = [col for col, p in profiles.items() if p['null_count'] > len(rows) * 0.3]
        if high_null_cols:
            anomalies.append(Anomaly(
                type='data_quality',
                severity=Severity.HIGH,
                description=f"{len(high_null_cols)} columns have >30% missing values: {', '.join(high_null_cols[:3])}",
                affected_data={'columns': high_null_cols},
                recommendation="Investigate data collection process for missing values"
            ))
        
        # Suspicious value patterns
        for col, profile in profiles.items():
            if profile['numeric'] and profile['stats']:
                # Check for suspicious zero concentration
                zero_count = sum(1 for row in rows if str(row.get(col, '')).strip() in ['0', '0.0', '0.00'])
                if zero_count > len(rows) * 0.3 and zero_count < len(rows) * 0.9:
                    anomalies.append(Anomaly(
                        type='suspicious_pattern',
                        severity=Severity.MEDIUM,
                        description=f"'{col}' has {round(zero_count/len(rows)*100)}% zero values - may indicate data issues",
                        affected_data={'column': col, 'zero_count': zero_count},
                        recommendation=f"Verify if zero values in '{col}' are intentional"
                    ))
        
        return anomalies[:8]
    
    def _extract_hebrew_document_data(self, all_text: str, rows: List[Dict]) -> Dict[str, Any]:
        """
        Extract structured data from Hebrew insurance/financial documents.
        Uses the HebrewDocumentExtractor for pattern-based extraction.
        """
        extracted = HebrewDocumentExtractor.extract_fields(all_text)
        
        # Also scan all row values for additional data
        for row in rows:
            row_text = ' '.join(str(v) for v in row.values() if v)
            row_extracted = HebrewDocumentExtractor.extract_fields(row_text)
            for key, value in row_extracted.items():
                if key not in extracted:
                    extracted[key] = value
        
        # Analyze policy age if start date found
        if 'start_date' in extracted:
            age_info = HebrewDocumentExtractor.analyze_policy_age(extracted['start_date'])
            extracted.update(age_info)
        
        # Calculate coverage ratio if premium and cover found
        if 'premium' in extracted and 'cover_amount' in extracted:
            try:
                premium = float(extracted['premium']) if isinstance(extracted['premium'], str) else extracted['premium']
                cover = float(extracted['cover_amount']) if isinstance(extracted['cover_amount'], str) else extracted['cover_amount']
                ratio_info = HebrewDocumentExtractor.calculate_coverage_ratio(premium, cover)
                extracted.update(ratio_info)
            except (ValueError, TypeError):
                pass
        
        return extracted
    
    def _create_hebrew_document_factors(self, hebrew_extracted: Dict[str, Any], lang: str) -> List[Factor]:
        """
        Create analysis factors from extracted Hebrew document data.
        """
        factors = []
        is_hebrew = lang == 'hebrew'
        
        # Policy Information Factor
        policy_info = {}
        if 'policy_number' in hebrew_extracted:
            policy_info['מספר פוליסה' if is_hebrew else 'policy_number'] = hebrew_extracted['policy_number']
        if 'insurance_type' in hebrew_extracted:
            policy_info['סוג ביטוח' if is_hebrew else 'insurance_type'] = hebrew_extracted['insurance_type']
        if 'product_type' in hebrew_extracted:
            policy_info['סוג מוצר' if is_hebrew else 'product_type'] = hebrew_extracted['product_type']
        
        if policy_info:
            factors.append(Factor(
                name='פרטי פוליסה' if is_hebrew else 'Policy Details',
                value=policy_info,
                importance=0.95,
                category='hebrew_insurance'
            ))
        
        # Financial Details Factor
        financial_info = {}
        if 'premium' in hebrew_extracted:
            financial_info['פרמיה חודשית' if is_hebrew else 'monthly_premium'] = f"₪{hebrew_extracted['premium']}"
        if 'cover_amount' in hebrew_extracted:
            financial_info['סכום כיסוי' if is_hebrew else 'cover_amount'] = f"₪{hebrew_extracted['cover_amount']:,}" if isinstance(hebrew_extracted['cover_amount'], (int, float)) else f"₪{hebrew_extracted['cover_amount']}"
        if 'coverage_ratio' in hebrew_extracted:
            financial_info['יחס כיסוי/פרמיה' if is_hebrew else 'coverage_ratio'] = hebrew_extracted['coverage_ratio']
        if 'efficiency_rating' in hebrew_extracted:
            rating_map = {'excellent': 'מצוין', 'good': 'טוב', 'fair': 'סביר', 'review_recommended': 'מומלץ לבדיקה'}
            financial_info['דירוג יעילות' if is_hebrew else 'efficiency'] = rating_map.get(hebrew_extracted['efficiency_rating'], hebrew_extracted['efficiency_rating']) if is_hebrew else hebrew_extracted['efficiency_rating']
        
        if financial_info:
            factors.append(Factor(
                name='נתונים כספיים' if is_hebrew else 'Financial Details',
                value=financial_info,
                importance=0.9,
                category='hebrew_insurance'
            ))
        
        # Policy Timeline Factor
        timeline_info = {}
        if 'start_date' in hebrew_extracted:
            timeline_info['תאריך תחילה' if is_hebrew else 'start_date'] = hebrew_extracted.get('start_date', hebrew_extracted.get('start_date'))
        if 'policy_age_years' in hebrew_extracted:
            timeline_info['ותק הפוליסה' if is_hebrew else 'policy_age'] = f"{hebrew_extracted['policy_age_years']} שנים" if is_hebrew else f"{hebrew_extracted['policy_age_years']} years"
        if 'status' in hebrew_extracted:
            status_map = {'veteran': 'ותיקה', 'mature': 'בשלה', 'established': 'מבוססת', 'new': 'חדשה'}
            timeline_info['סטטוס' if is_hebrew else 'status'] = status_map.get(hebrew_extracted['status'], hebrew_extracted['status']) if is_hebrew else hebrew_extracted['status']
        
        if timeline_info:
            factors.append(Factor(
                name='ציר זמן הפוליסה' if is_hebrew else 'Policy Timeline',
                value=timeline_info,
                importance=0.85,
                category='hebrew_insurance'
            ))
        
        # Insured Person Factor
        person_info = {}
        if 'id_number' in hebrew_extracted:
            # Mask ID for privacy
            id_num = hebrew_extracted['id_number']
            masked_id = id_num[:2] + '*****' + id_num[-2:] if len(id_num) >= 4 else '***'
            person_info['ת.ז.' if is_hebrew else 'id'] = masked_id
        if 'insured_name' in hebrew_extracted:
            person_info['שם מבוטח' if is_hebrew else 'insured_name'] = hebrew_extracted['insured_name']
        if 'beneficiary' in hebrew_extracted:
            person_info['מוטב' if is_hebrew else 'beneficiary'] = hebrew_extracted['beneficiary']
        
        if person_info:
            factors.append(Factor(
                name='פרטי מבוטח' if is_hebrew else 'Insured Details',
                value=person_info,
                importance=0.8,
                category='hebrew_insurance'
            ))
        
        return factors
    
    def _find_hebrew_patterns(self, hebrew_extracted: Dict[str, Any]) -> List[Pattern]:
        """
        Find patterns specific to Hebrew insurance documents.
        """
        patterns = []
        
        # Policy age pattern
        if 'policy_age_years' in hebrew_extracted:
            age = hebrew_extracted['policy_age_years']
            if age > 20:
                patterns.append(Pattern(
                    type='policy_veteran',
                    description=f"פוליסה ותיקה מאוד ({age} שנים) - מומלץ לבדוק תנאים מול מוצרים חדשים בשוק",
                    affected_rows=[],
                    significance=0.9
                ))
            elif age > 10:
                patterns.append(Pattern(
                    type='policy_mature',
                    description=f"פוליסה בשלה ({age} שנים) - ייתכן שצברה ערכים או בונוסים",
                    affected_rows=[],
                    significance=0.7
                ))
        
        # Coverage efficiency pattern
        if 'efficiency_rating' in hebrew_extracted:
            rating = hebrew_extracted['efficiency_rating']
            if rating == 'review_recommended':
                patterns.append(Pattern(
                    type='coverage_efficiency',
                    description="יחס כיסוי/פרמיה נמוך - מומלץ לבחון חלופות בשוק",
                    affected_rows=[],
                    significance=0.85
                ))
            elif rating == 'excellent':
                patterns.append(Pattern(
                    type='coverage_efficiency',
                    description="יחס כיסוי/פרמיה מצוין - הפוליסה מספקת ערך טוב",
                    affected_rows=[],
                    significance=0.6
                ))
        
        # Product type patterns
        if 'product_type' in hebrew_extracted:
            ptype = hebrew_extracted['product_type']
            if ptype == 'pension':
                patterns.append(Pattern(
                    type='pension_product',
                    description="מוצר פנסיוני - יש לבדוק דמי ניהול ומסלול השקעה",
                    affected_rows=[],
                    significance=0.8
                ))
            elif ptype == 'life':
                patterns.append(Pattern(
                    type='life_insurance',
                    description="ביטוח חיים - יש לוודא שסכום הכיסוי מתאים לצרכים הנוכחיים",
                    affected_rows=[],
                    significance=0.75
                ))
        
        return patterns
    
    def _generate_domain_insights(self, data_type: DataType, profiles: Dict[str, Dict],
                                  rows: List[Dict], lang: str, 
                                  hebrew_extracted: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate domain-specific insights based on data type.
        Enhanced with Hebrew document extraction insights.
        """
        insights = {
            'domain': data_type.value,
            'language': lang,
            'key_findings': [],
            'recommendations': [],
            'metrics': {},
            'hebrew_data': hebrew_extracted or {}
        }
        
        # Find currency/amount columns
        currency_cols = [col for col, p in profiles.items() 
                        if p.get('semantic_type') == 'currency' or 
                        any(x in col.lower() for x in ['amount', 'premium', 'price', 'סכום', 'פרמיה', 'מחיר'])]
        
        is_hebrew = lang == 'hebrew'
        
        # Add Hebrew document insights if available
        if hebrew_extracted:
            # Policy details finding
            if 'policy_number' in hebrew_extracted:
                insights['key_findings'].append({
                    'type': 'policy_identification',
                    'finding': f"זוהתה פוליסה מס': {hebrew_extracted['policy_number']}" if is_hebrew else f"Identified policy: {hebrew_extracted['policy_number']}",
                    'detail': hebrew_extracted.get('insurance_type', hebrew_extracted.get('product_type', ''))
                })
            
            # Financial metrics finding
            if 'premium' in hebrew_extracted:
                premium = hebrew_extracted['premium']
                insights['key_findings'].append({
                    'type': 'premium_analysis',
                    'finding': f"פרמיה חודשית: ₪{premium}" if is_hebrew else f"Monthly Premium: ₪{premium}",
                    'detail': f"עלות שנתית: ₪{float(premium) * 12:,.0f}" if is_hebrew else f"Annual cost: ₪{float(premium) * 12:,.0f}"
                })
                insights['metrics']['premium'] = {
                    'monthly': premium,
                    'annual': float(premium) * 12 if isinstance(premium, (int, float)) else premium
                }
            
            if 'cover_amount' in hebrew_extracted:
                cover = hebrew_extracted['cover_amount']
                insights['key_findings'].append({
                    'type': 'coverage_analysis',
                    'finding': f"סכום כיסוי: ₪{cover:,}" if is_hebrew and isinstance(cover, (int, float)) else f"Cover Amount: ₪{cover}",
                    'detail': ''
                })
                insights['metrics']['cover_amount'] = cover
            
            # Policy age finding
            if 'policy_age_years' in hebrew_extracted:
                age = hebrew_extracted['policy_age_years']
                status = hebrew_extracted.get('status', '')
                status_hebrew = {'veteran': 'ותיקה', 'mature': 'בשלה', 'established': 'מבוססת', 'new': 'חדשה'}.get(status, status)
                insights['key_findings'].append({
                    'type': 'policy_age',
                    'finding': f"ותק הפוליסה: {age} שנים" if is_hebrew else f"Policy Age: {age} years",
                    'detail': f"סטטוס: {status_hebrew}" if is_hebrew else f"Status: {status}"
                })
            
            # Coverage efficiency finding
            if 'efficiency_rating' in hebrew_extracted:
                rating = hebrew_extracted['efficiency_rating']
                rating_hebrew = {'excellent': 'מצוין', 'good': 'טוב', 'fair': 'סביר', 'review_recommended': 'מומלץ לבדיקה'}.get(rating, rating)
                insights['key_findings'].append({
                    'type': 'efficiency_rating',
                    'finding': f"דירוג יעילות: {rating_hebrew}" if is_hebrew else f"Efficiency Rating: {rating}",
                    'detail': f"יחס כיסוי/פרמיה: {hebrew_extracted.get('coverage_ratio', 'N/A')}" if is_hebrew else f"Coverage ratio: {hebrew_extracted.get('coverage_ratio', 'N/A')}"
                })
            
            # Hebrew-specific recommendations
            if is_hebrew:
                if hebrew_extracted.get('policy_age_years', 0) > 15:
                    insights['recommendations'].append('פוליסה ותיקה - מומלץ לבדוק האם התנאים עדיין תחרותיים')
                if hebrew_extracted.get('efficiency_rating') == 'review_recommended':
                    insights['recommendations'].append('יחס כיסוי/פרמיה נמוך - כדאי לקבל הצעות מחיר נוספות')
                if 'pension' in str(hebrew_extracted.get('product_type', '')):
                    insights['recommendations'].append('מוצר פנסיוני - בדוק דמי ניהול ומסלולי השקעה')
                insights['recommendations'].append('ודא שהמוטבים מעודכנים')
                insights['recommendations'].append('בדוק התאמת סכומי הכיסוי לצרכים הנוכחיים')
        
        if data_type == DataType.INSURANCE:
            insights['key_findings'].append({
                'type': 'domain_classification',
                'finding': 'ניתוח נתוני ביטוח' if is_hebrew else 'Insurance data analysis',
                'detail': f'{len(rows)} רשומות נותחו' if is_hebrew else f'{len(rows)} records analyzed'
            })
            
            # Insurance-specific metrics
            for col in currency_cols[:2]:
                if col in profiles and profiles[col]['stats']:
                    stats = profiles[col]['stats']
                    insights['metrics'][col] = {
                        'total': stats.get('sum', 0),
                        'average': stats.get('mean', 0),
                        'range': f"{stats.get('min', 0)} - {stats.get('max', 0)}"
                    }
            
            if is_hebrew and not hebrew_extracted:
                insights['recommendations'].append('בדוק כיסויים ביטוחיים מול צרכים')
                insights['recommendations'].append('השווה פרמיות לממוצע בשוק')
            elif not is_hebrew:
                insights['recommendations'].append('Review coverage adequacy against needs')
                insights['recommendations'].append('Compare premiums to market average')
                
        elif data_type == DataType.INVESTMENT:
            insights['key_findings'].append({
                'type': 'domain_classification',
                'finding': 'ניתוח תיק השקעות' if is_hebrew else 'Investment portfolio analysis',
                'detail': f'{len(rows)} נכסים נותחו' if is_hebrew else f'{len(rows)} assets analyzed'
            })
            
            if is_hebrew:
                insights['recommendations'].append('בדוק פיזור התיק')
                insights['recommendations'].append('נתח יחס תשואה/סיכון')
            else:
                insights['recommendations'].append('Review portfolio diversification')
                insights['recommendations'].append('Analyze return/risk ratio')
                
        elif data_type == DataType.SAVINGS:
            insights['key_findings'].append({
                'type': 'domain_classification',
                'finding': 'ניתוח חיסכון' if is_hebrew else 'Savings analysis',
                'detail': f'{len(rows)} רשומות' if is_hebrew else f'{len(rows)} records'
            })
            
        elif data_type == DataType.RISK:
            insights['key_findings'].append({
                'type': 'domain_classification',
                'finding': 'הערכת סיכונים' if is_hebrew else 'Risk assessment',
                'detail': f'{len(rows)} גורמי סיכון נותחו' if is_hebrew else f'{len(rows)} risk factors analyzed'
            })
        
        # Add data quality insight
        complete_cols = sum(1 for p in profiles.values() if p['null_count'] == 0)
        completeness = round(complete_cols / max(len(profiles), 1) * 100, 1)
        
        insights['key_findings'].append({
            'type': 'data_quality',
            'finding': f'שלמות נתונים: {completeness}%' if is_hebrew else f'Data completeness: {completeness}%',
            'detail': f'{complete_cols}/{len(profiles)} שדות מלאים' if is_hebrew else f'{complete_cols}/{len(profiles)} fields complete'
        })
        
        return insights
    
    def _calculate_risk_score_advanced(self, factors: List[Factor], patterns: List[Pattern],
                                       anomalies: List[Anomaly], correlations: List[Dict],
                                       domain_insights: Dict) -> float:
        """
        Calculate comprehensive risk score using multiple factors.
        """
        base_score = 35  # Start at low-medium risk
        
        # Factor-based adjustment
        for factor in factors:
            if factor.category == 'statistical':
                # High variance increases risk
                if isinstance(factor.value, dict) and factor.value.get('std_dev', 0) > factor.value.get('mean', 1) * 0.5:
                    base_score += 5
        
        # Pattern-based adjustment
        for pattern in patterns:
            if pattern.type == 'data_quality':
                base_score += 10
            elif pattern.type == 'outliers':
                base_score += pattern.significance * 8
            elif pattern.type == 'correlation':
                # Negative correlations in financial data can be risk indicators
                pass
        
        # Anomaly-based adjustment
        for anomaly in anomalies:
            if anomaly.severity == Severity.CRITICAL:
                base_score += 15
            elif anomaly.severity == Severity.HIGH:
                base_score += 10
            elif anomaly.severity == Severity.MEDIUM:
                base_score += 5
            else:
                base_score += 2
        
        # Domain-specific adjustment
        domain = domain_insights.get('domain', 'unknown')
        if domain == 'risk':
            base_score += 15  # Risk data inherently higher
        
        return min(max(base_score, 0), 100)
    
    def _generate_summary_advanced(self, lang: str, data_type: DataType, row_count: int,
                                   factors: List[Factor], risk_score: float,
                                   profiles: Dict[str, Dict], domain_insights: Dict) -> str:
        """
        Generate comprehensive language-aware summary.
        """
        risk_level = 'נמוך' if lang == 'hebrew' else 'Low'
        if risk_score >= 60:
            risk_level = 'גבוה' if lang == 'hebrew' else 'High'
        elif risk_score >= 30:
            risk_level = 'בינוני' if lang == 'hebrew' else 'Medium'
        
        numeric_cols = sum(1 for p in profiles.values() if p['numeric'])
        cat_cols = len(profiles) - numeric_cols
        
        if lang == 'hebrew':
            type_names = {
                'insurance': 'ביטוח',
                'investment': 'השקעות',
                'risk': 'סיכונים',
                'savings': 'חיסכון',
                'mixed': 'מעורב',
                'unknown': 'כללי'
            }
            type_name = type_names.get(data_type.value, 'נתונים')
            
            summary = f"""ניתוח AI מקיף של נתוני {type_name}:

📊 סטטיסטיקה:
• {row_count} רשומות נותחו
• {len(profiles)} שדות זוהו ({numeric_cols} מספריים, {cat_cols} קטגוריים)
• {len(factors)} גורמים מרכזיים חולצו

🎯 הערכת סיכון: {risk_score:.0f}/100 ({risk_level})

📈 תובנות עיקריות:
"""
            for finding in domain_insights.get('key_findings', [])[:3]:
                summary += f"• {finding['finding']}: {finding['detail']}\n"
            
        else:
            type_names = {
                'insurance': 'Insurance',
                'investment': 'Investment',
                'risk': 'Risk',
                'savings': 'Savings',
                'mixed': 'Mixed',
                'unknown': 'General'
            }
            type_name = type_names.get(data_type.value, 'Data')
            
            summary = f"""Comprehensive AI Analysis of {type_name} Data:

📊 Statistics:
• {row_count} records analyzed
• {len(profiles)} fields identified ({numeric_cols} numeric, {cat_cols} categorical)
• {len(factors)} key factors extracted

🎯 Risk Assessment: {risk_score:.0f}/100 ({risk_level})

📈 Key Insights:
"""
            for finding in domain_insights.get('key_findings', [])[:3]:
                summary += f"• {finding['finding']}: {finding['detail']}\n"
        
        return summary
    
    def _extract_key_metrics_advanced(self, rows: List[Dict], columns: List[str],
                                      data_type: DataType, profiles: Dict[str, Dict],
                                      correlations: List[Dict], domain_insights: Dict) -> Dict[str, Any]:
        """
        Extract comprehensive key metrics including BI indicators.
        """
        metrics = {
            'total_records': len(rows),
            'total_columns': len(columns),
            'data_type': data_type.value,
            'numeric_columns': sum(1 for p in profiles.values() if p['numeric']),
            'categorical_columns': sum(1 for p in profiles.values() if not p['numeric']),
            'correlation_count': len(correlations)
        }
        
        # Data quality metrics
        total_cells = len(rows) * len(columns)
        null_cells = sum(p['null_count'] for p in profiles.values())
        metrics['data_completeness'] = round((1 - null_cells / max(total_cells, 1)) * 100, 1)
        
        # Add column-specific metrics
        for col, profile in list(profiles.items())[:10]:
            if profile['numeric'] and profile['stats']:
                metrics[f'{col}_total'] = profile['stats'].get('sum', 0)
                metrics[f'{col}_avg'] = profile['stats'].get('mean', 0)
                metrics[f'{col}_min'] = profile['stats'].get('min', 0)
                metrics[f'{col}_max'] = profile['stats'].get('max', 0)
        
        # Add domain metrics
        metrics['domain_metrics'] = domain_insights.get('metrics', {})
        
        # Add top correlations
        if correlations:
            metrics['top_correlation'] = {
                'fields': f"{correlations[0]['column1']} ↔ {correlations[0]['column2']}",
                'strength': correlations[0]['correlation']
            }
        
        return metrics
    
    def _extract_factors(self, columns: List[str], rows: List[Dict], data_type: DataType) -> List[Factor]:
        """Extract key factors from the data"""
        factors = []
        
        # Identify numeric columns
        numeric_cols = []
        for col in columns:
            if rows:
                sample = rows[0].get(col, '')
                try:
                    float(str(sample).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                    numeric_cols.append(col)
                except (ValueError, TypeError):
                    pass
        
        # Calculate statistics for numeric columns
        for col in numeric_cols[:5]:  # Top 5 numeric columns
            values = []
            for row in rows:
                try:
                    val = float(str(row.get(col, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                    values.append(val)
                except (ValueError, TypeError):
                    continue
            
            if values:
                avg = sum(values) / len(values)
                total = sum(values)
                factors.append(Factor(
                    name=col,
                    value={'average': round(avg, 2), 'total': round(total, 2), 'count': len(values)},
                    importance=0.7 if data_type != DataType.UNKNOWN else 0.5,
                    category='numeric_metric'
                ))
        
        # Data type specific factors
        if data_type == DataType.INSURANCE:
            factors.append(Factor(
                name='Coverage Analysis',
                value={'records': len(rows), 'type': 'insurance'},
                importance=0.9,
                category='insurance'
            ))
        elif data_type == DataType.INVESTMENT:
            factors.append(Factor(
                name='Portfolio Analysis',
                value={'records': len(rows), 'type': 'investment'},
                importance=0.9,
                category='investment'
            ))
        elif data_type == DataType.RISK:
            factors.append(Factor(
                name='Risk Assessment',
                value={'records': len(rows), 'type': 'risk'},
                importance=0.95,
                category='risk'
            ))
        
        return factors
    
    def _find_patterns(self, rows: List[Dict], data_type: DataType) -> List[Pattern]:
        """Find patterns in the data"""
        patterns = []
        
        if len(rows) < 2:
            return patterns
        
        # Look for duplicate values
        for key in list(rows[0].keys())[:5]:
            values = [row.get(key) for row in rows if row.get(key)]
            unique_values = set(values)
            if len(values) > len(unique_values):
                dup_count = len(values) - len(unique_values)
                patterns.append(Pattern(
                    type='duplicate_values',
                    description=f'Found {dup_count} duplicate values in column "{key}"',
                    affected_rows=[i for i, v in enumerate(values) if values.count(v) > 1],
                    significance=0.6
                ))
        
        # Look for empty values
        empty_counts = {}
        for key in rows[0].keys():
            empty_count = sum(1 for row in rows if not row.get(key))
            if empty_count > 0:
                empty_counts[key] = empty_count
        
        if empty_counts:
            max_empty = max(empty_counts.values())
            if max_empty > len(rows) * 0.1:  # More than 10% empty
                patterns.append(Pattern(
                    type='missing_data',
                    description=f'Found columns with missing data: {list(empty_counts.keys())}',
                    affected_rows=list(range(len(rows))),
                    significance=0.7
                ))
        
        return patterns
    
    def _detect_anomalies(self, rows: List[Dict], data_type: DataType) -> List[Anomaly]:
        """Detect anomalies in the data"""
        anomalies = []
        
        if len(rows) < 3:
            return anomalies
        
        # Look for outliers in numeric columns
        for key in list(rows[0].keys())[:5]:
            values = []
            for row in rows:
                try:
                    val = float(str(row.get(key, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                    values.append(val)
                except (ValueError, TypeError):
                    continue
            
            if len(values) >= 3:
                avg = sum(values) / len(values)
                std_dev = (sum((x - avg) ** 2 for x in values) / len(values)) ** 0.5
                
                if std_dev > 0:
                    outliers = [v for v in values if abs(v - avg) > 2 * std_dev]
                    if outliers:
                        anomalies.append(Anomaly(
                            type='outlier',
                            severity=Severity.MEDIUM if len(outliers) < 3 else Severity.HIGH,
                            description=f'Found {len(outliers)} outlier values in column "{key}"',
                            affected_data={'column': key, 'outliers': outliers[:5]},
                            recommendation=f'Review the extreme values in {key} for accuracy'
                        ))
        
        return anomalies
    
    def _calculate_risk_score(self, factors: List[Factor], patterns: List[Pattern], anomalies: List[Anomaly]) -> float:
        """Calculate overall risk score (0-100)"""
        base_score = 50  # Start at medium risk
        
        # Adjust based on anomalies
        for anomaly in anomalies:
            if anomaly.severity == Severity.CRITICAL:
                base_score += 20
            elif anomaly.severity == Severity.HIGH:
                base_score += 10
            elif anomaly.severity == Severity.MEDIUM:
                base_score += 5
            else:
                base_score += 2
        
        # Adjust based on patterns
        for pattern in patterns:
            if pattern.type == 'missing_data':
                base_score += pattern.significance * 15
            elif pattern.type == 'duplicate_values':
                base_score += pattern.significance * 5
        
        # Adjust based on factors
        for factor in factors:
            if factor.category == 'risk':
                base_score += factor.importance * 10
        
        return min(max(base_score, 0), 100)
    
    def _generate_summary(self, lang: str, data_type: DataType, row_count: int, 
                          factors: List[Factor], risk_score: float) -> str:
        """Generate a summary of the analysis"""
        summaries = {
            'hebrew': {
                'insurance': f'ניתוח נתוני ביטוח: {row_count} רשומות נותחו. ציון סיכון: {risk_score:.1f}/100.',
                'investment': f'ניתוח תיק השקעות: {row_count} רשומות נותחו. ציון סיכון: {risk_score:.1f}/100.',
                'risk': f'הערכת סיכונים: {row_count} רשומות נותחו. ציון סיכון כולל: {risk_score:.1f}/100.',
                'savings': f'ניתוח חיסכון: {row_count} רשומות נותחו. ציון סיכון: {risk_score:.1f}/100.',
                'default': f'ניתוח נתונים: {row_count} רשומות נותחו. ציון סיכון: {risk_score:.1f}/100.'
            },
            'english': {
                'insurance': f'Insurance data analysis: {row_count} records analyzed. Risk score: {risk_score:.1f}/100.',
                'investment': f'Investment portfolio analysis: {row_count} records analyzed. Risk score: {risk_score:.1f}/100.',
                'risk': f'Risk assessment analysis: {row_count} records analyzed. Overall risk score: {risk_score:.1f}/100.',
                'savings': f'Savings analysis: {row_count} records analyzed. Risk score: {risk_score:.1f}/100.',
                'default': f'Data analysis: {row_count} records analyzed. Risk score: {risk_score:.1f}/100.'
            }
        }
        
        lang_summaries = summaries.get(lang, summaries['english'])
        return lang_summaries.get(data_type.value, lang_summaries['default'])
    
    def _extract_key_metrics(self, rows: List[Dict], columns: List[str], data_type: DataType) -> Dict[str, Any]:
        """Extract key metrics from the data"""
        metrics = {
            'total_records': len(rows),
            'columns_count': len(columns),
            'data_type': data_type.value
        }
        
        # Calculate totals for numeric columns
        for col in columns[:10]:
            values = []
            for row in rows:
                try:
                    val = float(str(row.get(col, 0)).replace(',', '').replace('₪', '').replace('$', '').replace('€', ''))
                    values.append(val)
                except (ValueError, TypeError):
                    continue
            
            if values:
                metrics[f'{col}_total'] = round(sum(values), 2)
                metrics[f'{col}_avg'] = round(sum(values) / len(values), 2)
                metrics[f'{col}_min'] = round(min(values), 2)
                metrics[f'{col}_max'] = round(max(values), 2)
        
        return metrics
    
    def generate_report(self, analysis_id: str, language: str = None) -> GeneratedReport:
        """Generate a comprehensive report from analysis results"""
        if analysis_id not in self.analyses:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        analysis = self.analyses[analysis_id]
        lang = language or analysis.language
        
        report_id = f"RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(1000, 9999)}"
        
        # Retrieve original document data for content analysis
        doc_data = None
        pension_data = None
        pension_report = None
        if analysis.document_id in self.documents:
            doc = self.documents[analysis.document_id]
            doc_data = doc.get('parsed_data', {})
            # Check for pension data from ZIP files
            if doc_data:
                pension_data = doc_data.get('pension_data')
                pension_report = doc_data.get('pension_report')
        
        # Build affiliated savings/coverage/ID summary for sections/charts/downloads
        affiliated_summary = self._extract_savings_cover_id_summary(doc_data, pension_data)

        # Generate sections based on data type (now with original data and pension data)
        sections = self._generate_sections(
            analysis,
            lang,
            doc_data,
            pension_data,
            pension_report,
            affiliated_summary
        )
        
        # Generate charts - pass pension_data and affiliated summary for specialized charts
        charts = self._generate_charts(analysis, pension_data, doc_data, affiliated_summary)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(analysis, lang)
        
        # Determine report title
        titles = {
            'hebrew': {
                'insurance': 'דו״ח ניתוח ביטוח',
                'investment': 'דו״ח ניתוח השקעות',
                'risk': 'דו״ח הערכת סיכונים',
                'savings': 'דו״ח ניתוח חיסכון',
                'default': 'דו״ח ניתוח נתונים'
            },
            'english': {
                'insurance': 'Insurance Analysis Report',
                'investment': 'Investment Analysis Report',
                'risk': 'Risk Assessment Report',
                'savings': 'Savings Analysis Report',
                'default': 'Data Analysis Report'
            }
        }
        
        lang_titles = titles.get(lang, titles['english'])
        title = lang_titles.get(analysis.data_classification.value, lang_titles['default'])
        
        report = GeneratedReport(
            id=report_id,
            analysis_id=analysis_id,
            report_type=analysis.data_classification.value,
            language=lang,
            title=title,
            sections=sections,
            charts=charts,
            recommendations=recommendations,
            generated_at=datetime.now().isoformat(),
            metadata={
                'document_id': analysis.document_id,
                'risk_score': analysis.risk_score,
                'confidence': analysis.confidence,
                'processing_time_ms': analysis.processing_time_ms,
                'pension_data': pension_data if pension_data else None,
                'is_pension_data': pension_data is not None or pension_report is not None,
                'affiliation_snapshot': self._build_affiliation_snapshot_metadata(),
                'savings_cover_id_summary': affiliated_summary,
                'report_model': self._get_report_model_metadata(),
            }
        )
        
        self.reports[report_id] = report
        
        # Auto-save for persistence
        self.save_data()

        _risk_report_audit('risk_report_generated', report_id, {
            'analysis_id': analysis_id,
            'document_id': getattr(analysis, 'document_id', None),
            'language': lang,
        })
        
        return report
    
    def _generate_sections(self, analysis: AnalysisResult, lang: str, 
                          doc_data: Dict[str, Any] = None,
                          pension_data: Dict[str, Any] = None,
                          pension_report: str = None,
                          savings_cover_id_summary: Dict[str, Any] = None) -> List[ReportSection]:
        """Generate comprehensive report sections with AI/BI insights and actual data content"""
        sections = []
        is_hebrew = lang == 'hebrew'
        
        # 1. Executive Summary
        sections.append(ReportSection(
            title='תקציר מנהלים' if is_hebrew else 'Executive Summary',
            content=analysis.summary,
            order=1
        ))
        
        # 2. PENSION DATA SECTION - Display pension report from PensionDataAgent
        if pension_data or pension_report:
            pension_section = self._generate_pension_section(pension_data, pension_report, is_hebrew)
            if pension_section:
                sections.append(ReportSection(
                    title='דו״ח ניתוח פנסיה וביטוח' if is_hebrew else 'Pension & Insurance Analysis Report',
                    content=pension_section,
                    order=2
                ))
            sections.extend(self._build_pension_affiliated_sections(pension_data, is_hebrew))
        
        # 3. ACTUAL DATA CONTENT SECTION - Show extracted data from the files
        if doc_data:
            data_content_section = self._generate_data_content_section(doc_data, analysis, is_hebrew)
            if data_content_section:
                sections.append(ReportSection(
                    title='תוכן הנתונים שהועלו' if is_hebrew else 'Uploaded Data Content',
                    content=data_content_section,
                    order=3
                ))

        # 3.5 Savings/Cover/ID affiliation section (table-oriented summary)
        if savings_cover_id_summary is None:
            savings_cover_id_summary = self._extract_savings_cover_id_summary(doc_data, pension_data)
        affiliated_summary_section = self._build_savings_cover_id_section(savings_cover_id_summary, is_hebrew)
        if affiliated_summary_section:
            sections.append(affiliated_summary_section)
        
        # 3. Hebrew Insurance Details (if extracted)
        hebrew_factors = [f for f in analysis.extracted_factors if f.category == 'hebrew_insurance']
        if hebrew_factors:
            hebrew_content = self._generate_hebrew_insurance_section(hebrew_factors, is_hebrew)
            sections.append(ReportSection(
                title='פרטי פוליסת ביטוח' if is_hebrew else 'Insurance Policy Details',
                content=hebrew_content,
                order=3
            ))
        
        # 4. Data Profile Overview
        total_records = analysis.key_metrics.get('total_records', 0)
        numeric_cols = analysis.key_metrics.get('numeric_columns', 0)
        cat_cols = analysis.key_metrics.get('categorical_columns', 0)
        completeness = analysis.key_metrics.get('data_completeness', 100)
        
        if is_hebrew:
            profile_content = f"""📊 פרופיל הנתונים:

• סה"כ רשומות: {total_records}
• שדות מספריים: {numeric_cols}
• שדות קטגוריים: {cat_cols}
• שלמות נתונים: {completeness}%
• סוג נתונים: {analysis.data_classification.value}
• שפה: {analysis.language_name}
• רמת ביטחון: {analysis.confidence:.0%}"""
        else:
            profile_content = f"""📊 Data Profile:

• Total Records: {total_records}
• Numeric Fields: {numeric_cols}
• Categorical Fields: {cat_cols}
• Data Completeness: {completeness}%
• Data Type: {analysis.data_classification.value}
• Language: {analysis.language_name}
• Confidence Level: {analysis.confidence:.0%}"""
        
        sections.append(ReportSection(
            title='פרופיל נתונים' if is_hebrew else 'Data Profile',
            content=profile_content,
            order=2
        ))
        
        # 3. Statistical Analysis (BI Metrics)
        # SKIP for pension data - the pension report already shows meaningful data clearly
        # Statistical analysis of IDs/policy numbers is meaningless
        if not pension_report:  # Only show statistical analysis for non-pension data
            stat_factors = [f for f in analysis.extracted_factors if f.category == 'statistical']
            if stat_factors:
                if is_hebrew:
                    stats_lines = ['📈 ניתוח סטטיסטי מפורט:\n']
                    for f in stat_factors[:8]:
                        if isinstance(f.value, dict):
                            stats_lines.append(f"🔹 {f.name}:")
                            stats_lines.append(f"   • ממוצע: {f.value.get('mean', 'N/A')}")
                            stats_lines.append(f"   • חציון: {f.value.get('median', 'N/A')}")
                            stats_lines.append(f"   • טווח: {f.value.get('range', 'N/A')}")
                            stats_lines.append(f"   • סטיית תקן: {f.value.get('std_dev', 'N/A')}")
                            stats_lines.append(f"   • התפלגות: {f.value.get('distribution', 'N/A')}")
                            stats_lines.append("")
                else:
                    stats_lines = ['📈 Detailed Statistical Analysis:\n']
                    for f in stat_factors[:8]:
                        if isinstance(f.value, dict):
                            stats_lines.append(f"🔹 {f.name}:")
                            stats_lines.append(f"   • Mean: {f.value.get('mean', 'N/A')}")
                            stats_lines.append(f"   • Median: {f.value.get('median', 'N/A')}")
                            stats_lines.append(f"   • Range: {f.value.get('range', 'N/A')}")
                            stats_lines.append(f"   • Std Dev: {f.value.get('std_dev', 'N/A')}")
                            stats_lines.append(f"   • Distribution: {f.value.get('distribution', 'N/A')}")
                            stats_lines.append("")
                
                sections.append(ReportSection(
                    title='ניתוח סטטיסטי' if is_hebrew else 'Statistical Analysis',
                    content='\n'.join(stats_lines),
                    order=3
                ))
        
        # 4. Correlation Insights
        top_corr = analysis.key_metrics.get('top_correlation')
        if top_corr:
            if is_hebrew:
                corr_content = f"""🔗 מתאמים שזוהו:

• הקשר החזק ביותר: {top_corr.get('fields', '')}
• עוצמת המתאם: {top_corr.get('strength', 0)}

💡 משמעות: מתאמים אלו מצביעים על קשרים פוטנציאליים בין משתנים שיש לקחת בחשבון בניתוח."""
            else:
                corr_content = f"""🔗 Correlations Discovered:

• Strongest Relationship: {top_corr.get('fields', '')}
• Correlation Strength: {top_corr.get('strength', 0)}

💡 Significance: These correlations indicate potential relationships between variables that should be considered in the analysis."""
            
            sections.append(ReportSection(
                title='ניתוח מתאמים' if is_hebrew else 'Correlation Analysis',
                content=corr_content,
                order=4
            ))
        
        # 5. Patterns & Trends
        if analysis.patterns_found:
            if is_hebrew:
                patterns_lines = ['🔍 דפוסים ומגמות שזוהו:\n']
                for i, p in enumerate(analysis.patterns_found, 1):
                    patterns_lines.append(f"{i}. [{p.type}] {p.description}")
                    patterns_lines.append(f"   משמעות: {p.significance:.0%}")
                    patterns_lines.append("")
            else:
                patterns_lines = ['🔍 Identified Patterns & Trends:\n']
                for i, p in enumerate(analysis.patterns_found, 1):
                    patterns_lines.append(f"{i}. [{p.type}] {p.description}")
                    patterns_lines.append(f"   Significance: {p.significance:.0%}")
                    patterns_lines.append("")
            
            sections.append(ReportSection(
                title='דפוסים ומגמות' if is_hebrew else 'Patterns & Trends',
                content='\n'.join(patterns_lines),
                order=5
            ))
        
        # 6. Anomalies & Warnings
        if analysis.anomalies:
            if is_hebrew:
                anomaly_lines = ['⚠️ חריגות ואזהרות:\n']
                severity_map = {'critical': '🔴 קריטי', 'high': '🟠 גבוה', 'medium': '🟡 בינוני', 'low': '🟢 נמוך'}
                for a in analysis.anomalies:
                    sev_label = severity_map.get(a.severity.value, a.severity.value)
                    anomaly_lines.append(f"• {sev_label}: {a.description}")
                    anomaly_lines.append(f"  📋 המלצה: {a.recommendation}")
                    anomaly_lines.append("")
            else:
                anomaly_lines = ['⚠️ Anomalies & Warnings:\n']
                severity_map = {'critical': '🔴 Critical', 'high': '🟠 High', 'medium': '🟡 Medium', 'low': '🟢 Low'}
                for a in analysis.anomalies:
                    sev_label = severity_map.get(a.severity.value, a.severity.value)
                    anomaly_lines.append(f"• {sev_label}: {a.description}")
                    anomaly_lines.append(f"  📋 Recommendation: {a.recommendation}")
                    anomaly_lines.append("")
            
            sections.append(ReportSection(
                title='חריגות ואזהרות' if is_hebrew else 'Anomalies & Warnings',
                content='\n'.join(anomaly_lines),
                order=6
            ))
        
        # 7. Risk Assessment
        risk_score = analysis.risk_score
        if risk_score < 30:
            risk_level = 'נמוך' if is_hebrew else 'Low'
            risk_color = '🟢'
            risk_desc = 'הנתונים מצביעים על רמת סיכון נמוכה' if is_hebrew else 'Data indicates low risk level'
        elif risk_score < 60:
            risk_level = 'בינוני' if is_hebrew else 'Medium'
            risk_color = '🟡'
            risk_desc = 'יש לשים לב לגורמי סיכון מסוימים' if is_hebrew else 'Some risk factors require attention'
        else:
            risk_level = 'גבוה' if is_hebrew else 'High'
            risk_color = '🔴'
            risk_desc = 'נדרשת בדיקה מעמיקה של גורמי הסיכון' if is_hebrew else 'In-depth review of risk factors required'
        
        if is_hebrew:
            risk_content = f"""🎯 הערכת סיכון כוללת:

{risk_color} ציון סיכון: {risk_score:.0f}/100
📊 רמת סיכון: {risk_level}

{risk_desc}

גורמים המשפיעים על הציון:
• מספר חריגות: {len(analysis.anomalies)}
• דפוסים חריגים: {len([p for p in analysis.patterns_found if p.significance > 0.5])}
• שלמות נתונים: {completeness}%"""
        else:
            risk_content = f"""🎯 Overall Risk Assessment:

{risk_color} Risk Score: {risk_score:.0f}/100
📊 Risk Level: {risk_level}

{risk_desc}

Factors Affecting Score:
• Number of anomalies: {len(analysis.anomalies)}
• Unusual patterns: {len([p for p in analysis.patterns_found if p.significance > 0.5])}
• Data completeness: {completeness}%"""
        
        sections.append(ReportSection(
            title='הערכת סיכון' if is_hebrew else 'Risk Assessment',
            content=risk_content,
            order=7
        ))
        
        # 8. Key Metrics Table
        metrics_items = []
        for k, v in list(analysis.key_metrics.items())[:15]:
            if not k.startswith('domain_') and k != 'top_correlation':
                if isinstance(v, float):
                    metrics_items.append(f"• {k}: {v:.2f}")
                else:
                    metrics_items.append(f"• {k}: {v}")
        
        sections.append(ReportSection(
            title='מדדים מרכזיים' if is_hebrew else 'Key Metrics',
            content='\n'.join(metrics_items),
            data_table=analysis.key_metrics,
            order=8
        ))

        # 9. Affiliation Snapshot (Mislaka schema codes)
        affiliation_section = self._build_affiliation_mapping_section(is_hebrew)
        if affiliation_section:
            sections.append(affiliation_section)
        
        return sections

    def _build_affiliation_snapshot_metadata(self) -> Dict[str, Any]:
        """Build compact affiliation metadata without mutating source mappings."""
        try:
            from services.pension_data_agent import MislakaSchemaMapping

            return {
                'interface_codes': len(MislakaSchemaMapping.INTERFACE_CODES),
                'product_types': len(MislakaSchemaMapping.PRODUCT_TYPE_CODES),
                'entity_types': len(MislakaSchemaMapping.ENTITY_TYPE_CODES),
                'status_codes': len(MislakaSchemaMapping.STATUS_CODES),
                'id_types': len(MislakaSchemaMapping.ID_TYPE_CODES),
                'environment_codes': len(MislakaSchemaMapping.ENVIRONMENT_CODES),
            }
        except Exception:
            return {}

    def _get_report_model_metadata(self) -> Optional[Dict[str, Any]]:
        """Include the Swiftness report model section keys for frontend structuring."""
        try:
            from services.swiftness_data_service import get_swiftness_data_service
            svc = get_swiftness_data_service()
            model = svc.get_report_model()
            return {
                'section_keys': [s['key'] for s in model.get('sections', [])],
                'model_version': model.get('metadata', {}).get('model_version'),
                'total_sections': model.get('metadata', {}).get('total_sections'),
            }
        except Exception:
            return None

    def _build_pension_affiliated_sections(self, pension_data: Dict[str, Any], is_hebrew: bool) -> List[ReportSection]:
        """Build table-oriented sections aligned with the Nituach Tik report model."""
        sections: List[ReportSection] = []
        if not pension_data:
            return sections

        accounts = pension_data.get('accounts', []) or []
        contributions = pension_data.get('contributions', []) or []
        totals = pension_data.get('totals', {}) or {}
        employers = pension_data.get('employers', []) or []
        client = pension_data.get('client', {}) or {}
        if isinstance(client, list):
            client = client[0] if client else {}
        if isinstance(client, dict) and client:
            client = self._normalize_client_profile_fields(client)
            profile_rows = [{
                'שם מלא' if is_hebrew else 'Full Name': client.get('full_name', client.get('client_name', '')),
                'מזהה לקוח' if is_hebrew else 'Customer ID': client.get('id_number', ''),
                'תאריך לידה' if is_hebrew else 'Birth Date': client.get('birth_date', ''),
                'פורמט גולמי' if is_hebrew else 'Birth Date Raw': client.get('birth_date_raw', ''),
                'אימות מזהה' if is_hebrew else 'ID Validation': (
                    'תקין' if client.get('id_israeli_valid') else 'דורש בדיקה'
                ) if is_hebrew else (
                    'Valid' if client.get('id_israeli_valid') else 'Needs review'
                ),
            }]
            sections.append(ReportSection(
                title='פרופיל לקוח (שיוך)' if is_hebrew else 'Customer Profile (Affiliated)',
                content='פרטי לקוח מאומתים מתוך קבצים מסונפים.' if is_hebrew else 'Customer identity fields validated from affiliated files.',
                data_table={
                    'columns': list(profile_rows[0].keys()),
                    'rows': profile_rows
                },
                order=2
            ))

        if accounts:
            status_rows = []
            for acct in accounts[:80]:
                status_rows.append({
                    'מספר פוליסה' if is_hebrew else 'Policy Number': acct.get('policy_number', ''),
                    'יצרן' if is_hebrew else 'Provider': acct.get('provider', ''),
                    'סוג מוצר' if is_hebrew else 'Product Type': acct.get('product_type_name', acct.get('product_type', '')),
                    'סטטוס' if is_hebrew else 'Status': acct.get('status', acct.get('status_en', '')),
                    'יתרה כוללת' if is_hebrew else 'Total Balance': acct.get('total_balance', 0),
                    'פיצויים' if is_hebrew else 'Severance': acct.get('severance_balance', 0),
                    'מעסיק' if is_hebrew else 'Employer': acct.get('employer_name', ''),
                    'סעיף 14' if is_hebrew else 'Section 14': ('כן' if acct.get('section14') else 'לא') if is_hebrew else ('Yes' if acct.get('section14') else 'No'),
                })

            sections.append(ReportSection(
                title='סטטוס פוליסות (טבלת שיוכים)' if is_hebrew else 'Policy Status (Affiliation Table)',
                content='מבט טבלאי על פוליסות לפי שיוכי מסלקה.' if is_hebrew else 'Table view of policies by Mislaka affiliation mappings.',
                data_table={
                    'columns': list(status_rows[0].keys()) if status_rows else [],
                    'rows': status_rows
                },
                order=3
            ))

            plan_rows = []
            for acct in accounts[:80]:
                plan_rows.append({
                    'מספר פוליסה' if is_hebrew else 'Policy Number': acct.get('policy_number', ''),
                    'תאריך תחילה' if is_hebrew else 'Start Date': acct.get('start_date', ''),
                    'דמי ניהול מצבירה %' if is_hebrew else 'Mgmt Fee Savings %': acct.get('management_fee_savings', 0),
                    'דמי ניהול מהפקדה %' if is_hebrew else 'Mgmt Fee Deposits %': acct.get('management_fee_deposits', 0),
                    'כיסוי חיים' if is_hebrew else 'Life Coverage': acct.get('death_coverage', 0),
                    'כיסוי אכ"ע' if is_hebrew else 'Disability Coverage': acct.get('disability_coverage', 0),
                    'תגמולים' if is_hebrew else 'Savings': acct.get('savings_balance', 0),
                    'פיצויים' if is_hebrew else 'Severance': acct.get('severance_balance', 0),
                })

            sections.append(ReportSection(
                title='רשימת תוכניות (פירוט טבלאי)' if is_hebrew else 'Plan Details (Tabular)',
                content='פירוט תוכניות לפי מודל הדוח המסונף.' if is_hebrew else 'Detailed plan view aligned with the affiliated report model.',
                data_table={
                    'columns': list(plan_rows[0].keys()) if plan_rows else [],
                    'rows': plan_rows
                },
                order=4
            ))

        if contributions:
            contribution_rows = []
            for contrib in contributions[:120]:
                contribution_rows.append({
                    'תקופה' if is_hebrew else 'Period': contrib.get('period', ''),
                    'מעסיק' if is_hebrew else 'Employer': contrib.get('employer_name', ''),
                    'הפקדת עובד' if is_hebrew else 'Employee Amount': contrib.get('employee_amount', 0),
                    'הפקדת מעסיק' if is_hebrew else 'Employer Amount': contrib.get('employer_amount', 0),
                    'פיצויים' if is_hebrew else 'Severance': contrib.get('severance_amount', 0),
                    'סה״כ' if is_hebrew else 'Total': contrib.get('total_amount', 0),
                })

            sections.append(ReportSection(
                title='פירוט הפקדות וחובות' if is_hebrew else 'Contribution Details',
                content='רצף הפקדות לפי תקופה לצורכי בקרה ותאימות.' if is_hebrew else 'Period-level contribution trail for control and reconciliation.',
                data_table={
                    'columns': list(contribution_rows[0].keys()) if contribution_rows else [],
                    'rows': contribution_rows
                },
                order=5
            ))

        if employers or accounts:
            employer_values = []
            seen = set()
            for acct in accounts:
                emp_name = (acct.get('employer_name') or '').strip()
                if emp_name and emp_name not in seen:
                    seen.add(emp_name)
                    employer_values.append({
                        'שם מעסיק' if is_hebrew else 'Employer Name': emp_name,
                        'מספר מעסיק' if is_hebrew else 'Employer ID': acct.get('employer_id', ''),
                        'סעיף 14' if is_hebrew else 'Section 14': ('כן' if acct.get('section14') else 'לא') if is_hebrew else ('Yes' if acct.get('section14') else 'No'),
                    })
            for emp in employers:
                emp_name = (emp.get('name') or '').strip() if isinstance(emp, dict) else str(emp).strip()
                if emp_name and emp_name not in seen:
                    seen.add(emp_name)
                    employer_values.append({
                        'שם מעסיק' if is_hebrew else 'Employer Name': emp_name,
                        'מספר מעסיק' if is_hebrew else 'Employer ID': emp.get('id', '') if isinstance(emp, dict) else '',
                        'סעיף 14' if is_hebrew else 'Section 14': '',
                    })

            if employer_values:
                sections.append(ReportSection(
                    title='פרטי מעסיקים' if is_hebrew else 'Employer Information',
                    content='שיוך מעסיקים לחשבונות ולזכויות.' if is_hebrew else 'Employer affiliation to accounts and severance rights.',
                    data_table={
                        'columns': list(employer_values[0].keys()),
                        'rows': employer_values[:80]
                    },
                    order=6
                ))

        if totals:
            totals_rows = [{
                'שדה' if is_hebrew else 'Metric': 'סה״כ צבירה' if is_hebrew else 'Total Balance',
                'ערך' if is_hebrew else 'Value': totals.get('total_balance', 0)
            }, {
                'שדה' if is_hebrew else 'Metric': 'סה״כ חסכונות' if is_hebrew else 'Total Savings',
                'ערך' if is_hebrew else 'Value': totals.get('total_savings', totals.get('total_savings_balance', 0))
            }, {
                'שדה' if is_hebrew else 'Metric': 'סה״כ פיצויים' if is_hebrew else 'Total Severance',
                'ערך' if is_hebrew else 'Value': totals.get('total_severance', totals.get('total_severance_balance', 0))
            }, {
                'שדה' if is_hebrew else 'Metric': 'מספר פוליסות' if is_hebrew else 'Policy Count',
                'ערך' if is_hebrew else 'Value': totals.get('account_count', len(accounts))
            }]

            sections.append(ReportSection(
                title='סיכום כספי (מודל דוח)' if is_hebrew else 'Financial Summary (Model-Aligned)',
                content='תקציר כספי לצורך השוואה מול מודל הדוח המסונף.' if is_hebrew else 'Financial summary aligned with the affiliated report model.',
                data_table={
                    'columns': list(totals_rows[0].keys()),
                    'rows': totals_rows
                },
                order=7
            ))

        return sections

    def _build_affiliation_mapping_section(self, is_hebrew: bool) -> Optional[ReportSection]:
        """Build a compact affiliations map section from authoritative schema constants."""
        try:
            from services.pension_data_agent import MislakaSchemaMapping
        except Exception:
            return None

        rows: List[Dict[str, Any]] = []

        for code, info in sorted(MislakaSchemaMapping.INTERFACE_CODES.items(), key=lambda item: int(item[0]))[:20]:
            rows.append({
                'קבוצה' if is_hebrew else 'Group': 'ממשק' if is_hebrew else 'Interface',
                'קוד' if is_hebrew else 'Code': code,
                'שם' if is_hebrew else 'Name': info.get('he', info.get('name', '')),
                'שיוך' if is_hebrew else 'Affiliation': info.get('schema', info.get('name', ''))
            })

        for code, info in sorted(MislakaSchemaMapping.PRODUCT_TYPE_CODES.items(), key=lambda item: str(item[0]))[:20]:
            rows.append({
                'קבוצה' if is_hebrew else 'Group': 'מוצר' if is_hebrew else 'Product',
                'קוד' if is_hebrew else 'Code': code,
                'שם' if is_hebrew else 'Name': info.get('he', info.get('name', '')),
                'שיוך' if is_hebrew else 'Affiliation': info.get('en', '')
            })

        for code, info in sorted(MislakaSchemaMapping.STATUS_CODES.items(), key=lambda item: str(item[0]))[:10]:
            rows.append({
                'קבוצה' if is_hebrew else 'Group': 'סטטוס' if is_hebrew else 'Status',
                'קוד' if is_hebrew else 'Code': code,
                'שם' if is_hebrew else 'Name': info.get('he', info.get('name', '')),
                'שיוך' if is_hebrew else 'Affiliation': info.get('en', '')
            })

        for code, info in sorted(MislakaSchemaMapping.ID_TYPE_CODES.items(), key=lambda item: str(item[0]))[:10]:
            rows.append({
                'קבוצה' if is_hebrew else 'Group': 'זיהוי' if is_hebrew else 'ID Type',
                'קוד' if is_hebrew else 'Code': code,
                'שם' if is_hebrew else 'Name': info.get('he', ''),
                'שיוך' if is_hebrew else 'Affiliation': info.get('en', '')
            })

        return ReportSection(
            title='מפת שיוכים (Affiliations)' if is_hebrew else 'Affiliation Mapping Snapshot',
            content='טבלת שיוכים לפי מסלקה: ממשקים, מוצרים, סטטוסים וסוגי זיהוי.' if is_hebrew
            else 'Affiliation map by Mislaka schema: interfaces, products, statuses, and ID types.',
            data_table={
                'columns': list(rows[0].keys()) if rows else [],
                'rows': rows
            },
            order=9
        )

    @staticmethod
    def _to_float_amount(value: Any) -> float:
        """Best-effort numeric conversion for monetary/coverage fields."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)

        text_value = str(value).strip()
        if not text_value:
            return 0.0

        cleaned = (
            text_value
            .replace(',', '')
            .replace('₪', '')
            .replace('$', '')
            .replace('€', '')
            .replace('%', '')
        )
        cleaned = re.sub(r'[^0-9\.\-]', '', cleaned)
        if cleaned in ['', '-', '.', '-.']:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    @staticmethod
    def _mask_identifier(identifier: Any) -> str:
        """Mask identifiers for privacy in report tables."""
        identifier_str = str(identifier or '').strip()
        if len(identifier_str) <= 4:
            return identifier_str
        return f"{identifier_str[:2]}****{identifier_str[-2:]}"

    @staticmethod
    def _column_matches(column_name: str, tokens: List[str]) -> bool:
        column_lower = (column_name or '').lower()
        return any(token in column_lower for token in tokens)

    def _extract_savings_cover_id_summary(
        self,
        doc_data: Optional[Dict[str, Any]],
        pension_data: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Extract affiliated summary metrics focused on savings, cover and ID dimensions.
        The output is intentionally source-agnostic (no credentials or external links).
        """
        rows = []
        columns = []
        if isinstance(doc_data, dict):
            rows = doc_data.get('rows', []) or []
            columns = doc_data.get('columns', []) or []

        id_tokens = [
            'id', 'identity', 'id_number', 'customer_id', 'policyholder_id',
            'ת.ז', 'ת"ז', 'תז', 'תעודת זהות', 'מספר זהות'
        ]
        savings_tokens = [
            'saving', 'savings', 'balance', 'accumulated', 'total_balance',
            'יתרה', 'צבירה', 'חיסכון', 'תגמולים'
        ]
        cover_tokens = [
            'cover', 'coverage', 'insured_amount', 'sum_insured',
            'death_coverage', 'disability_coverage', 'כיסוי', 'סכום ביטוח'
        ]

        id_columns = [c for c in columns if self._column_matches(str(c), id_tokens)]
        savings_columns = [c for c in columns if self._column_matches(str(c), savings_tokens)]
        cover_columns = [c for c in columns if self._column_matches(str(c), cover_tokens)]

        total_savings = 0.0
        total_cover = 0.0
        records_with_savings = 0
        records_with_cover = 0
        id_values: List[str] = []
        sample_rows: List[Dict[str, Any]] = []

        # Prefer pension structured data when available (Mislaka-aligned source).
        if isinstance(pension_data, dict) and pension_data.get('accounts'):
            accounts = pension_data.get('accounts', []) or []
            client_data = pension_data.get('client', {})
            if isinstance(client_data, list):
                client_data = client_data[0] if client_data else {}
            shared_client_id = ''
            if isinstance(client_data, dict):
                shared_client_id = str(client_data.get('id_number', '') or '').strip()

            for account in accounts[:500]:
                account_id = str(account.get('id_number') or shared_client_id or account.get('policy_number') or '').strip()
                savings_value = (
                    self._to_float_amount(account.get('savings_balance'))
                    or self._to_float_amount(account.get('total_balance'))
                )
                cover_value = (
                    self._to_float_amount(account.get('death_coverage'))
                    + self._to_float_amount(account.get('disability_coverage'))
                )

                if account_id:
                    id_values.append(account_id)
                if savings_value > 0:
                    total_savings += savings_value
                    records_with_savings += 1
                if cover_value > 0:
                    total_cover += cover_value
                    records_with_cover += 1

                if account_id or savings_value > 0 or cover_value > 0:
                    sample_rows.append({
                        'id': self._mask_identifier(account_id) if account_id else '',
                        'savings': round(savings_value, 2),
                        'cover': round(cover_value, 2),
                        'reference': str(account.get('policy_number', '') or '')
                    })
        else:
            for row in rows[:500]:
                if not isinstance(row, dict):
                    continue

                id_value = ''
                for col in id_columns:
                    candidate = str(row.get(col, '') or '').strip()
                    if candidate:
                        id_value = candidate
                        break

                savings_value = sum(self._to_float_amount(row.get(col)) for col in savings_columns)
                cover_value = sum(self._to_float_amount(row.get(col)) for col in cover_columns)

                if id_value:
                    id_values.append(id_value)
                if savings_value > 0:
                    total_savings += savings_value
                    records_with_savings += 1
                if cover_value > 0:
                    total_cover += cover_value
                    records_with_cover += 1

                if id_value or savings_value > 0 or cover_value > 0:
                    sample_rows.append({
                        'id': self._mask_identifier(id_value) if id_value else '',
                        'savings': round(savings_value, 2),
                        'cover': round(cover_value, 2),
                        'reference': ''
                    })

        unique_ids = sorted({v for v in id_values if v})
        customer_id = ''
        customer_id_valid = None
        customer_birth_date = ''
        customer_birth_date_raw = ''
        integrity_issues: List[str] = []

        client_data = {}
        if isinstance(pension_data, dict):
            client_data = pension_data.get('client', {}) or {}
            if isinstance(client_data, list):
                client_data = client_data[0] if client_data else {}
            if not isinstance(client_data, dict):
                client_data = {}
        client_data = self._normalize_client_profile_fields(client_data)
        if client_data.get('id_number'):
            customer_id = str(client_data.get('id_number') or '').strip()
        elif unique_ids:
            customer_id = self._normalize_customer_identifier(unique_ids[0])

        if customer_id and customer_id.isdigit() and len(customer_id) == 9:
            customer_id_valid = self._is_valid_israeli_id(customer_id)
            if not customer_id_valid:
                integrity_issues.append('Customer ID failed Israeli checksum validation')
        elif customer_id:
            customer_id_valid = False
            integrity_issues.append('Customer ID is not a 9-digit value')

        customer_birth_date = str(client_data.get('birth_date') or '').strip()
        customer_birth_date_raw = str(client_data.get('birth_date_raw') or '').strip()

        if not customer_birth_date:
            birth_tokens = ['birth', 'birth_date', 'date_of_birth', 'dob', 'תאריך לידה', 'לידה']
            birth_columns = [c for c in columns if self._column_matches(str(c), birth_tokens)]
            for row in rows[:500]:
                if not isinstance(row, dict):
                    continue
                found_raw = ''
                for birth_col in birth_columns:
                    candidate = str(row.get(birth_col, '') or '').strip()
                    if candidate:
                        found_raw = candidate
                        break
                if found_raw:
                    birth_raw, birth_display = self._normalize_birth_date(found_raw)
                    customer_birth_date_raw = birth_raw or customer_birth_date_raw
                    customer_birth_date = birth_display or customer_birth_date
                    if customer_birth_date:
                        break

        if customer_birth_date_raw and not customer_birth_date:
            integrity_issues.append('Birth date could not be normalized to DD/MM/YYYY')

        records_analyzed = max(len(rows), len(sample_rows), 0)
        id_rows_count = len([entry for entry in sample_rows if entry.get('id')])

        return {
            'records_analyzed': records_analyzed,
            'id_columns': id_columns,
            'savings_columns': savings_columns,
            'cover_columns': cover_columns,
            'ids_with_values': len(id_values),
            'unique_id_count': len(unique_ids),
            'id_row_coverage': id_rows_count,
            'total_savings': round(total_savings, 2),
            'average_savings': round(total_savings / records_with_savings, 2) if records_with_savings else 0.0,
            'total_cover': round(total_cover, 2),
            'average_cover': round(total_cover / records_with_cover, 2) if records_with_cover else 0.0,
            'coverage_to_savings_ratio': round(total_cover / total_savings, 2) if total_savings > 0 else None,
            'sample_rows': sample_rows[:120],
            'customer_id': customer_id,
            'customer_id_masked': self._mask_identifier(customer_id) if customer_id else '',
            'customer_id_valid': customer_id_valid,
            'birth_date': customer_birth_date,
            'birth_date_raw': customer_birth_date_raw,
            'integrity_issues': integrity_issues,
        }

    def _build_savings_cover_id_section(
        self,
        summary: Optional[Dict[str, Any]],
        is_hebrew: bool
    ) -> Optional[ReportSection]:
        """Build a compact affiliated section for savings/cover/ID analysis."""
        if not summary:
            return None

        records_analyzed = int(summary.get('records_analyzed', 0) or 0)
        total_savings = float(summary.get('total_savings', 0) or 0)
        total_cover = float(summary.get('total_cover', 0) or 0)
        unique_id_count = int(summary.get('unique_id_count', 0) or 0)
        if records_analyzed <= 0 and total_savings <= 0 and total_cover <= 0 and unique_id_count <= 0:
            return None

        if is_hebrew:
            content = (
                "סיכום מסונף לחיסכון וביטוח (על בסיס שיוכי מסלקה):\n\n"
                f"• מזהה לקוח: {summary.get('customer_id', 'לא זמין')}\n"
                f"• תאריך לידה: {summary.get('birth_date', 'לא זמין')}"
                + (f" (מקור: {summary.get('birth_date_raw')})" if summary.get('birth_date_raw') else "")
                + "\n"
                f"• רשומות שנותחו: {records_analyzed}\n"
                f"• סך חיסכון: ₪{total_savings:,.2f}\n"
                f"• סך כיסוי: ₪{total_cover:,.2f}\n"
                f"• מזהים ייחודיים: {unique_id_count}\n"
                f"• יחס כיסוי/חיסכון: {summary.get('coverage_to_savings_ratio', 'N/A')}\n"
                f"• תקינות מזהה: {'✓ תקין' if summary.get('customer_id_valid') else '⚠ דורש בדיקה'}"
            )
            title = 'סיכום מסונף - חיסכון, כיסוי וזיהוי'
            columns = ['מזהה לקוח', 'תאריך לידה', 'מזהה (מוסתר)', 'חיסכון', 'כיסוי', 'אסמכתא']
            data_rows = [{
                'מזהה לקוח': summary.get('customer_id', ''),
                'תאריך לידה': summary.get('birth_date', ''),
                'מזהה (מוסתר)': row.get('id', ''),
                'חיסכון': row.get('savings', 0),
                'כיסוי': row.get('cover', 0),
                'אסמכתא': row.get('reference', ''),
            } for row in summary.get('sample_rows', [])[:60]]
        else:
            content = (
                "Affiliated savings and insurance snapshot (Mislaka-aligned):\n\n"
                f"• Customer ID: {summary.get('customer_id', 'N/A')}\n"
                f"• Birth Date: {summary.get('birth_date', 'N/A')}"
                + (f" (source: {summary.get('birth_date_raw')})" if summary.get('birth_date_raw') else "")
                + "\n"
                f"• Records analyzed: {records_analyzed}\n"
                f"• Total savings: ₪{total_savings:,.2f}\n"
                f"• Total cover: ₪{total_cover:,.2f}\n"
                f"• Unique IDs: {unique_id_count}\n"
                f"• Cover/Savings ratio: {summary.get('coverage_to_savings_ratio', 'N/A')}\n"
                f"• ID validation: {'Valid' if summary.get('customer_id_valid') else 'Needs review'}"
            )
            title = 'Affiliated Summary - Savings, Cover & ID'
            columns = ['Customer ID', 'Birth Date', 'Masked ID', 'Savings', 'Cover', 'Reference']
            data_rows = [{
                'Customer ID': summary.get('customer_id', ''),
                'Birth Date': summary.get('birth_date', ''),
                'Masked ID': row.get('id', ''),
                'Savings': row.get('savings', 0),
                'Cover': row.get('cover', 0),
                'Reference': row.get('reference', ''),
            } for row in summary.get('sample_rows', [])[:60]]

        # Fall back to metric table when we don't have row-level samples.
        if not data_rows:
            metric_key = 'מדד' if is_hebrew else 'Metric'
            value_key = 'ערך' if is_hebrew else 'Value'
            metrics_rows = [
                {metric_key: 'Records', value_key: records_analyzed},
                {metric_key: 'Total Savings', value_key: total_savings},
                {metric_key: 'Total Cover', value_key: total_cover},
                {metric_key: 'Unique IDs', value_key: unique_id_count},
                {metric_key: 'Cover/Savings Ratio', value_key: summary.get('coverage_to_savings_ratio', 'N/A')},
            ]
            return ReportSection(
                title=title,
                content=content,
                data_table={'columns': [metric_key, value_key], 'rows': metrics_rows},
                order=4
            )

        return ReportSection(
            title=title,
            content=content,
            data_table={'columns': columns, 'rows': data_rows},
            order=4
        )

    def _build_savings_cover_id_charts(
        self,
        summary: Optional[Dict[str, Any]],
        lang_code: str
    ) -> List[ChartConfig]:
        """Generate supplementary charts focused on savings, cover and ID availability."""
        if not summary:
            return []

        charts: List[ChartConfig] = []
        is_hebrew = lang_code == 'hebrew'
        total_savings = float(summary.get('total_savings', 0) or 0)
        total_cover = float(summary.get('total_cover', 0) or 0)
        records_analyzed = int(summary.get('records_analyzed', 0) or 0)
        id_rows = int(summary.get('id_row_coverage', 0) or 0)

        if total_savings > 0 or total_cover > 0:
            charts.append(ChartConfig(
                type=ChartType.BAR,
                title='חיסכון מול כיסוי' if is_hebrew else 'Savings vs Cover',
                data={
                    'labels': ['חיסכון' if is_hebrew else 'Savings', 'כיסוי' if is_hebrew else 'Cover'],
                    'values': [total_savings, total_cover]
                },
                options={
                    'colors': ['#10b981', '#1a237e'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))

        if records_analyzed > 0:
            missing_ids = max(records_analyzed - id_rows, 0)
            charts.append(ChartConfig(
                type=ChartType.DOUGHNUT,
                title='כיסוי שדות זיהוי' if is_hebrew else 'ID Field Coverage',
                data={
                    'labels': ['כולל מזהה' if is_hebrew else 'With ID', 'ללא מזהה' if is_hebrew else 'Without ID'],
                    'values': [id_rows, missing_ids]
                },
                options={'colors': ['#3b82f6', '#cbd5e1']}
            ))

        return charts
    
    def _generate_swiftness_resources_section(self, is_hebrew: bool) -> str:
        """Generate a report section with Swiftness affiliated links and resources."""
        try:
            from services.swiftness_data_service import get_swiftness_data_service
            svc = get_swiftness_data_service()
            catalog = svc.get_resource_catalog()
            model = svc.get_report_model()
        except Exception:
            return ''
        
        lines = []
        meta = catalog.get('metadata', {})
        
        if is_hebrew:
            lines.append('📥 משאבי נתונים מ-Swiftness לעבודה מול המסלקה:\n')
            lines.append(f'סה"כ משאבים: {meta.get("total_resources", 0)}')
            lines.append(f'ממשקים: {", ".join(meta.get("interfaces", []))}')
            lines.append(f'סוגי קבצים: {", ".join(t.upper() for t in meta.get("file_types", []))}')
            lines.append('')
            lines.append('🔗 קישורים ישירים:')
            for link in catalog.get('quick_links', []):
                lines.append(f'  • {link.get("label_he", link["label"])}: {link["url"]}')
            lines.append('')
            lines.append('📋 כללי מערכת (סכימות וטבלאות קודים):')
            for res in catalog.get('system_general', [])[:6]:
                lines.append(f'  • {res["name"]} ({res.get("file_type", "").upper()}'
                             f'{" v" + res["version"] if res.get("version") else ""})')
            lines.append('')
            lines.append('📂 קבצים עדכניים לעבודה מול המסלקה:')
            for res in catalog.get('mislaka_work_files', [])[:6]:
                lines.append(f'  • {res["name"]} ({res.get("file_type", "").upper()})')
            
            # Report model summary
            model_meta = model.get('metadata', {})
            sections_count = len(model.get('sections', []))
            lines.append('')
            lines.append(f'📊 מודל דוח מקיף: {sections_count} חלקים')
            for sec in model.get('sections', []):
                fields_count = len(sec.get('data_fields', []))
                lines.append(f'  {sec["order"]}. {sec["title_he"]} ({fields_count} שדות)')
            
            # Data integrity
            rules = model_meta.get('data_integrity_rules', [])
            if rules:
                lines.append('')
                lines.append('🛡️ כללי שלמות נתונים:')
                for rule in rules:
                    lines.append(f'  ✓ {rule}')
        else:
            lines.append('📥 Swiftness Data Resources for Mislaka Integration:\n')
            lines.append(f'Total Resources: {meta.get("total_resources", 0)}')
            lines.append(f'Interfaces: {", ".join(meta.get("interfaces", []))}')
            lines.append(f'File Types: {", ".join(t.upper() for t in meta.get("file_types", []))}')
            lines.append('')
            lines.append('🔗 Direct Links:')
            for link in catalog.get('quick_links', []):
                lines.append(f'  • {link["label"]}: {link["url"]}')
            lines.append('')
            lines.append('📋 System General (Schemas & Code Tables):')
            for res in catalog.get('system_general', [])[:6]:
                lines.append(f'  • {res.get("name_en", res["name"])} ({res.get("file_type", "").upper()}'
                             f'{" v" + res["version"] if res.get("version") else ""})')
            lines.append('')
            lines.append('📂 Latest Mislaka Work Files:')
            for res in catalog.get('mislaka_work_files', [])[:6]:
                lines.append(f'  • {res.get("name_en", res["name"])} ({res.get("file_type", "").upper()})')
            
            # Report model summary
            model_meta = model.get('metadata', {})
            sections_count = len(model.get('sections', []))
            lines.append('')
            lines.append(f'📊 Comprehensive Report Model: {sections_count} sections')
            for sec in model.get('sections', []):
                fields_count = len(sec.get('data_fields', []))
                lines.append(f'  {sec["order"]}. {sec["title_en"]} ({fields_count} fields)')
            
            # Data integrity
            rules = model_meta.get('data_integrity_rules', [])
            if rules:
                lines.append('')
                lines.append('🛡️ Data Integrity Rules:')
                for rule in rules:
                    lines.append(f'  ✓ {rule}')
        
        return '\n'.join(lines)
    
    def _generate_data_content_section(self, doc_data: Dict[str, Any], 
                                        analysis: AnalysisResult, is_hebrew: bool) -> str:
        """
        Generate a section showing actual data content from uploaded files.
        This displays the real values from CSV/ZIP files, not just statistics.
        """
        content_lines = []
        
        columns = doc_data.get('columns', [])
        rows = doc_data.get('rows', [])
        files = doc_data.get('files', [])
        
        # If from ZIP, show file list
        if files:
            if is_hebrew:
                content_lines.append("📁 קבצים שנותחו מתוך ה-ZIP:")
            else:
                content_lines.append("📁 Files analyzed from ZIP:")
            
            for f in files:
                content_lines.append(f"  • {f.get('name', 'Unknown')} ({f.get('row_count', 0)} שורות)" if is_hebrew else f"  • {f.get('name', 'Unknown')} ({f.get('row_count', 0)} rows)")
            content_lines.append("")
        
        # Show column headers
        if columns:
            if is_hebrew:
                content_lines.append(f"📋 עמודות הנתונים ({len(columns)}):")
            else:
                content_lines.append(f"📋 Data Columns ({len(columns)}):")
            
            # Display columns in a formatted way
            col_display = []
            for col in columns[:20]:  # Limit to 20 columns
                col_display.append(f"  • {col}")
            content_lines.extend(col_display)
            if len(columns) > 20:
                content_lines.append(f"  ... ועוד {len(columns) - 20} עמודות" if is_hebrew else f"  ... and {len(columns) - 20} more columns")
            content_lines.append("")
        
        # Show sample data rows as table
        if rows:
            if is_hebrew:
                content_lines.append(f"📊 נתונים שחולצו ({len(rows)} רשומות):")
                content_lines.append("=" * 50)
            else:
                content_lines.append(f"📊 Extracted Data ({len(rows)} records):")
                content_lines.append("=" * 50)
            
            # Display first 10 rows with all their values
            for i, row in enumerate(rows[:15], 1):
                if is_hebrew:
                    content_lines.append(f"\n🔹 רשומה {i}:")
                else:
                    content_lines.append(f"\n🔹 Record {i}:")
                
                for key, value in row.items():
                    if value and str(value).strip():
                        # Clean and format the value
                        val_str = str(value).strip()
                        # Detect if it's a numeric value
                        try:
                            num_val = float(val_str.replace(',', '').replace('₪', '').replace('$', ''))
                            if num_val > 1000:
                                val_str = f"₪{num_val:,.0f}" if any(x in key.lower() for x in ['premium', 'cover', 'amount', 'סכום', 'פרמיה', 'כיסוי']) else f"{num_val:,.0f}"
                        except ValueError:
                            pass
                        content_lines.append(f"    {key}: {val_str}")
            
            if len(rows) > 15:
                content_lines.append(f"\n... ועוד {len(rows) - 15} רשומות" if is_hebrew else f"\n... and {len(rows) - 15} more records")
        
        # Extract and highlight key insurance/financial fields
        key_fields = self._extract_key_fields_from_data(rows, is_hebrew)
        if key_fields:
            content_lines.append("")
            if is_hebrew:
                content_lines.append("🎯 שדות מרכזיים שזוהו:")
            else:
                content_lines.append("🎯 Key Fields Identified:")
            
            for field_name, field_value in key_fields.items():
                content_lines.append(f"  • {field_name}: {field_value}")
        
        return '\n'.join(content_lines) if content_lines else ""
    
    def _generate_pension_section(self, pension_data: Dict[str, Any], 
                                   pension_report: str, is_hebrew: bool) -> str:
        """
        Generate a comprehensive pension and insurance report section.
        Uses data from the enhanced PensionDataAgent for Mislaka XML files.
        
        Supports the full Mislaka interface standards:
        - Holdings Interface (v9.7.7)
        - Severance Interface (v5.9.38)
        - Event Interface (v7.6.30)
        - Transference Interface (v3.7.2)
        
        This displays:
        - Professional Mislaka report generated by PensionDataAgent
        - Account summaries with balances by provider/product type
        - Section 14 status and severance details
        - Health score and AI recommendations
        - Contribution analysis and trends
        """
        content_lines = []
        
        # If we have the pre-generated Mislaka pension report, include it
        if pension_report:
            content_lines.append(pension_report)
            content_lines.append("")
            content_lines.append("─" * 50)
            content_lines.append("")
        
        # If we have structured pension data, add detailed breakdown
        if pension_data:
            # Support both 'totals' (new) and 'summary' (legacy) keys
            totals = pension_data.get('totals', pension_data.get('summary', {}))
            accounts = pension_data.get('accounts', [])
            clients = pension_data.get('client', {})
            if isinstance(clients, list) and clients:
                clients = clients[0]
            header = pension_data.get('header', {})
            contributions = pension_data.get('contributions', [])
            severance = pension_data.get('severance', [])
            
            if is_hebrew:
                # Financial summary section with health score
                health_score = totals.get('health_score', {})
                if health_score:
                    content_lines.append("🎯 ציון בריאות פיננסית:")
                    content_lines.append("=" * 40)
                    content_lines.append(f"• ציון כולל: {health_score.get('overall', 0)}/100 ({health_score.get('rating_he', 'לא ידוע')})")
                    content_lines.append(f"• ציון חסכונות: {health_score.get('savings', 0)}/100")
                    content_lines.append(f"• ציון פיזור: {health_score.get('diversification', 0)}/100")
                    content_lines.append(f"• ציון סעיף 14: {health_score.get('section14', 0)}/100")
                    content_lines.append("")
                
                # Financial summary section
                content_lines.append("💰 סיכום כספי מפורט:")
                content_lines.append("=" * 40)
                content_lines.append(f"• סה״כ יתרה בחשבונות: {totals.get('total_balance_formatted', '₪0')}")
                content_lines.append(f"• סה״כ חסכונות: {totals.get('total_savings_formatted', '₪0')}")
                content_lines.append(f"• סה״כ פיצויים צבורים: {totals.get('total_severance_formatted', '₪0')}")
                content_lines.append(f"• מספר חשבונות/פוליסות: {totals.get('account_count', 0)}")
                content_lines.append(f"• מספר יצרנים/חברות: {totals.get('provider_count', 0)}")
                
                # Providers
                providers = totals.get('providers', [])
                if providers:
                    content_lines.append(f"• יצרנים: {', '.join(providers)}")
                content_lines.append("")
                
                # Section 14 status
                content_lines.append("📌 סעיף 14 (פיצויים):")
                if totals.get('section14_coverage'):
                    content_lines.append("• סטטוס: ✅ מכוסה")
                    content_lines.append("• ✅ הלקוח מכוסה תחת סעיף 14 - פיצויים מובטחים")
                    content_lines.append(f"• מספר חשבונות עם סעיף 14: {totals.get('section14_accounts', 0)}")
                else:
                    content_lines.append("• סטטוס: ⚠️ לא מכוסה")
                    content_lines.append("• ⚠️ אין כיסוי סעיף 14 - יש לבדוק עם המעסיק")
                content_lines.append("")
                
                # Contribution summary
                contrib_totals = totals.get('contributions', {})
                if contrib_totals:
                    content_lines.append("📈 סיכום הפקדות:")
                    content_lines.append("=" * 40)
                    content_lines.append(f"• הפקדות עובד: ₪{contrib_totals.get('employee_total', 0):,.2f}")
                    content_lines.append(f"• הפקדות מעסיק: ₪{contrib_totals.get('employer_total', 0):,.2f}")
                    content_lines.append(f"• הפקדות פיצויים: ₪{contrib_totals.get('severance_total', 0):,.2f}")
                    content_lines.append(f"• סה״כ הפקדות: ₪{contrib_totals.get('grand_total', 0):,.2f}")
                    content_lines.append(f"• תקופות: {contrib_totals.get('periods_count', 0)}")
                    content_lines.append("")
                
                # Contribution trend
                trend = totals.get('contribution_trend')
                if trend:
                    trend_he = totals.get('contribution_trend_he', trend)
                    content_lines.append("📈 מגמת הפקדות:")
                    content_lines.append(f"• מגמה: {trend_he}")
                    content_lines.append("")
                
                # Missing months warning
                missing = totals.get('missing_contribution_months', [])
                if missing:
                    content_lines.append("⚠️ אזהרה - חודשים חסרים:")
                    content_lines.append(f"• נמצאו {len(missing)} חודשים ללא הפקדות")
                    content_lines.append(f"• חודשים: {', '.join(missing[:6])}{'...' if len(missing) > 6 else ''}")
                    content_lines.append("")
                
                # Account details
                if accounts:
                    content_lines.append("📁 פירוט חשבונות:")
                    content_lines.append("-" * 40)
                    total_balance = totals.get('total_balance', 1)
                    for i, acct in enumerate(accounts[:10], 1):
                        balance = acct.get('total_balance', acct.get('balance', 0))
                        pct = (balance / total_balance * 100) if total_balance > 0 else 0
                        content_lines.append(f"\n🔹 חשבון {i}:")
                        content_lines.append(f"   • מספר פוליסה: {acct.get('policy_number', 'לא ידוע')}")
                        if acct.get('provider'):
                            content_lines.append(f"   • יצרן: {acct.get('provider')}")
                        if acct.get('product_type_name') or acct.get('product_name') or acct.get('product_type'):
                            content_lines.append(f"   • סוג מוצר: {acct.get('product_type_name', acct.get('product_name', acct.get('product_type', 'לא ידוע')))}")
                        if acct.get('status'):
                            content_lines.append(f"   • סטטוס: {acct.get('status')}")
                        content_lines.append(f"   • יתרה: ₪{balance:,.2f} ({pct:.1f}% מהכולל)")
                        if acct.get('savings_balance', 0) > 0:
                            content_lines.append(f"   • חיסכון: ₪{acct.get('savings_balance', 0):,.2f}")
                        if acct.get('severance_balance', 0) > 0:
                            content_lines.append(f"   • פיצויים: ₪{acct.get('severance_balance', 0):,.2f}")
                        if acct.get('section14'):
                            content_lines.append(f"   • סעיף 14: ✅ מכוסה")
                        if acct.get('management_fee_savings', 0) > 0:
                            content_lines.append(f"   • דמי ניהול: {acct.get('management_fee_savings', 0):.2f}%")
                        if acct.get('employer_name'):
                            content_lines.append(f"   • מעסיק: {acct.get('employer_name')}")
                    
                    if len(accounts) > 10:
                        content_lines.append(f"\n   ... ועוד {len(accounts) - 10} חשבונות")
                    content_lines.append("")
            
            else:
                # English version
                health_score = totals.get('health_score', {})
                if health_score:
                    content_lines.append("🎯 Financial Health Score:")
                    content_lines.append("=" * 40)
                    content_lines.append(f"• Overall Score: {health_score.get('overall', 0)}/100 ({health_score.get('rating', 'unknown')})")
                    content_lines.append(f"• Savings Score: {health_score.get('savings', 0)}/100")
                    content_lines.append(f"• Diversification Score: {health_score.get('diversification', 0)}/100")
                    content_lines.append(f"• Section 14 Score: {health_score.get('section14', 0)}/100")
                    content_lines.append("")
                
                content_lines.append("💰 Detailed Financial Summary:")
                content_lines.append("=" * 40)
                content_lines.append(f"• Total Account Balance: {totals.get('total_balance_formatted', '₪0')}")
                content_lines.append(f"• Total Savings: {totals.get('total_savings_formatted', '₪0')}")
                content_lines.append(f"• Total Severance Accrued: {totals.get('total_severance_formatted', '₪0')}")
                content_lines.append(f"• Number of Accounts/Policies: {totals.get('account_count', 0)}")
                content_lines.append(f"• Number of Providers: {totals.get('provider_count', 0)}")
                
                providers = totals.get('providers', [])
                if providers:
                    content_lines.append(f"• Providers: {', '.join(providers)}")
                content_lines.append("")
                
                # Section 14 status
                content_lines.append("📌 Section 14 (Severance):")
                content_lines.append(f"• Covered: {'Yes' if totals.get('section14_coverage') else 'No'}")
                if totals.get('section14_coverage'):
                    content_lines.append("• ✅ Client is covered under Section 14 - severance is secured")
                    content_lines.append(f"• Accounts with Section 14: {totals.get('section14_accounts', 0)}")
                else:
                    content_lines.append("• ⚠️ No Section 14 coverage - verify with employer")
                content_lines.append("")
                
                # Contribution summary
                contrib_totals = totals.get('contributions', {})
                if contrib_totals:
                    content_lines.append("📈 Contribution Summary:")
                    content_lines.append("=" * 40)
                    content_lines.append(f"• Employee Contributions: ₪{contrib_totals.get('employee_total', 0):,.2f}")
                    content_lines.append(f"• Employer Contributions: ₪{contrib_totals.get('employer_total', 0):,.2f}")
                    content_lines.append(f"• Severance Contributions: ₪{contrib_totals.get('severance_total', 0):,.2f}")
                    content_lines.append(f"• Total Contributions: ₪{contrib_totals.get('grand_total', 0):,.2f}")
                    content_lines.append(f"• Periods: {contrib_totals.get('periods_count', 0)}")
                    content_lines.append("")
                
                # Contribution trend
                trend = totals.get('contribution_trend')
                if trend:
                    content_lines.append("📈 Contribution Trend:")
                    content_lines.append(f"• Trend: {trend.capitalize()}")
                    content_lines.append("")
                
                # Missing months warning
                missing = totals.get('missing_contribution_months', [])
                if missing:
                    content_lines.append("⚠️ Warning - Missing Months:")
                    content_lines.append(f"• Found {len(missing)} months without contributions")
                    content_lines.append(f"• Months: {', '.join(missing[:6])}{'...' if len(missing) > 6 else ''}")
                    content_lines.append("")
                
                # Account details
                if accounts:
                    content_lines.append("📁 Account Details:")
                    content_lines.append("-" * 40)
                    total_balance = totals.get('total_balance', 1)
                    for i, acct in enumerate(accounts[:10], 1):
                        balance = acct.get('total_balance', acct.get('balance', 0))
                        pct = (balance / total_balance * 100) if total_balance > 0 else 0
                        content_lines.append(f"\n🔹 Account {i}:")
                        content_lines.append(f"   • Policy Number: {acct.get('policy_number', 'Unknown')}")
                        content_lines.append(f"   • Balance: ₪{balance:,.2f} ({pct:.1f}% of total)")
                        if acct.get('provider'):
                            content_lines.append(f"   • Provider: {acct.get('provider')}")
                        if acct.get('product_name') or acct.get('product_type'):
                            content_lines.append(f"   • Product: {acct.get('product_name', acct.get('product_type', 'Unknown'))}")
                        if acct.get('status'):
                            content_lines.append(f"   • Status: {acct.get('status')}")
                        content_lines.append(f"   • Balance: ₪{acct.get('balance', 0):,.2f}")
                        if acct.get('severance_balance', 0) > 0:
                            content_lines.append(f"   • Severance: ₪{acct.get('severance_balance', 0):,.2f}")
                        if acct.get('employer'):
                            emp = acct['employer']
                            if isinstance(emp, dict):
                                content_lines.append(f"   • Employer: {emp.get('name', '')}")
                    
                    if len(accounts) > 10:
                        content_lines.append(f"\n   ... and {len(accounts) - 10} more accounts")
                    content_lines.append("")
        
        return '\n'.join(content_lines) if content_lines else ""
    
    def _extract_key_fields_from_data(self, rows: List[Dict], is_hebrew: bool) -> Dict[str, Any]:
        """
        Extract key financial/insurance fields from the actual data rows.
        """
        key_fields = {}
        
        # Keywords to look for
        important_keys = {
            'policy': ['policy', 'פוליסה', 'מספר פוליסה', 'policy_number'],
            'premium': ['premium', 'פרמיה', 'תשלום', 'payment', 'חודשי'],
            'cover': ['cover', 'כיסוי', 'סכום ביטוח', 'סכום', 'coverage', 'amount'],
            'date': ['date', 'תאריך', 'start', 'תחילה', 'end', 'סיום'],
            'id': ['id', 'ת.ז', 'תעודת זהות', 'מספר זהות', 'identity'],
            'name': ['name', 'שם', 'מבוטח', 'insured'],
            'type': ['type', 'סוג', 'תוכנית', 'plan', 'מסלול'],
            'pension': ['pension', 'פנסיה', 'גמל', 'קרן'],
            'beneficiary': ['beneficiary', 'מוטב', 'מוטבים'],
        }
        
        for row in rows[:20]:  # Check first 20 rows
            for col_name, value in row.items():
                if not value or not str(value).strip():
                    continue
                
                col_lower = col_name.lower()
                value_str = str(value).strip()
                
                for field_type, keywords in important_keys.items():
                    if any(kw in col_lower for kw in keywords):
                        label_map = {
                            'policy': 'מספר פוליסה' if is_hebrew else 'Policy Number',
                            'premium': 'פרמיה' if is_hebrew else 'Premium',
                            'cover': 'סכום כיסוי' if is_hebrew else 'Cover Amount',
                            'date': col_name,
                            'id': 'מספר זהות' if is_hebrew else 'ID Number',
                            'name': 'שם' if is_hebrew else 'Name',
                            'type': 'סוג' if is_hebrew else 'Type',
                            'pension': 'פנסיה' if is_hebrew else 'Pension',
                            'beneficiary': 'מוטב' if is_hebrew else 'Beneficiary',
                        }
                        
                        label = label_map.get(field_type, col_name)
                        
                        # Format numeric values
                        if field_type in ['premium', 'cover']:
                            try:
                                num_val = float(value_str.replace(',', '').replace('₪', '').replace('$', ''))
                                value_str = f"₪{num_val:,.0f}"
                            except ValueError:
                                pass
                        
                        # Mask ID numbers for privacy
                        if field_type == 'id' and len(value_str) >= 6:
                            value_str = value_str[:2] + '****' + value_str[-2:]
                        
                        if label not in key_fields:
                            key_fields[label] = value_str
                        break
        
        return key_fields
    
    def _generate_hebrew_insurance_section(self, hebrew_factors: List[Factor], 
                                            is_hebrew: bool) -> str:
        """
        Generate a detailed section showing extracted Hebrew insurance policy details.
        """
        content_lines = []
        
        if is_hebrew:
            content_lines.append("📋 פרטי הפוליסה שחולצו מהמסמכים:\n")
        else:
            content_lines.append("📋 Policy Details Extracted from Documents:\n")
        
        for factor in hebrew_factors:
            # Add factor name as header
            content_lines.append(f"▸ {factor.name}:")
            
            if isinstance(factor.value, dict):
                for key, val in factor.value.items():
                    content_lines.append(f"    • {key}: {val}")
            else:
                content_lines.append(f"    {factor.value}")
            
            content_lines.append("")
        
        # Add importance rating
        if hebrew_factors:
            avg_importance = sum(f.importance for f in hebrew_factors) / len(hebrew_factors)
            if is_hebrew:
                content_lines.append(f"📈 רמת חשיבות ממוצעת: {avg_importance:.0%}")
            else:
                content_lines.append(f"📈 Average Importance: {avg_importance:.0%}")
        
        return '\n'.join(content_lines)
    
    def _generate_charts(
        self,
        analysis: AnalysisResult,
        pension_data: Dict = None,
        doc_data: Dict[str, Any] = None,
        savings_cover_id_summary: Dict[str, Any] = None
    ) -> List[ChartConfig]:
        """
        Generate chart configurations.
        
        For pension data, generates meaningful financial charts:
        - Cumulative savings by provider
        - Savings vs Severance breakdown
        - Insurance coverage breakdown
        """
        charts = []
        
        if savings_cover_id_summary is None:
            savings_cover_id_summary = self._extract_savings_cover_id_summary(doc_data, pension_data)

        # Check if we have pension data for specialized charts
        if pension_data:
            charts.extend(self._generate_pension_charts(pension_data, analysis.language))
            charts.extend(self._build_savings_cover_id_charts(savings_cover_id_summary, analysis.language))
            return charts
        
        # Risk Score Gauge (for non-pension data)
        charts.append(ChartConfig(
            type=ChartType.GAUGE,
            title='Risk Score',
            data={
                'value': analysis.risk_score,
                'min': 0,
                'max': 100,
                'thresholds': [30, 60, 80]
            },
            options={'colors': ['#4caf50', '#ffeb3b', '#ff9800', '#f44336']}
        ))
        
        # Factors Importance Bar Chart - only for meaningful factors
        meaningful_factors = [f for f in analysis.extracted_factors[:8] 
                            if f.category in ['domain', 'currency', 'financial']]
        if meaningful_factors:
            charts.append(ChartConfig(
                type=ChartType.BAR,
                title='Factors Importance',
                data={
                    'labels': [f.name for f in meaningful_factors],
                    'values': [f.importance * 100 for f in meaningful_factors]
                },
                options={'horizontal': True}
            ))
        
        # Data Distribution Pie Chart - only meaningful totals
        if analysis.key_metrics:
            # Filter to only financial totals, not counts
            financial_metrics = {k: v for k, v in analysis.key_metrics.items() 
                               if isinstance(v, (int, float)) and 
                               (k.endswith('_total') or 'balance' in k.lower() or 
                                'savings' in k.lower() or 'coverage' in k.lower()) and
                               v > 0}
            if financial_metrics:
                charts.append(ChartConfig(
                    type=ChartType.PIE,
                    title='Data Distribution',
                    data={
                        'labels': list(financial_metrics.keys())[:6],
                        'values': list(financial_metrics.values())[:6]
                    }
                ))

        charts.extend(self._build_savings_cover_id_charts(savings_cover_id_summary, analysis.language))
        
        return charts
    
    def _generate_pension_charts(self, pension_data: Dict, lang_code: str) -> List[ChartConfig]:
        """
        Generate specialized charts for pension/Mislaka data.
        
        Creates meaningful visualizations:
        1. Cumulative savings by provider (bar chart)
        2. Savings vs Severance breakdown (doughnut)
        3. Insurance coverage breakdown (pie chart)
        """
        charts = []
        is_hebrew = lang_code == 'hebrew'
        
        totals = pension_data.get('totals', {})
        accounts = pension_data.get('accounts', [])
        
        # 1. Cumulative Savings by Provider (Bar Chart)
        provider_totals = {}
        for acct in accounts:
            provider = acct.get('provider', 'לא ידוע' if is_hebrew else 'Unknown')
            balance = acct.get('total_balance', 0) or acct.get('savings_balance', 0) or 0
            if provider and balance > 0:
                provider_totals[provider] = provider_totals.get(provider, 0) + balance
        
        if provider_totals:
            charts.append(ChartConfig(
                type=ChartType.BAR,
                title='צבירה לפי יצרן' if is_hebrew else 'Savings by Provider',
                data={
                    'labels': list(provider_totals.keys()),
                    'values': list(provider_totals.values())
                },
                options={
                    'horizontal': False,
                    'colors': ['#2196F3', '#4CAF50', '#FF9800', '#9C27B0', '#00BCD4'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))
        
        # 2. Savings vs Severance Breakdown (Doughnut Chart)
        total_savings = totals.get('total_savings_balance', 0)
        total_severance = totals.get('total_severance_balance', 0)
        
        if not total_savings and not total_severance:
            # Calculate from accounts
            total_savings = sum(a.get('savings_balance', 0) or 0 for a in accounts)
            total_severance = sum(a.get('severance_balance', 0) or 0 for a in accounts)
        
        if total_savings > 0 or total_severance > 0:
            labels = ['תגמולים', 'פיצויים'] if is_hebrew else ['Savings', 'Severance']
            charts.append(ChartConfig(
                type=ChartType.DOUGHNUT,
                title='תגמולים מול פיצויים' if is_hebrew else 'Savings vs Severance',
                data={
                    'labels': labels,
                    'values': [total_savings, total_severance]
                },
                options={
                    'colors': ['#4CAF50', '#FF9800'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))
        
        # 3. Insurance Coverage Breakdown (Pie Chart)
        coverage_totals = {}
        for acct in accounts:
            death_coverage = acct.get('death_coverage', 0) or 0
            disability_coverage = acct.get('disability_coverage', 0) or 0
            
            if death_coverage > 0:
                label = 'ביטוח חיים' if is_hebrew else 'Life Insurance'
                coverage_totals[label] = coverage_totals.get(label, 0) + death_coverage
            if disability_coverage > 0:
                label = 'אובדן כושר' if is_hebrew else 'Disability'
                coverage_totals[label] = coverage_totals.get(label, 0) + disability_coverage
        
        if coverage_totals:
            charts.append(ChartConfig(
                type=ChartType.PIE,
                title='כיסויים ביטוחיים' if is_hebrew else 'Insurance Coverage',
                data={
                    'labels': list(coverage_totals.keys()),
                    'values': list(coverage_totals.values())
                },
                options={
                    'colors': ['#E91E63', '#3F51B5'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))
        
        # 4. Product Type Distribution (Pie Chart)
        product_balances = {}
        for acct in accounts:
            product_type = acct.get('product_type_name', '') or acct.get('product_type', '')
            if not product_type:
                product_type = 'לא מוגדר' if is_hebrew else 'Undefined'
            balance = acct.get('total_balance', 0) or acct.get('savings_balance', 0) or 0
            if balance > 0:
                product_balances[product_type] = product_balances.get(product_type, 0) + balance
        
        if product_balances and len(product_balances) > 1:
            charts.append(ChartConfig(
                type=ChartType.PIE,
                title='צבירה לפי סוג מוצר' if is_hebrew else 'Savings by Product Type',
                data={
                    'labels': list(product_balances.keys()),
                    'values': list(product_balances.values())
                },
                options={
                    'colors': ['#009688', '#795548', '#607D8B', '#FF5722', '#673AB7'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))
        
        # 5. Total Summary Gauge (if we have total balance)
        total_balance = totals.get('total_balance', 0)
        if not total_balance:
            total_balance = sum(a.get('total_balance', 0) or 0 for a in accounts)
        
        if total_balance > 0:
            charts.append(ChartConfig(
                type=ChartType.GAUGE,
                title='סה״כ חיסכון' if is_hebrew else 'Total Savings',
                data={
                    'value': total_balance,
                    'display_value': f'₪{total_balance:,.0f}',
                    'min': 0,
                    'max': total_balance * 1.2,  # 20% headroom
                    'thresholds': [total_balance * 0.25, total_balance * 0.5, total_balance * 0.75]
                },
                options={
                    'colors': ['#ffeb3b', '#8BC34A', '#4CAF50', '#2196F3'],
                    'currency': True,
                    'currency_symbol': '₪'
                }
            ))
        
        return charts
    
    def _generate_recommendations(self, analysis: AnalysisResult, lang: str) -> List[Recommendation]:
        """Generate actionable recommendations"""
        recommendations = []
        rec_id = 1
        
        # High risk score recommendation
        if analysis.risk_score > 70:
            recommendations.append(Recommendation(
                id=f"REC-{rec_id}",
                category='risk',
                priority=Priority.URGENT,
                title='סקירת סיכונים דחופה' if lang == 'hebrew' else 'Urgent Risk Review Required',
                description='ציון הסיכון הכולל גבוה ומצריך התייחסות מיידית' if lang == 'hebrew' 
                           else 'The overall risk score is high and requires immediate attention',
                action_items=[
                    'סקור את כל החריגות שזוהו' if lang == 'hebrew' else 'Review all identified anomalies',
                    'בדוק את הנתונים החריגים' if lang == 'hebrew' else 'Verify outlier data points',
                    'עדכן את הערכת הסיכון' if lang == 'hebrew' else 'Update risk assessment'
                ],
                expected_impact='הפחתת רמת הסיכון ב-20-30%' if lang == 'hebrew' else 'Risk level reduction of 20-30%'
            ))
            rec_id += 1
        
        # Missing data recommendation
        missing_patterns = [p for p in analysis.patterns_found if p.type == 'missing_data']
        if missing_patterns:
            recommendations.append(Recommendation(
                id=f"REC-{rec_id}",
                category='data_quality',
                priority=Priority.HIGH,
                title='השלמת נתונים חסרים' if lang == 'hebrew' else 'Complete Missing Data',
                description='זוהו שדות עם נתונים חסרים המשפיעים על איכות הניתוח' if lang == 'hebrew'
                           else 'Fields with missing data detected affecting analysis quality',
                action_items=[
                    'אסוף את הנתונים החסרים' if lang == 'hebrew' else 'Collect missing data',
                    'עדכן את המערכת' if lang == 'hebrew' else 'Update the system',
                    'הרץ ניתוח מחדש' if lang == 'hebrew' else 'Re-run analysis'
                ],
                expected_impact='שיפור דיוק הניתוח ב-15-25%' if lang == 'hebrew' else 'Analysis accuracy improvement of 15-25%'
            ))
            rec_id += 1
        
        # Anomalies recommendation
        if analysis.anomalies:
            high_severity = [a for a in analysis.anomalies if a.severity in [Severity.HIGH, Severity.CRITICAL]]
            if high_severity:
                recommendations.append(Recommendation(
                    id=f"REC-{rec_id}",
                    category='anomalies',
                    priority=Priority.HIGH,
                    title='טיפול בחריגות קריטיות' if lang == 'hebrew' else 'Address Critical Anomalies',
                    description=f'זוהו {len(high_severity)} חריגות ברמה גבוהה או קריטית' if lang == 'hebrew'
                               else f'{len(high_severity)} high or critical anomalies detected',
                    action_items=[a.recommendation for a in high_severity[:3]],
                    expected_impact='הפחתת סיכון והגברת אמינות הנתונים' if lang == 'hebrew' 
                                   else 'Risk reduction and improved data reliability'
                ))
                rec_id += 1
        
        # Data type specific recommendations
        if analysis.data_classification == DataType.INSURANCE:
            recommendations.append(Recommendation(
                id=f"REC-{rec_id}",
                category='insurance',
                priority=Priority.MEDIUM,
                title='סקירת כיסויים ביטוחיים' if lang == 'hebrew' else 'Review Insurance Coverage',
                description='מומלץ לבדוק התאמת הכיסויים לצרכים' if lang == 'hebrew'
                           else 'Review coverage adequacy against needs',
                action_items=[
                    'השווה כיסויים לסיכונים' if lang == 'hebrew' else 'Compare coverage to risks',
                    'בדוק חפיפות בפוליסות' if lang == 'hebrew' else 'Check for policy overlaps',
                    'עדכן סכומי ביטוח' if lang == 'hebrew' else 'Update coverage amounts'
                ],
                expected_impact='אופטימיזציה של הוצאות ביטוח' if lang == 'hebrew' else 'Insurance expense optimization'
            ))
            rec_id += 1
        elif analysis.data_classification == DataType.INVESTMENT:
            recommendations.append(Recommendation(
                id=f"REC-{rec_id}",
                category='investment',
                priority=Priority.MEDIUM,
                title='איזון תיק השקעות' if lang == 'hebrew' else 'Portfolio Rebalancing',
                description='בדוק את פיזור התיק ואיזון הסיכון' if lang == 'hebrew'
                           else 'Review portfolio diversification and risk balance',
                action_items=[
                    'נתח פיזור נכסים' if lang == 'hebrew' else 'Analyze asset allocation',
                    'בדוק התאמה לפרופיל סיכון' if lang == 'hebrew' else 'Check risk profile alignment',
                    'שקול איזון מחדש' if lang == 'hebrew' else 'Consider rebalancing'
                ],
                expected_impact='שיפור יחס תשואה/סיכון' if lang == 'hebrew' else 'Improved return/risk ratio'
            ))
            rec_id += 1
        
        return recommendations
    
    def get_documents_for_user(self, user_id: str, user_role: str) -> List[Dict]:
        """
        Get all documents accessible to a user.
        Admins can see all documents, customers only see their own.
        
        Args:
            user_id: The user's ID
            user_role: The user's role (admin, customer, etc.)
            
        Returns:
            List of document metadata (without parsed_data for efficiency)
        """
        results = []
        is_admin = user_role in ['admin', 'actuary', 'underwriter', 'analyst']
        
        for doc_id, doc in self.documents.items():
            # Admin roles can see all documents
            if is_admin or doc.get('owner_id') == user_id:
                # Return summary without heavy parsed_data
                results.append({
                    'document_id': doc['document_id'],
                    'filename': doc['filename'],
                    'file_type': doc['file_type'],
                    'file_size': doc['file_size'],
                    'status': doc['status'],
                    'row_count': doc.get('row_count', 0),
                    'column_count': doc.get('column_count', 0),
                    'owner_id': doc.get('owner_id'),
                    'created_at': doc.get('created_at')
                })
        
        # Sort by created_at descending
        results.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        return results
    
    def get_reports_for_user(self, user_id: str, user_role: str) -> List[Dict]:
        """
        Get all reports accessible to a user.
        Admins can see all reports, customers only see their own.
        
        Args:
            user_id: The user's ID
            user_role: The user's role
            
        Returns:
            List of report summaries
        """
        results = []
        is_admin = user_role in ['admin', 'actuary', 'underwriter', 'analyst']
        
        for report_id, report in self.reports.items():
            # Get the associated analysis to check ownership
            analysis = self.analyses.get(report.analysis_id)
            if not analysis:
                continue
            
            doc = self.documents.get(analysis.document_id)
            if not doc:
                continue
            
            # Check access permission
            if is_admin or doc.get('owner_id') == user_id:
                results.append({
                    'report_id': report.id,
                    'title': report.title,
                    'report_type': report.report_type,
                    'language': report.language,
                    'generated_at': report.generated_at,
                    'analysis_id': report.analysis_id,
                    'document_id': analysis.document_id,
                    'filename': doc.get('filename'),
                    'risk_score': report.metadata.get('risk_score'),
                    'confidence': report.metadata.get('confidence'),
                    'owner_id': doc.get('owner_id')
                })
        
        # Sort by generated_at descending
        results.sort(key=lambda x: x.get('generated_at', ''), reverse=True)
        return results

    def revoke_reports_for_date(self, user_id: str, user_role: str, target_date: date,
                                scope: str = "self") -> Dict[str, Any]:
        """
        Revoke (delete) reports generated on a specific date.
        Also removes orphaned analyses/documents when possible.

        Args:
            user_id: The requesting user ID
            user_role: The requesting user role
            target_date: Date to revoke (date object)
            scope: "self" (default) or "all" (admin only)

        Returns:
            Summary dict with counts of removed items.
        """
        is_admin = user_role in ['admin', 'actuary', 'underwriter', 'analyst']
        allow_all = is_admin and scope == "all"

        reports_to_remove = []
        analyses_to_check = set()
        documents_to_check = set()

        for report_id, report in list(self.reports.items()):
            report_dt = self._parse_report_datetime(report.generated_at)
            if not report_dt or report_dt.date() != target_date:
                continue

            analysis = self.analyses.get(report.analysis_id)
            if not analysis:
                continue

            doc = self.documents.get(analysis.document_id)
            if not doc:
                continue

            if not allow_all and doc.get('owner_id') != user_id:
                continue

            reports_to_remove.append(report_id)
            analyses_to_check.add(report.analysis_id)
            documents_to_check.add(analysis.document_id)

        # Remove reports
        for report_id in reports_to_remove:
            self.reports.pop(report_id, None)

        # Remove analyses not referenced by any remaining report
        analyses_removed = 0
        for analysis_id in analyses_to_check:
            if not any(r.analysis_id == analysis_id for r in self.reports.values()):
                self.analyses.pop(analysis_id, None)
                analyses_removed += 1

        # Remove documents not referenced by any remaining analysis
        documents_removed = 0
        for document_id in documents_to_check:
            if not any(a.document_id == document_id for a in self.analyses.values()):
                self.documents.pop(document_id, None)
                documents_removed += 1

        # Persist changes
        self.save_data()

        _risk_report_audit('risk_report_revoked', None, {
            'target_date': str(target_date),
            'reports_removed': len(reports_to_remove),
            'analyses_removed': analyses_removed,
            'documents_removed': documents_removed,
            'scope': scope,
            'user_id': user_id,
            'user_role': user_role,
        })

        return {
            'success': True,
            'target_date': target_date.isoformat(),
            'reports_removed': len(reports_to_remove),
            'analyses_removed': analyses_removed,
            'documents_removed': documents_removed
        }

    def _parse_report_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse report datetime from stored string values."""
        if not date_str:
            return None
        try:
            if date_str.endswith('Z'):
                return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return datetime.fromisoformat(date_str)
        except ValueError:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return None
    
    def authorize_access(self, resource_type: str, resource_id: str, 
                        user_id: str, user_role: str) -> Tuple[bool, Optional[str]]:
        """
        Check if a user is authorized to access a specific resource.
        
        Args:
            resource_type: 'document', 'analysis', or 'report'
            resource_id: The ID of the resource
            user_id: The user's ID
            user_role: The user's role
            
        Returns:
            Tuple of (is_authorized, error_message)
        """
        is_admin = user_role in ['admin', 'actuary', 'underwriter', 'analyst']
        
        if is_admin:
            return True, None
        
        if resource_type == 'document':
            doc = self.documents.get(resource_id)
            if not doc:
                return False, f"Document {resource_id} not found"
            if doc.get('owner_id') != user_id:
                return False, "Access denied: You can only access your own documents"
            return True, None
        
        elif resource_type == 'analysis':
            analysis = self.analyses.get(resource_id)
            if not analysis:
                return False, f"Analysis {resource_id} not found"
            doc = self.documents.get(analysis.document_id)
            if not doc:
                return False, "Associated document not found"
            if doc.get('owner_id') != user_id:
                return False, "Access denied: You can only access your own analyses"
            return True, None
        
        elif resource_type == 'report':
            report = self.reports.get(resource_id)
            if not report:
                return False, f"Report {resource_id} not found"
            analysis = self.analyses.get(report.analysis_id)
            if not analysis:
                return False, "Associated analysis not found"
            doc = self.documents.get(analysis.document_id)
            if not doc:
                return False, "Associated document not found"
            if doc.get('owner_id') != user_id:
                return False, "Access denied: You can only access your own reports"
            return True, None
        
        return False, f"Unknown resource type: {resource_type}"
    
    def get_report_by_id(self, report_id: str, user_id: str = None, user_role: str = None) -> Optional[GeneratedReport]:
        """
        Get a specific report by ID with access control.
        
        Args:
            report_id: The report ID
            user_id: The requesting user's ID (for access control)
            user_role: The requesting user's role
            
        Returns:
            The report if found and authorized, None otherwise
        """
        if user_id and user_role:
            is_authorized, error = self.authorize_access('report', report_id, user_id, user_role)
            if not is_authorized:
                return None
        
        return self.reports.get(report_id)

    def build_report_download_summary(self, report_id: str, user_id: str, user_role: str) -> Dict[str, Any]:
        """
        Build a sanitized report summary payload for downloadable exports.
        This payload excludes source URLs/credentials and focuses on report metrics.
        """
        report = self.get_report_by_id(report_id, user_id=user_id, user_role=user_role)
        if not report:
            raise ValueError('Report not found or access denied')

        analysis = self.analyses.get(report.analysis_id)
        if not analysis:
            raise ValueError('Associated analysis not found')

        doc = self.documents.get(analysis.document_id, {}) if analysis.document_id else {}
        doc_data = doc.get('parsed_data', {}) if isinstance(doc, dict) else {}
        pension_data = doc_data.get('pension_data') if isinstance(doc_data, dict) else None
        summary = self._extract_savings_cover_id_summary(doc_data, pension_data)

        table_sections: List[Dict[str, Any]] = []
        for section in report.sections:
            if not section.data_table:
                continue

            section_title = section.title or ''
            title_lower = section_title.lower()
            if 'swiftness' in title_lower or 'resource' in title_lower:
                continue

            data_table = section.data_table if isinstance(section.data_table, dict) else {}
            columns = data_table.get('columns', [])
            rows = data_table.get('rows', [])

            # Support key/value maps (e.g. key metrics) in addition to tabular structures.
            if not isinstance(rows, list):
                rows = []
            if not rows and data_table:
                rows = [{'Metric': key, 'Value': value} for key, value in data_table.items() if not isinstance(value, dict)]
                columns = ['Metric', 'Value']

            cleaned_rows: List[Dict[str, Any]] = []
            for row in rows[:80]:
                if not isinstance(row, dict):
                    continue
                cleaned_row = {}
                for key, value in row.items():
                    if isinstance(value, str):
                        cleaned_value = re.sub(r'https?://\S+', '[redacted]', value)
                    else:
                        cleaned_value = value
                    cleaned_row[str(key)] = cleaned_value
                cleaned_rows.append(cleaned_row)

            if not isinstance(columns, list) or not columns:
                columns = list(cleaned_rows[0].keys()) if cleaned_rows else []

            table_sections.append({
                'title': section_title,
                'columns': [str(c) for c in columns],
                'rows': cleaned_rows
            })

        chart_summaries: List[Dict[str, Any]] = []
        for chart in report.charts:
            chart_data = chart.data if isinstance(chart.data, dict) else {}
            labels = chart_data.get('labels', [])
            values = chart_data.get('values', [])
            if isinstance(labels, list) and isinstance(values, list):
                series = [
                    {
                        'label': str(label),
                        'value': values[index] if index < len(values) else None
                    }
                    for index, label in enumerate(labels[:30])
                ]
            else:
                series = [{'label': 'value', 'value': chart_data.get('value')}]

            chart_type = chart.type.value if isinstance(chart.type, Enum) else str(chart.type)
            chart_summaries.append({
                'title': chart.title,
                'type': chart_type,
                'series': series
            })

        recommendations = [{
            'priority': rec.priority.value if isinstance(rec.priority, Enum) else str(rec.priority),
            'title': rec.title,
            'description': rec.description,
            'action_items': rec.action_items,
            'expected_impact': rec.expected_impact,
        } for rec in report.recommendations]

        return {
            'report_id': report.id,
            'title': report.title,
            'language': report.language,
            'generated_at': report.generated_at,
            'report_type': report.report_type,
            'risk_score': report.metadata.get('risk_score'),
            'confidence': report.metadata.get('confidence'),
            'savings_cover_id_summary': summary,
            'table_sections': table_sections,
            'chart_summaries': chart_summaries,
            'recommendations': recommendations,
        }
    
    def to_dict(self, obj) -> Dict:
        """Convert dataclass objects to dictionaries for JSON serialization"""
        if hasattr(obj, '__dataclass_fields__'):
            result = {}
            for field_name in obj.__dataclass_fields__:
                value = getattr(obj, field_name)
                result[field_name] = self.to_dict(value)
            return result
        elif isinstance(obj, list):
            return [self.to_dict(item) for item in obj]
        elif isinstance(obj, Enum):
            return obj.value
        elif isinstance(obj, dict):
            return {k: self.to_dict(v) for k, v in obj.items()}
        else:
            return obj
    
    # =========================================================================
    # PERSISTENCE - Save and Load Data
    # =========================================================================
    
    def save_data(self, filepath: str = None) -> bool:
        """
        Save all documents, analyses, and reports to a JSON file.
        
        Args:
            filepath: Optional custom filepath. Defaults to AI_REPORTS_DATA_FILE.
            
        Returns:
            True if successful, False otherwise
        """
        if filepath is None:
            filepath = AI_REPORTS_DATA_FILE
        
        try:
            # Convert analyses to serializable format
            analyses_data = {}
            for aid, analysis in self.analyses.items():
                analyses_data[aid] = self.to_dict(analysis)
            
            # Convert reports to serializable format
            reports_data = {}
            for rid, report in self.reports.items():
                reports_data[rid] = self.to_dict(report)
            
            # Prepare documents (remove parsed_data to save space - can be re-parsed)
            documents_data = {}
            for did, doc in self.documents.items():
                doc_copy = doc.copy()
                # Keep only metadata, not the full parsed data
                if 'parsed_data' in doc_copy:
                    doc_copy['has_parsed_data'] = doc_copy['parsed_data'] is not None
                    # Store column info but not full row data to save space
                    if doc_copy['parsed_data']:
                        doc_copy['columns'] = doc_copy['parsed_data'].get('columns', [])
                    del doc_copy['parsed_data']
                documents_data[did] = doc_copy
            
            data = {
                'saved_at': datetime.now().isoformat(),
                'version': '1.0',
                'documents': documents_data,
                'analyses': analyses_data,
                'reports': reports_data,
                'stats': {
                    'total_documents': len(self.documents),
                    'total_analyses': len(self.analyses),
                    'total_reports': len(self.reports)
                }
            }
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else '.', exist_ok=True)
            
            # Write to temp file first for atomic operation
            temp_file = filepath + '.tmp'
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, default=str, indent=2, ensure_ascii=False)
            
            # Atomic rename
            os.rename(temp_file, filepath)
            print(f"[AI_REPORTS] Saved data to {filepath} ({len(self.documents)} docs, {len(self.analyses)} analyses, {len(self.reports)} reports)")
            return True
            
        except Exception as e:
            print(f"[AI_REPORTS] Error saving data: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def load_data(self, filepath: str = None) -> bool:
        """
        Load documents, analyses, and reports from a JSON file.
        
        Args:
            filepath: Optional custom filepath. Defaults to AI_REPORTS_DATA_FILE.
            
        Returns:
            True if successful, False otherwise
        """
        if filepath is None:
            filepath = AI_REPORTS_DATA_FILE
        
        if not os.path.exists(filepath):
            print(f"[AI_REPORTS] No saved data file found at {filepath}")
            return False
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Load documents
            documents_data = data.get('documents', {})
            for did, doc in documents_data.items():
                # Restore minimal document structure
                doc['parsed_data'] = None  # Will need to re-upload to re-parse
                self.documents[did] = doc
            
            # Load analyses - reconstruct AnalysisResult objects
            analyses_data = data.get('analyses', {})
            for aid, analysis_dict in analyses_data.items():
                try:
                    # Reconstruct factors
                    factors = [Factor(**f) for f in analysis_dict.get('extracted_factors', [])]
                    
                    # Reconstruct patterns
                    patterns = [Pattern(**p) for p in analysis_dict.get('patterns_found', [])]
                    
                    # Reconstruct anomalies
                    anomalies = []
                    for a in analysis_dict.get('anomalies', []):
                        a_copy = a.copy()
                        a_copy['severity'] = Severity(a_copy['severity'])
                        anomalies.append(Anomaly(**a_copy))
                    
                    # Reconstruct analysis
                    analysis = AnalysisResult(
                        id=analysis_dict['id'],
                        document_id=analysis_dict['document_id'],
                        language=analysis_dict['language'],
                        language_name=analysis_dict['language_name'],
                        data_classification=DataType(analysis_dict['data_classification']),
                        extracted_factors=factors,
                        patterns_found=patterns,
                        anomalies=anomalies,
                        risk_score=analysis_dict['risk_score'],
                        confidence=analysis_dict['confidence'],
                        processing_time_ms=analysis_dict['processing_time_ms'],
                        summary=analysis_dict['summary'],
                        key_metrics=analysis_dict['key_metrics']
                    )
                    self.analyses[aid] = analysis
                except Exception as e:
                    print(f"[AI_REPORTS] Error loading analysis {aid}: {e}")
            
            # Load reports - reconstruct GeneratedReport objects
            reports_data = data.get('reports', {})
            for rid, report_dict in reports_data.items():
                try:
                    # Reconstruct sections
                    sections = [ReportSection(**s) for s in report_dict.get('sections', [])]
                    
                    # Reconstruct charts
                    charts = []
                    for c in report_dict.get('charts', []):
                        c_copy = c.copy()
                        c_copy['type'] = ChartType(c_copy['type'])
                        charts.append(ChartConfig(**c_copy))
                    
                    # Reconstruct recommendations
                    recommendations = []
                    for r in report_dict.get('recommendations', []):
                        r_copy = r.copy()
                        r_copy['priority'] = Priority(r_copy['priority'])
                        recommendations.append(Recommendation(**r_copy))
                    
                    # Reconstruct report
                    report = GeneratedReport(
                        id=report_dict['id'],
                        analysis_id=report_dict['analysis_id'],
                        report_type=report_dict['report_type'],
                        language=report_dict['language'],
                        title=report_dict['title'],
                        sections=sections,
                        charts=charts,
                        recommendations=recommendations,
                        generated_at=report_dict['generated_at'],
                        metadata=report_dict.get('metadata', {})
                    )
                    self.reports[rid] = report
                except Exception as e:
                    print(f"[AI_REPORTS] Error loading report {rid}: {e}")
            
            print(f"[AI_REPORTS] Loaded data from {filepath} ({len(self.documents)} docs, {len(self.analyses)} analyses, {len(self.reports)} reports)")
            return True
            
        except Exception as e:
            print(f"[AI_REPORTS] Error loading data: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_persistence_stats(self) -> Dict[str, Any]:
        """Get statistics about stored data"""
        return {
            'total_documents': len(self.documents),
            'total_analyses': len(self.analyses),
            'total_reports': len(self.reports),
            'documents_by_owner': self._count_by_owner(self.documents),
            'reports_by_type': self._count_by_type()
        }
    
    def _count_by_owner(self, collection: Dict) -> Dict[str, int]:
        """Count items by owner"""
        counts = {}
        for item in collection.values():
            owner = item.get('owner_id', 'unknown')
            counts[owner] = counts.get(owner, 0) + 1
        return counts
    
    def _count_by_type(self) -> Dict[str, int]:
        """Count reports by type"""
        counts = {}
        for report in self.reports.values():
            rtype = report.report_type
            counts[rtype] = counts.get(rtype, 0) + 1
        return counts


# Persistence file path
AI_REPORTS_DATA_FILE = os.environ.get('AI_REPORTS_DATA_FILE', 'data/ai_reports_data.json')

# Singleton instance
_ai_reports_service: AIRiskReportsService = None


def get_ai_reports_service() -> AIRiskReportsService:
    """Get or create the AI reports service singleton, loading saved data if available"""
    global _ai_reports_service
    if _ai_reports_service is None:
        _ai_reports_service = AIRiskReportsService()
        # Load persisted data on first access
        _ai_reports_service.load_data()
    return _ai_reports_service


def init_ai_reports_service(load_persisted: bool = True) -> AIRiskReportsService:
    """
    Initialize the AI reports service.
    
    Args:
        load_persisted: If True, load previously saved data from disk.
                       Set to False for testing with fresh state.
    """
    global _ai_reports_service
    _ai_reports_service = AIRiskReportsService()
    if load_persisted:
        _ai_reports_service.load_data()
    return _ai_reports_service
